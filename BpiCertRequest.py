import os
import json
from io import BytesIO
from datetime import datetime
from copy import copy

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# =========================
# PAGE CONFIG
# =========================
st.set_page_config(
    page_title="BPI Cert Request Drive",
    page_icon="📄",
    layout="wide",
)

# =========================
# CUSTOM CSS - DARK MODE ONLY
# =========================
CSS = """
<style>
    .stApp {
        background: linear-gradient(180deg, #0b0f19 0%, #111827 100%);
        color: #f3f4f6;
    }

    [data-testid="stAppViewContainer"] {
        background: linear-gradient(180deg, #0b0f19 0%, #111827 100%);
    }

    [data-testid="stHeader"] {
        background: rgba(0, 0, 0, 0);
    }

    [data-testid="stSidebar"] {
        background: #0f172a;
        border-right: 1px solid #1f2937;
    }

    .main-title {
        font-size: 2.2rem;
        font-weight: 800;
        color: #f8fafc;
        margin-bottom: 0.2rem;
        letter-spacing: -0.02em;
    }

    .sub-title {
        font-size: 1rem;
        color: #94a3b8;
        margin-bottom: 1.4rem;
    }

    .section-card {
        background: rgba(17, 24, 39, 0.92);
        border: 1px solid #1f2937;
        border-radius: 20px;
        padding: 1.2rem;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.35);
        margin-bottom: 1rem;
    }

    .hero-card {
        background: linear-gradient(135deg, rgba(91, 61, 245, 0.16) 0%, rgba(30, 41, 59, 0.95) 100%);
        border: 1px solid rgba(122, 90, 248, 0.25);
        border-radius: 22px;
        padding: 1.4rem 1.4rem 1.1rem 1.4rem;
        box-shadow: 0 12px 28px rgba(0, 0, 0, 0.30);
        margin-bottom: 1rem;
    }

    .metric-card {
        background: linear-gradient(180deg, #111827 0%, #0f172a 100%);
        border: 1px solid #243041;
        border-radius: 18px;
        padding: 1rem;
        box-shadow: 0 8px 20px rgba(0, 0, 0, 0.28);
        text-align: center;
    }

    .metric-value {
        font-size: 1.7rem;
        font-weight: 800;
        color: #a78bfa;
        margin-bottom: 0.15rem;
    }

    .metric-label {
        font-size: 0.95rem;
        color: #94a3b8;
    }

    .result-card {
        background: linear-gradient(180deg, #111827 0%, #0f172a 100%);
        border: 1px solid #243041;
        border-radius: 18px;
        padding: 1rem;
        box-shadow: 0 8px 20px rgba(0, 0, 0, 0.24);
        margin-bottom: 0.9rem;
    }

    .result-title {
        font-size: 1rem;
        font-weight: 700;
        color: #f8fafc;
        margin-bottom: 0.25rem;
    }

    .result-meta {
        color: #94a3b8;
        font-size: 0.92rem;
        margin-bottom: 0.75rem;
    }

    .sidebar-note {
        background: #111827;
        border: 1px solid #243041;
        border-radius: 14px;
        padding: 0.9rem;
        color: #cbd5e1;
        font-size: 0.92rem;
    }

    .small-chip {
        display: inline-block;
        background: rgba(167, 139, 250, 0.14);
        color: #c4b5fd;
        border: 1px solid rgba(167, 139, 250, 0.20);
        border-radius: 999px;
        padding: 0.28rem 0.65rem;
        font-size: 0.8rem;
        font-weight: 700;
        margin-right: 0.35rem;
        margin-top: 0.2rem;
    }

    .stButton > button,
    .stDownloadButton > button {
        border-radius: 12px !important;
        font-weight: 700 !important;
        padding: 0.68rem 1rem !important;
    }

    .stButton > button {
        background: linear-gradient(135deg, #6d4aff 0%, #8b5cf6 100%) !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 10px 24px rgba(109, 74, 255, 0.25);
    }

    .stDownloadButton > button {
        background: #111827 !important;
        color: #e5e7eb !important;
        border: 1px solid #2b3750 !important;
        box-shadow: none !important;
    }

    .stTextInput input,
    .stNumberInput input,
    .stFileUploader,
    .stSelectbox div[data-baseweb="select"] > div {
        background-color: #0f172a !important;
        color: #f3f4f6 !important;
        border-radius: 12px !important;
    }

    div[data-testid="stFileUploaderDropzone"] {
        background: #0f172a !important;
        border: 1px dashed #334155 !important;
        border-radius: 16px !important;
    }

    .stAlert {
        border-radius: 14px;
    }

    h1, h2, h3, h4, h5, h6, p, label, div {
        color: inherit;
    }
</style>
"""

# =========================
# CONSTANTS
# =========================
SHEET_NAME = "BPI RECOVERIES REGULAR AND RECO"
COUNTER_FILE = "ctrl_counters.json"
START_CTRL_NUMBER = 280  # first generated CTRL will be 000281

COUNTER_KEYS = [
    "RECO_1",
    "RECO_2",
    "RECO_3",
]

RULES = [
    # ALL PLACEMENT RECOV AND SPECIAL PROJECTS
    {
        "type_of_cert": "Request for Closure & Cert (OTP)",
        "payment_type": "OTP",
        "levels": None,
        "file_name": "REQUEST FOR CLOSURE & CERT (OTP) (WOFF - NON LEGAL) - MADRID - {date}",
        "label": "Download OTP File",
        "use_ctrl_number": False,
        "counter_key": "OTP_FILE",
    },
    {
        "type_of_cert": "Request for Closure & Cert (TERM)",
        "payment_type": "TERM",
        "levels": None,
        "file_name": "REQUEST FOR CLOSURE & CERT (TERM) (WOFF - NON LEGAL) - MADRID - {date}",
        "label": "Download TERM File",
        "use_ctrl_number": False,
        "counter_key": "TERM_FILE",
    },
    # RECOM (OTP)
    {
        "type_of_cert": "Late PA Encoding - For Recom, Late Payment - For Recom, Late Request For Closure, Manual Debit by the Bank",
        "payment_type": "OTP",
        "levels": ["BPI RECOV 1", "BPI SPECIAL PROJECT - REVIVAL"],
        "file_name": "RUSH RECOM FOR CLOSURE & FOR CERT (WOFF - NON LEGAL) - MADRID - RECO 1 (OTP) - {date} - CTRL#{ctrl_no}.xlsx",
        "label": "Late PA Encoding - RECO 1 OTP",
        "use_ctrl_number": True,
        "counter_key": "RECO_1",
    },
    {
        "type_of_cert": "Late PA Encoding - For Recom, Late Payment - For Recom, Late Request For Closure, Manual Debit by the Bank",
        "payment_type": "OTP",
        "levels": ["BPI RECOV 2"],
        "file_name": "RUSH RECOM FOR CLOSURE & FOR CERT (WOFF - NON LEGAL) - MADRID - RECO 2 (OTP) - {date} - CTRL#{ctrl_no}.xlsx",
        "label": "Late PA Encoding - RECO 2 OTP",
        "use_ctrl_number": True,
        "counter_key": "RECO_2",
    },
    {
        "type_of_cert": "Late PA Encoding - For Recom, Late Payment - For Recom, Late Request For Closure, Manual Debit by the Bank",
        "payment_type": "OTP",
        "levels": ["BPI RECOV 3"],
        "file_name": "RUSH RECOM FOR CLOSURE & FOR CERT (WOFF - NON LEGAL) - MADRID - RECO 3 (OTP) - {date} - CTRL#{ctrl_no}.xlsx",
        "label": "Late PA Encoding - RECO 3 OTP",
        "use_ctrl_number": True,
        "counter_key": "RECO_3",
    },
    # RECOM (TERM)
    {
        "type_of_cert": "Late PA Encoding - For Recom, Late Payment - For Recom, Late Request For Closure, Manual Debit by the Bank",
        "payment_type": "TERM",
        "levels": ["BPI RECOV 1", "BPI SPECIAL PROJECT - REVIVAL"],
        "file_name": "RUSH RECOM FOR CLOSURE & FOR CERT (WOFF - NON LEGAL) - MADRID - RECO 1 (TERM) - {date} - CTRL#{ctrl_no}.xlsx",
        "label": "Late PA Encoding - RECO 1 TERM",
        "use_ctrl_number": True,
        "counter_key": "RECO_1",
    },
    {
        "type_of_cert": "Late PA Encoding - For Recom, Late Payment - For Recom, Late Request For Closure, Manual Debit by the Bank",
        "payment_type": "TERM",
        "levels": ["BPI RECOV 2"],
        "file_name": "RUSH RECOM FOR CLOSURE & FOR CERT (WOFF - NON LEGAL) - MADRID - RECO 2 (TERM) - {date} - CTRL#{ctrl_no}.xlsx",
        "label": "Late PA Encoding - RECO 2 TERM",
        "use_ctrl_number": True,
        "counter_key": "RECO_2",
    },
    {
        "type_of_cert": "Late PA Encoding - For Recom, Late Payment - For Recom, Late Request For Closure, Manual Debit by the Bank",
        "payment_type": "TERM",
        "levels": ["BPI RECOV 3"],
        "file_name": "RUSH RECOM FOR CLOSURE & FOR CERT (WOFF - NON LEGAL) - MADRID - RECO 3 (TERM) - {date} - CTRL#{ctrl_no}.xlsx",
        "label": "Late PA Encoding - RECO 3 TERM",
        "use_ctrl_number": True,
        "counter_key": "RECO_3",
    },
]

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

HEADER_FILL = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")

FIXED_SOURCE_WIDTHS = {
    "V": 20,
    "W": 25,
}
NO_COLUMN_WIDTH = 8

# =========================
# HELPERS
# =========================
def normalize(value):
    return "" if value is None else str(value).strip()


def get_header_map(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = normalize(ws.cell(row=1, column=col).value)
        if header:
            headers[header] = col
    return headers


def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)
    target_cell.border = THIN_BORDER


def is_row_allowed(ws, row, da_remarks_col, leaders_approval_col):
    da_remarks = normalize(ws.cell(row=row, column=da_remarks_col).value)
    leaders_approval = normalize(ws.cell(row=row, column=leaders_approval_col).value)

    if da_remarks != "":
        return False

    if leaders_approval == "" or leaders_approval.casefold() in ["disapproved", "on hold"]:
        return False

    return True


def get_matching_rows(
    ws,
    cert_col,
    payment_type_col,
    level_col,
    da_remarks_col,
    leaders_approval_col,
    cert_status_2_col,
    type_of_cert,
    payment_type,
    allowed_levels=None,
):
    rows = []

    allowed_type_of_certs = {
        normalize(value).casefold()
        for value in str(type_of_cert).split(",")
        if normalize(value)
    }

    normalized_levels = None
    if allowed_levels:
        normalized_levels = {normalize(level).casefold() for level in allowed_levels}

    for row in range(2, ws.max_row + 1):
        if not is_row_allowed(ws, row, da_remarks_col, leaders_approval_col):
            continue

        cert_value = normalize(ws.cell(row=row, column=cert_col).value).casefold()
        payment_type_value = normalize(ws.cell(row=row, column=payment_type_col).value).casefold()
        level_value = normalize(ws.cell(row=row, column=level_col).value).casefold()

        if cert_value not in allowed_type_of_certs:
            continue

        if payment_type_value != payment_type.casefold():
            continue

        if normalized_levels is not None and level_value not in normalized_levels:
            continue

        rows.append(row)

    rows.sort(
        key=lambda row: (
            0 if normalize(ws.cell(row=row, column=cert_status_2_col).value).casefold() == "rush" else 1,
            row,
        )
    )

    return rows


def adjust_column_widths(source_ws, output_ws, selected_cols):
    output_ws.column_dimensions["A"].width = NO_COLUMN_WIDTH

    for output_col, source_col in enumerate(selected_cols, start=2):
        source_letter = get_column_letter(source_col)
        output_letter = get_column_letter(output_col)

        if source_letter in FIXED_SOURCE_WIDTHS:
            output_ws.column_dimensions[output_letter].width = FIXED_SOURCE_WIDTHS[source_letter]
            continue

        source_width = source_ws.column_dimensions[source_letter].width or 0

        max_length = 0
        for row in range(1, output_ws.max_row + 1):
            value = output_ws.cell(row=row, column=output_col).value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        output_ws.column_dimensions[output_letter].width = max(source_width, max_length + 2, 12)


def build_output_file(source_ws, selected_cols, matched_rows):
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = source_ws.title

    first_header_source = source_ws.cell(row=1, column=selected_cols[0])

    no_header = output_ws.cell(row=1, column=1, value="No.")
    no_header.border = THIN_BORDER
    no_header.fill = HEADER_FILL
    no_header.alignment = copy(first_header_source.alignment)
    no_header_font = copy(first_header_source.font)
    no_header_font.bold = True
    no_header.font = no_header_font

    for output_col, source_col in enumerate(selected_cols, start=2):
        source_cell = source_ws.cell(row=1, column=source_col)
        target_cell = output_ws.cell(row=1, column=output_col)

        copy_cell(source_cell, target_cell)
        target_cell.fill = HEADER_FILL
        header_font = copy(target_cell.font)
        header_font.bold = True
        target_cell.font = header_font

    output_row = 2
    for index, source_row in enumerate(matched_rows, start=1):
        no_cell = output_ws.cell(row=output_row, column=1, value=index)
        no_cell.border = THIN_BORDER
        no_cell.alignment = copy(source_ws.cell(row=source_row, column=selected_cols[0]).alignment)
        no_cell.font = copy(source_ws.cell(row=source_row, column=selected_cols[0]).font)

        for output_col, source_col in enumerate(selected_cols, start=2):
            copy_cell(
                source_ws.cell(row=source_row, column=source_col),
                output_ws.cell(row=output_row, column=output_col),
            )
        output_row += 1

    if source_ws.row_dimensions[1].height:
        output_ws.row_dimensions[1].height = source_ws.row_dimensions[1].height

    for output_row_index, source_row in enumerate(matched_rows, start=2):
        if source_ws.row_dimensions[source_row].height:
            output_ws.row_dimensions[output_row_index].height = source_ws.row_dimensions[source_row].height

    adjust_column_widths(source_ws, output_ws, selected_cols)

    buffer = BytesIO()
    output_wb.save(buffer)
    buffer.seek(0)
    return buffer


def load_counters():
    if not os.path.exists(COUNTER_FILE):
        return {key: START_CTRL_NUMBER for key in COUNTER_KEYS}

    try:
        with open(COUNTER_FILE, "r", encoding="utf-8") as f:
            counters = json.load(f)
    except Exception:
        counters = {}

    numeric_values = []
    for value in counters.values():
        try:
            numeric_values.append(int(value))
        except Exception:
            pass

    seed_value = max(numeric_values) if numeric_values else START_CTRL_NUMBER

    for key in COUNTER_KEYS:
        counters.setdefault(key, seed_value)

    return counters


def save_counters(counters):
    with open(COUNTER_FILE, "w", encoding="utf-8") as f:
        json.dump(counters, f)


def _get_file_bytes(uploaded_source):
    if isinstance(uploaded_source, (bytes, bytearray)):
        return bytes(uploaded_source)

    if hasattr(uploaded_source, "getvalue"):
        return uploaded_source.getvalue()

    if hasattr(uploaded_source, "read"):
        current_pos = None
        if hasattr(uploaded_source, "tell"):
            try:
                current_pos = uploaded_source.tell()
            except Exception:
                current_pos = None

        data = uploaded_source.read()

        if current_pos is not None and hasattr(uploaded_source, "seek"):
            try:
                uploaded_source.seek(current_pos)
            except Exception:
                pass

        return data

    raise TypeError("Unsupported uploaded file type.")


def ensure_excel_extension(file_name):
    if str(file_name).lower().endswith(".xlsx"):
        return file_name
    return f"{file_name}.xlsx"


def prepare_outputs(uploaded_source, progress_bar=None):
    if progress_bar is not None:
        progress_bar.progress(5, text="Reading workbook...")

    file_bytes = _get_file_bytes(uploaded_source)
    workbook = load_workbook(BytesIO(file_bytes))

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f'Sheet "{SHEET_NAME}" was not found. Available sheets: {", ".join(workbook.sheetnames)}'
        )

    ws = workbook[SHEET_NAME]
    headers = get_header_map(ws)

    required_headers = [
        "CUSTOMER NUMBER",
        "LEVEL",
        "TYPE OF CERT",
        "PAYMENT TYPE",
        "DA Remarks",
        "LEADERS APPROVAL",
        "CERT STATUS 2",
    ]
    missing = [header for header in required_headers if header not in headers]

    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")

    start_col = headers["CUSTOMER NUMBER"]
    end_col = headers["LEVEL"]
    cert_col = headers["TYPE OF CERT"]
    payment_type_col = headers["PAYMENT TYPE"]
    level_col = headers["LEVEL"]
    da_remarks_col = headers["DA Remarks"]
    leaders_approval_col = headers["LEADERS APPROVAL"]
    cert_status_2_col = headers["CERT STATUS 2"]

    if start_col > end_col:
        raise ValueError('"CUSTOMER NUMBER" is after "LEVEL". Please check the sheet.')

    selected_cols = list(range(start_col, end_col + 1))
    today = datetime.today().strftime("%d%B%Y").upper()
    outputs = []
    counters = load_counters()

    total_rules = len(RULES)

    for idx, rule in enumerate(RULES, start=1):
        if progress_bar is not None:
            percent = 10 + int(((idx - 1) / total_rules) * 75)
            progress_bar.progress(percent, text=f"Generating file logic {idx} of {total_rules}...")

        matched_rows = get_matching_rows(
            ws=ws,
            cert_col=cert_col,
            payment_type_col=payment_type_col,
            level_col=level_col,
            da_remarks_col=da_remarks_col,
            leaders_approval_col=leaders_approval_col,
            cert_status_2_col=cert_status_2_col,
            type_of_cert=rule["type_of_cert"],
            payment_type=rule["payment_type"],
            allowed_levels=rule.get("levels"),
        )

        if matched_rows:
            output_file = build_output_file(ws, selected_cols, matched_rows)

            if rule.get("use_ctrl_number"):
                counter_key = rule["counter_key"]
                counters[counter_key] = int(counters.get(counter_key, START_CTRL_NUMBER)) + 1
                ctrl_no = f"{counters[counter_key]:06d}"
                file_name = rule["file_name"].format(date=today, ctrl_no=ctrl_no)
            else:
                file_name = rule["file_name"].format(date=today)

            file_name = ensure_excel_extension(file_name)

            outputs.append(
                {
                    "label": rule["label"],
                    "count": len(matched_rows),
                    "file_name": file_name,
                    "data": output_file.getvalue(),
                }
            )

    save_counters(counters)

    if progress_bar is not None:
        progress_bar.progress(100, text="Done generating files.")

    return outputs


def _save_value(key):
    st.session_state[key] = st.session_state["_" + key]


def _load_value(key, default=None):
    if key not in st.session_state:
        st.session_state[key] = default
    st.session_state["_" + key] = st.session_state[key]


# =========================
# MAIN APP
# =========================
def run():
    st.markdown(CSS, unsafe_allow_html=True)

    if "bpi_cert_uploaded_bytes" not in st.session_state:
        st.session_state["bpi_cert_uploaded_bytes"] = None
    if "bpi_cert_uploaded_name" not in st.session_state:
        st.session_state["bpi_cert_uploaded_name"] = None
    if "bpi_cert_generated_outputs" not in st.session_state:
        st.session_state["bpi_cert_generated_outputs"] = []

    current_counters = load_counters()

    for counter_key in COUNTER_KEYS:
        permanent_key = f"bpi_{counter_key}"
        default_value = int(current_counters.get(counter_key, START_CTRL_NUMBER))
        _load_value(permanent_key, default_value)

    with st.sidebar:
        st.markdown("## ⚙️ Control Center")
        st.markdown(
            '<div class="sidebar-note">Manage the current CTRL counters here. Enter the <b>last used number</b>. The next generated file will use the next value.</div>',
            unsafe_allow_html=True,
        )

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("### Current CTRL Counters")
        current_display = {
            key: st.session_state.get(f"bpi_{key}", START_CTRL_NUMBER)
            for key in COUNTER_KEYS
        }
        st.json(current_display)

        st.markdown("### Edit CTRL Counters")
        for counter_key in COUNTER_KEYS:
            permanent_key = f"bpi_{counter_key}"
            st.number_input(
                counter_key,
                min_value=0,
                step=1,
                key=f"_{permanent_key}",
                on_change=_save_value,
                args=(permanent_key,),
            )

        if st.button("Save CTRL Counters", use_container_width=True, key="bpi_cert_save_counters"):
            counters_to_save = {
                key: int(st.session_state[f"bpi_{key}"])
                for key in COUNTER_KEYS
            }
            save_counters(counters_to_save)
            st.success("CTRL counters updated.")

    st.markdown(
        """
        <div class="hero-card">
            <div class="main-title">BPI Cert Request Drive</div>
            <div class="sub-title">Automated Certificate Request Processing</div>
            <span class="small-chip">Dark Mode</span>
            <span class="small-chip">RUSH Priority</span>
            <span class="small-chip">Per-File CTRL</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Upload Source File")

    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=["xlsx"],
        key="_bpi_cert_file_uploader",
    )

    if uploaded_file is not None:
        st.session_state["bpi_cert_uploaded_bytes"] = uploaded_file.getvalue()
        st.session_state["bpi_cert_uploaded_name"] = uploaded_file.name

    if st.session_state["bpi_cert_uploaded_name"]:
        st.info(f"Loaded file kept in memory: {st.session_state['bpi_cert_uploaded_name']}")

    col1, col2 = st.columns(2)

    with col1:
        generate = st.button("Generate Files", use_container_width=True, key="bpi_cert_generate")

    with col2:
        clear_loaded = st.button("Clear Loaded File", use_container_width=True, key="bpi_cert_clear")

    st.markdown("</div>", unsafe_allow_html=True)

    if generate:
        file_bytes = st.session_state["bpi_cert_uploaded_bytes"]
        if file_bytes is None:
            st.error("Please upload a file first.")
        else:
            progress_placeholder = st.empty()
            try:
                with st.spinner("Generating files, please wait..."):
                    progress_bar = progress_placeholder.progress(0, text="Starting...")
                    st.session_state["bpi_cert_generated_outputs"] = prepare_outputs(
                        file_bytes,
                        progress_bar=progress_bar
                    )
                progress_placeholder.empty()
            except Exception as e:
                progress_placeholder.empty()
                st.session_state["bpi_cert_generated_outputs"] = []
                st.error(str(e))

    if clear_loaded:
        st.session_state["bpi_cert_uploaded_bytes"] = None
        st.session_state["bpi_cert_uploaded_name"] = None
        st.session_state["bpi_cert_generated_outputs"] = []
        if "_bpi_cert_file_uploader" in st.session_state:
            del st.session_state["_bpi_cert_file_uploader"]
        st.rerun()

    outputs = st.session_state["bpi_cert_generated_outputs"]

    if outputs:
        total_files = len(outputs)
        total_rows = sum(item["count"] for item in outputs)

        col1, col2 = st.columns(2)
        with col1:
            st.markdown(
                f'''
                <div class="metric-card">
                    <div class="metric-value">{total_files}</div>
                    <div class="metric-label">Generated Files</div>
                </div>
                ''',
                unsafe_allow_html=True,
            )

        with col2:
            st.markdown(
                f'''
                <div class="metric-card">
                    <div class="metric-value">{total_rows}</div>
                    <div class="metric-label">Total Matched Rows</div>
                </div>
                ''',
                unsafe_allow_html=True,
            )

        st.markdown("<br>", unsafe_allow_html=True)
        st.success("Files generated successfully.")

        for i, output in enumerate(outputs, start=1):
            st.markdown('<div class="result-card">', unsafe_allow_html=True)
            st.markdown(f'<div class="result-title">{output["label"]}</div>', unsafe_allow_html=True)
            st.markdown(
                f'<div class="result-meta">{output["count"]} matched rows • {output["file_name"]}</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                label=f"Download File {i}",
                data=output["data"],
                file_name=output["file_name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"bpi_cert_download_{i}",
            )
            st.markdown("</div>", unsafe_allow_html=True)
    else:
        st.info("Upload a file and click Generate Files to see results.")


if __name__ == "__main__":
    run()
