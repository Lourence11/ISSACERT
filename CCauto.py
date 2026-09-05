import math
import re
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from copy import copy
from io import BytesIO
from datetime import datetime, date
import streamlit as st
import pandas as pd
import msoffcrypto
import openpyxl
from msoffcrypto.format.ooxml import OOXMLFile
from openpyxl.utils import get_column_letter


# ==============================
# CONFIGURATION
# ==============================

TPAP_PASSWORD = "MAD_1Q2026"
BPI_PASSWORD = "Madrid*123"
OUTPUT_FILE_PASSWORD = TPAP_PASSWORD
OUTPUT_FILE_NAME_TEMPLATE = "TPAP Monitoring - MADRID RECOVERY AND SPECIAL PROJECT - {date}.xlsx"
OUTPUT_FILE_DATE_FORMAT = "%d%B%Y"
PROTECTED_STATUSES = {"COMPLIED"}
AUDIT_SKIP_STATUSES = {"DEFAULTED", "UNDER NEGO", "REFUSED/NOT AVAILING"}
AUDIT_STATUS_COLUMN = "STAT (AVAILED, UNDER NEGO, REFUSED/NOT AVAILING, COMPLIED, DEFAULTED)"
SHORT_DATE_NUMBER_FORMAT = "MM/DD/YYYY"
WHOLE_NUMBER_FORMAT = "#,##0"
STATUS_HEADER_CANDIDATES = [AUDIT_STATUS_COLUMN, "Status"]
DATE_UPDATED_HEADER_CANDIDATES = ["Date Updated (nego date/review)", "Date Updated"]
PTP_DATE_HEADER_CANDIDATES = ["PTP Date/Due date", "PTP Date"]
PRIN_HEADER = "PRIN"
FACE_AMOUNT_AUDIT_HEADER = "FACE AMOUNT (OTP/EPA)"
UNDER_NEGO_PAYMENT_AUDIT_STATUSES = {"UNDER NEGO", "REFUSED/NOT AVAILING"}
PARTIAL_PAP_CODE = "PARTIAL"
PL_PAYMENT_WEEK_COLUMN_PATTERN = re.compile(r"^Payment Amount .*?\(Week [1-4]\)$", re.IGNORECASE)

# ==============================
# HELPERS
# ==============================

def decrypt_file(uploaded_file, password):
    uploaded_file.seek(0)
    decrypted_file = BytesIO()
    office_file = msoffcrypto.OfficeFile(uploaded_file)
    office_file.load_key(password=password)
    office_file.decrypt(decrypted_file)
    decrypted_file.seek(0)
    return decrypted_file


def build_output_filename(current_datetime=None):
    current_datetime = current_datetime or datetime.today()
    date_text = current_datetime.strftime(OUTPUT_FILE_DATE_FORMAT).upper()
    return OUTPUT_FILE_NAME_TEMPLATE.format(date=date_text)


def encrypt_output_workbook(workbook_buffer, password):
    workbook_buffer.seek(0)
    encrypted_output = BytesIO()
    OOXMLFile(workbook_buffer).encrypt(password, encrypted_output)
    encrypted_output.seek(0)
    return encrypted_output


def normalize_key(value):
    """Normalize account/customer numbers so matching is more reliable."""
    if pd.isna(value):
        return None

    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return str(int(value))

    if isinstance(value, int):
        return str(value)

    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"\s+", "", text)
    return text


def has_meaningful_value(value):
    if pd.isna(value):
        return False
    return str(value).strip() != ""


def parse_numeric_value(value):
    if pd.isna(value):
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    normalized = (
        text.replace("â‚±", "")
        .replace("₱", "")
        .replace(",", "")
        .replace(" ", "")
    )
    if normalized.endswith("%"):
        normalized = normalized[:-1]

    try:
        return float(normalized)
    except ValueError:
        return None


def parse_percentage_value(value):
    numeric_value = parse_numeric_value(value)
    if numeric_value is None:
        return None

    text = "" if pd.isna(value) else str(value).strip()
    if text.endswith("%") or abs(numeric_value) > 1:
        return numeric_value / 100

    return numeric_value


def round_up_to_nearest_hundred(value):
    numeric_value = parse_numeric_value(value)
    if numeric_value is None:
        return None

    decimal_value = Decimal(str(numeric_value))
    if decimal_value == 0:
        return 0

    hundred = Decimal("100")
    rounded_value = (decimal_value / hundred).to_integral_value(rounding=ROUND_CEILING) * hundred
    return int(rounded_value)


def has_payment_value(value):
    numeric_value = parse_numeric_value(value)
    if numeric_value is not None:
        return numeric_value != 0
    return has_meaningful_value(value)


def get_pl_payment_columns(df):
    return [
        column_name
        for column_name in df.columns
        if PL_PAYMENT_WEEK_COLUMN_PATTERN.match(str(column_name).strip())
    ]


def populate_pl_payment_from_weekly_columns(df):
    df = df.copy()
    payment_columns = get_pl_payment_columns(df)
    if not payment_columns:
        return df

    weekly_payments = df[payment_columns].apply(pd.to_numeric, errors="coerce")
    df["Payment"] = weekly_payments.sum(axis=1, min_count=1)
    return df


def apply_under_nego_payment_audit(df, sheet_name):
    df = df.copy()

    audit_status_source = df[AUDIT_STATUS_COLUMN] if AUDIT_STATUS_COLUMN in df.columns else get_audit_status_series(df)
    status_series = audit_status_source.fillna("").astype(str).str.strip().str.upper()
    current_status_series = df["Status"].fillna("").astype(str).str.strip().str.upper()
    pap_code_series = df["PAP_Code"].fillna("").astype(str).str.strip().str.upper()
    eligible_status_mask = status_series.isin(UNDER_NEGO_PAYMENT_AUDIT_STATUSES)

    if sheet_name == "PL":
        payment_columns = get_pl_payment_columns(df)
        if not payment_columns:
            raise ValueError(
                "PL sheet is missing payment amount week columns required for the UNDER NEGO audit check."
            )
        payment_has_value_mask = df[payment_columns].apply(lambda column: column.apply(has_payment_value)).any(axis=1)
    else:
        payment_columns = ["Payment"]
        payment_has_value_mask = df["Payment"].apply(has_payment_value)

    rows_with_payment_mask = eligible_status_mask & payment_has_value_mask
    rows_to_force_under_nego_mask = rows_with_payment_mask & ~current_status_series.eq("COMPLIED")
    changed_to_under_nego_mask = rows_to_force_under_nego_mask & ~current_status_series.eq("UNDER NEGO")
    df.loc[rows_to_force_under_nego_mask, "Status"] = "UNDER NEGO"

    partial_rows_mask = rows_to_force_under_nego_mask & pap_code_series.eq(PARTIAL_PAP_CODE)
    if "Face Amount" in df.columns:
        df.loc[partial_rows_mask, "Face Amount"] = pd.NA
    if FACE_AMOUNT_AUDIT_HEADER in df.columns:
        df.loc[partial_rows_mask, FACE_AMOUNT_AUDIT_HEADER] = pd.NA

    return df, {
        "type": "under_nego_payment_review",
        "category": "AUDIT",
        "priority": "HIGH",
        "payment_columns": payment_columns,
        "rows_reviewed": int(eligible_status_mask.sum()),
        "rows_with_payment": int(rows_with_payment_mask.sum()),
        "rows_changed_to_under_nego": int(changed_to_under_nego_mask.sum()),
        "partial_rows_blank": int(partial_rows_mask.sum()),
    }, partial_rows_mask


def populate_face_amount_from_discount_principal(df, exclude_mask=None):
    df = df.copy()
    if exclude_mask is None:
        exclude_mask = pd.Series(False, index=df.index)

    discount_source_mask = df["Discount Rate Principal (%)"].apply(has_meaningful_value)
    discount_rates = pd.to_numeric(
        df["Discount Rate Principal (%)"].apply(parse_percentage_value),
        errors="coerce",
    )
    principal_values = pd.to_numeric(
        df[PRIN_HEADER].apply(parse_numeric_value),
        errors="coerce",
    )

    computed_mask = discount_rates.notna() & principal_values.notna() & ~exclude_mask
    computed_values = (discount_rates * principal_values).apply(round_up_to_nearest_hundred)
    if computed_mask.any():
        df.loc[computed_mask, FACE_AMOUNT_AUDIT_HEADER] = computed_values[computed_mask]

    missing_prin_mask = discount_source_mask & principal_values.isna() & ~exclude_mask
    missing_prin_details = df.loc[
        missing_prin_mask,
        ["Customer No.", "PAP_Code", "Discount Rate Principal (%)", PRIN_HEADER],
    ].copy()
    if not missing_prin_details.empty:
        missing_prin_details.insert(0, "Excel Row", missing_prin_details.index + 2)
        missing_prin_details.reset_index(drop=True, inplace=True)

    return df, {
        "type": "face_amount_calculated_from_prin",
        "category": "AUDIT",
        "priority": "HIGH",
        "title": '"FACE AMOUNT (OTP/EPA)" was calculated from "Discount Rate Principal (%)" x "PRIN"',
        "source_rows_found": int(discount_source_mask.sum()),
        "rows_updated": int(computed_mask.sum()),
        "missing_prin_count": int(missing_prin_mask.sum()),
        "details": missing_prin_details,
    }


def get_audit_status_series(df):
    if "Status" in df.columns:
        return df["Status"]

    if AUDIT_STATUS_COLUMN in df.columns:
        return df[AUDIT_STATUS_COLUMN]

    return pd.Series("", index=df.index, dtype="object")


def get_first_existing_column(df, header_candidates):
    for header in header_candidates:
        if header in df.columns:
            return header
    return None


def ensure_canonical_column(df, canonical_name, header_candidates, default_value=pd.NA):
    source_col = get_first_existing_column(df, header_candidates)

    if source_col is not None:
        if canonical_name != source_col or canonical_name not in df.columns:
            df[canonical_name] = df[source_col]
    elif canonical_name not in df.columns:
        df[canonical_name] = default_value

    return df


def coerce_to_excel_datetime(value):
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.to_pydatetime()

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None

        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return None

        return parsed.to_pydatetime()

    return None


def get_decimal_places(value):
    try:
        decimal_value = Decimal(str(value)).normalize()
    except (InvalidOperation, ValueError, TypeError):
        return 0

    return max(0, -decimal_value.as_tuple().exponent)


def get_payment_number_format(value, source_number_format=None):
    normalized_source_format = str(source_number_format).strip() if source_number_format else ""
    numeric_value = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric_value):
        return normalized_source_format or None

    decimal_places = get_decimal_places(numeric_value)
    comma_number_format = "#,##0"
    if decimal_places > 0:
        comma_number_format += "." + ("0" * min(decimal_places, 6))

    if normalized_source_format and normalized_source_format.lower() != "general":
        if "," in normalized_source_format:
            return normalized_source_format
        return comma_number_format

    return comma_number_format


def build_bpi_payment_number_format_lookup(bpi_ws):
    header_map = build_header_map(bpi_ws, header_row=1)
    if "Account No" not in header_map:
        raise ValueError("Column 'Account No' not found in BPI Payment Extraction sheet.")

    account_col = header_map["Account No"]
    payment_col = 4
    number_format_lookup = {}

    for row_idx in range(2, bpi_ws.max_row + 1):
        account_key = normalize_key(bpi_ws.cell(row_idx, account_col).value)
        payment_cell = bpi_ws.cell(row_idx, payment_col)
        payment_value = pd.to_numeric(payment_cell.value, errors="coerce")

        if account_key is None or pd.isna(payment_value):
            continue

        number_format_lookup[account_key] = payment_cell.number_format

    return number_format_lookup


def load_excel_files(tpap_file, bpi_file):
    tpap_decrypted = decrypt_file(tpap_file, TPAP_PASSWORD)
    bpi_decrypted = decrypt_file(bpi_file, BPI_PASSWORD)

    tpap_wb = openpyxl.load_workbook(tpap_decrypted)
    required_tpap_sheets = ["CC", "PL"]
    missing_tpap_sheets = [sheet_name for sheet_name in required_tpap_sheets if sheet_name not in tpap_wb.sheetnames]
    if missing_tpap_sheets:
        raise ValueError(
            "Missing required TPAP Monitoring sheet(s): "
            + ", ".join(f"'{sheet_name}'" for sheet_name in missing_tpap_sheets)
        )
    tpap_ws_map = {sheet_name: tpap_wb[sheet_name] for sheet_name in required_tpap_sheets}

    bpi_wb = openpyxl.load_workbook(bpi_decrypted, data_only=True)
    if "PAYMENTS" not in bpi_wb.sheetnames:
        raise ValueError("Sheet 'PAYMENTS' not found in BPI Payment Extraction file.")
    bpi_ws = bpi_wb["PAYMENTS"]
    bpi_payment_number_formats = build_bpi_payment_number_format_lookup(bpi_ws)

    tpap_decrypted.seek(0)
    bpi_decrypted.seek(0)

    tpap_sheets = pd.read_excel(
        tpap_decrypted,
        sheet_name=required_tpap_sheets,
        engine="openpyxl",
    )
    bpi_df = pd.read_excel(bpi_decrypted, sheet_name="PAYMENTS", engine="openpyxl")

    for sheet_name, tpap_df in tpap_sheets.items():
        tpap_df.columns = tpap_df.columns.astype(str).str.strip()
        tpap_sheets[sheet_name] = tpap_df
    bpi_df.columns = bpi_df.columns.astype(str).str.strip()

    return tpap_sheets, bpi_df, tpap_wb, tpap_ws_map, bpi_payment_number_formats


def validate_required_columns(tpap_df, bpi_df=None, sheet_name="CC", require_bpi_columns=False):
    required_tpap_columns = [
        "Customer No.",
        "PAP",
        "PAP_Code",
        PRIN_HEADER,
        "EPA TERM",
        "Program Start Date",
        "Program End Date",
        FACE_AMOUNT_AUDIT_HEADER,
        "Discount Rate Principal (%)",
        "Discount Rate Charges (%)",
    ]
    missing_tpap_columns = [col for col in required_tpap_columns if col not in tpap_df.columns]
    if missing_tpap_columns:
        st.error(
            f"Error: Missing required column(s) in TPAP Monitoring sheet '{sheet_name}': "
            + ", ".join(f"'{col}'" for col in missing_tpap_columns)
        )
        return False

    if require_bpi_columns and (bpi_df is None or "Account No" not in bpi_df.columns):
        st.error("Error: 'Account No' column not found in BPI Payment Extraction sheet.")
        return False

    return True


def ensure_columns_exist(df):
    df = df.copy()
    df = ensure_canonical_column(df, "Payment", ["Payment"])
    df = ensure_canonical_column(df, "Face Amount", ["Face Amount", FACE_AMOUNT_AUDIT_HEADER])
    df = ensure_canonical_column(df, "Status", STATUS_HEADER_CANDIDATES)
    df = ensure_canonical_column(df, "Date Updated", DATE_UPDATED_HEADER_CANDIDATES)
    df = ensure_canonical_column(df, "PTP Date", PTP_DATE_HEADER_CANDIDATES)
    return df


def get_bpi_payment_column(bpi_df):
    """
    Use COLUMN D from the PAYMENTS sheet of BPI PYTS EXTRACTION.
    Excel column D = pandas column index 3.
    """
    if len(bpi_df.columns) < 4:
        raise ValueError(
            "BPI PAYMENTS sheet does not have enough columns. "
            "Column D was expected but not found."
        )

    return bpi_df.columns[3]


def copy_cell_style(source_cell, target_cell):
    target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format


def write_cell_value_preserving_format(cell, value, number_format_override=None):
    original_style = copy(cell._style)
    original_font = copy(cell.font)
    original_fill = copy(cell.fill)
    original_border = copy(cell.border)
    original_alignment = copy(cell.alignment)
    original_protection = copy(cell.protection)
    original_number_format = cell.number_format

    if pd.isna(value):
        cell.value = None
    elif isinstance(value, pd.Timestamp):
        cell.value = value.to_pydatetime()
    else:
        cell.value = value

    cell._style = original_style
    cell.font = original_font
    cell.fill = original_fill
    cell.border = original_border
    cell.alignment = original_alignment
    cell.protection = original_protection
    cell.number_format = number_format_override or original_number_format


def build_header_map(ws, header_row=1):
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if value is not None:
            header_map[str(value).strip()] = col_idx
    return header_map


def ensure_worksheet_column(ws, header_candidates, create_header=None, header_row=1):
    header_map = build_header_map(ws, header_row=header_row)

    for header in header_candidates:
        if header in header_map:
            return header_map[header]

    new_header = create_header or header_candidates[0]
    ensure_worksheet_columns(ws, [new_header], header_row=header_row)
    return build_header_map(ws, header_row=header_row)[new_header]


def ensure_worksheet_columns(ws, required_headers, header_row=1):
    header_map = build_header_map(ws, header_row)

    for header in required_headers:
        if header in header_map:
            continue

        template_col = ws.max_column
        new_col = ws.max_column + 1

        template_header_cell = ws.cell(header_row, template_col)
        new_header_cell = ws.cell(header_row, new_col)
        new_header_cell.value = header
        copy_cell_style(template_header_cell, new_header_cell)

        template_letter = get_column_letter(template_col)
        new_letter = get_column_letter(new_col)
        ws.column_dimensions[new_letter].width = ws.column_dimensions[template_letter].width

        for row_idx in range(header_row + 1, ws.max_row + 1):
            copy_cell_style(ws.cell(row_idx, template_col), ws.cell(row_idx, new_col))

        header_map[header] = new_col

    return header_map


# ==============================
# PAYMENT UPDATE
# ==============================

def update_payment_data(tpap_df, bpi_df, bpi_payment_number_formats=None):
    payment_col = get_bpi_payment_column(bpi_df)

    temp_bpi = bpi_df.copy()
    temp_bpi["Account Key"] = temp_bpi["Account No"].apply(normalize_key)
    temp_bpi[payment_col] = pd.to_numeric(temp_bpi[payment_col], errors="coerce")

    # Keep only valid account number rows
    temp_bpi = temp_bpi.dropna(subset=["Account Key"])

    # Keep only rows with actual payment values
    temp_bpi = temp_bpi.dropna(subset=[payment_col])

    # Sum all payment rows that belong to the same account number.
    temp_bpi = temp_bpi.groupby("Account Key", as_index=False)[payment_col].sum()

    payment_lookup = temp_bpi.set_index("Account Key")[payment_col]

    temp_tpap = tpap_df.copy()
    temp_tpap["Customer Key"] = temp_tpap["Customer No."].apply(normalize_key)

    mapped_payment = temp_tpap["Customer Key"].map(payment_lookup)
    payment_updated_mask = mapped_payment.notna()

    # Update only if BPI has a payment; otherwise keep current CC payment
    temp_tpap["Payment"] = temp_tpap["Payment"].where(mapped_payment.isna(), mapped_payment)

    payment_format_overrides = {}
    if bpi_payment_number_formats is None:
        bpi_payment_number_formats = {}

    for df_index in temp_tpap.index[payment_updated_mask]:
        customer_key = temp_tpap.at[df_index, "Customer Key"]
        source_number_format = bpi_payment_number_formats.get(customer_key)
        payment_value = temp_tpap.at[df_index, "Payment"]
        payment_format_overrides[df_index] = get_payment_number_format(
            payment_value,
            source_number_format=source_number_format,
        )

    temp_tpap.drop(columns=["Customer Key"], inplace=True)
    return temp_tpap, payment_format_overrides


# ==============================
# PAYMENT LOGIC
# ==============================

def apply_payment_logic(df):
    df = df.copy()

    current_datetime = datetime.today()
    today_value = pd.Timestamp(current_datetime.date())
    today_date = current_datetime.date()

    df["Payment"] = pd.to_numeric(df["Payment"], errors="coerce")
    df["Face Amount"] = pd.to_numeric(df["Face Amount"], errors="coerce")
    pap_code_series = df["PAP_Code"].fillna("").astype(str).str.strip().str.upper()

    ptp_dates = pd.to_datetime(df["PTP Date"], errors="coerce") if "PTP Date" in df.columns else pd.Series([pd.NaT] * len(df))

    for index in df.index:
        current_status = str(df.at[index, "Status"]).strip().upper() if pd.notna(df.at[index, "Status"]) else ""
        if current_status in PROTECTED_STATUSES:
            continue

        payment = df.at[index, "Payment"]
        face_amount = df.at[index, "Face Amount"]
        ptp_date = ptp_dates.loc[index] if index in ptp_dates.index else pd.NaT
        ptp_due_date = ptp_date.date() if pd.notna(ptp_date) else None
        ptp_is_past_due = ptp_due_date is not None and ptp_due_date < today_date
        has_actual_payment = pd.notna(payment) and payment > 0

        if "BAU" in pap_code_series.loc[index]:
            continue

        if current_status in UNDER_NEGO_PAYMENT_AUDIT_STATUSES and not has_actual_payment:
            continue

        if has_actual_payment and pd.notna(face_amount) and payment >= face_amount:
            df.at[index, "Status"] = "COMPLIED"
            df.at[index, "Date Updated"] = today_value
            continue

        if has_actual_payment:
            df.at[index, "Status"] = "AVAILED"
            df.at[index, "Date Updated"] = today_value
            continue

        if pd.notna(ptp_date):
            if (pd.isna(payment) or payment <= 0) and ptp_due_date < today_date:
                df.at[index, "Status"] = "DEFAULTED"
                df.at[index, "Date Updated"] = today_value

            elif (pd.isna(payment) or payment <= 0) and ptp_due_date >= today_date:
                df.at[index, "Status"] = "AVAILED"
                df.at[index, "Date Updated"] = today_value

    return df


def apply_audit_checks(df, sheet_name="CC"):
    df = df.copy()
    df, under_nego_payment_audit, partial_face_amount_blank_mask = apply_under_nego_payment_audit(
        df,
        sheet_name=sheet_name,
    )
    df, face_amount_calculation_audit = populate_face_amount_from_discount_principal(
        df,
        exclude_mask=partial_face_amount_blank_mask,
    )

    pap_code_series = df["PAP_Code"].fillna("").astype(str)
    pap_series = df["PAP"].fillna("").astype(str).str.strip().str.upper()
    audit_status_series = get_audit_status_series(df).fillna("").astype(str).str.strip().str.upper()
    audit_skip_mask = audit_status_series.isin(AUDIT_SKIP_STATUSES)
    bau_mask = pap_code_series.str.contains("BAU", case=False, na=False) & ~audit_skip_mask
    tpap_mask = pap_code_series.str.contains("TPAP", case=False, na=False) & ~audit_skip_mask
    epa_mask = pap_series.eq("EPA") & ~audit_skip_mask

    start_has_value = df["Program Start Date"].apply(has_meaningful_value)
    end_has_value = df["Program End Date"].apply(has_meaningful_value)
    epa_term_has_value = df["EPA TERM"].apply(has_meaningful_value)
    tpap_face_amount_has_value = df["FACE AMOUNT (OTP/EPA)"].apply(has_meaningful_value)
    tpap_discount_principal_has_value = df["Discount Rate Principal (%)"].apply(has_meaningful_value)
    cleared_mask = bau_mask & (start_has_value | end_has_value)
    discount_has_value = df["Discount Rate Principal (%)"].apply(has_meaningful_value)
    discount_violation_mask = bau_mask & discount_has_value
    charges_has_value = df["Discount Rate Charges (%)"].apply(has_meaningful_value)
    charges_filled_mask = bau_mask & ~charges_has_value
    epa_missing_term_mask = epa_mask & ~epa_term_has_value
    tpap_missing_dates_mask = tpap_mask & ~(start_has_value & end_has_value)
    tpap_missing_face_amount_mask = tpap_mask & ~tpap_face_amount_has_value
    tpap_missing_discount_principal_mask = tpap_mask & ~tpap_discount_principal_has_value

    df.loc[bau_mask, "Program Start Date"] = pd.NA
    df.loc[bau_mask, "Program End Date"] = pd.NA
    df.loc[charges_filled_mask, "Discount Rate Charges (%)"] = "0%"

    discount_violations = df.loc[
        discount_violation_mask,
        ["Customer No.", "PAP_Code", "Discount Rate Principal (%)"],
    ].copy()
    if not discount_violations.empty:
        discount_violations.insert(0, "Excel Row", discount_violations.index + 2)
        discount_violations.reset_index(drop=True, inplace=True)

    tpap_date_violations = df.loc[
        tpap_missing_dates_mask,
        ["Customer No.", "PAP_Code", "Program Start Date", "Program End Date"],
    ].copy()
    if not tpap_date_violations.empty:
        missing_fields = []
        for row_index in tpap_date_violations.index:
            fields = []
            if not start_has_value.loc[row_index]:
                fields.append("Program Start Date")
            if not end_has_value.loc[row_index]:
                fields.append("Program End Date")
            missing_fields.append(", ".join(fields))

        tpap_date_violations.insert(0, "Excel Row", tpap_date_violations.index + 2)
        tpap_date_violations["Missing Fields"] = missing_fields
        tpap_date_violations.reset_index(drop=True, inplace=True)

    tpap_face_amount_violations = df.loc[
        tpap_missing_face_amount_mask,
        ["Customer No.", "PAP_Code", "FACE AMOUNT (OTP/EPA)"],
    ].copy()
    if not tpap_face_amount_violations.empty:
        tpap_face_amount_violations.insert(0, "Excel Row", tpap_face_amount_violations.index + 2)
        tpap_face_amount_violations.reset_index(drop=True, inplace=True)

    tpap_discount_principal_violations = df.loc[
        tpap_missing_discount_principal_mask,
        ["Customer No.", "PAP_Code", "Discount Rate Principal (%)"],
    ].copy()
    if not tpap_discount_principal_violations.empty:
        tpap_discount_principal_violations.insert(
            0,
            "Excel Row",
            tpap_discount_principal_violations.index + 2,
        )
        tpap_discount_principal_violations.reset_index(drop=True, inplace=True)

    epa_term_violations = df.loc[
        epa_missing_term_mask,
        ["Customer No.", "PAP", "EPA TERM"],
    ].copy()
    if not epa_term_violations.empty:
        epa_term_violations.insert(0, "Excel Row", epa_term_violations.index + 2)
        epa_term_violations.reset_index(drop=True, inplace=True)

    audit_results = [
        under_nego_payment_audit,
        {
            "type": "epa_term_required_for_epa",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": 'EPA rows must have "EPA TERM"',
            "epa_rows_found": int(epa_mask.sum()),
            "violations_count": int(epa_missing_term_mask.sum()),
            "details": epa_term_violations,
        },
        {
            "type": "discount_rate_principal_required_for_tpap",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": 'TPAP rows must have "Discount Rate Principal (%)"',
            "tpap_rows_found": int(tpap_mask.sum()),
            "violations_count": int(tpap_missing_discount_principal_mask.sum()),
            "details": tpap_discount_principal_violations,
        },
        face_amount_calculation_audit,
        {
            "type": "face_amount_required_for_tpap",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": 'TPAP rows must have "FACE AMOUNT (OTP/EPA)"',
            "tpap_rows_found": int(tpap_mask.sum()),
            "violations_count": int(tpap_missing_face_amount_mask.sum()),
            "details": tpap_face_amount_violations,
        },
        {
            "type": "program_dates_required_for_tpap",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": "TPAP rows must have Program Start Date and Program End Date",
            "tpap_rows_found": int(tpap_mask.sum()),
            "violations_count": int(tpap_missing_dates_mask.sum()),
            "details": tpap_date_violations,
        },
        {
            "type": "discount_rate_charges_zero_for_bau",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": 'BAU rows with blank "Discount Rate Charges (%)" were set to "0%"',
            "bau_rows_found": int(bau_mask.sum()),
            "rows_updated": int(charges_filled_mask.sum()),
        },
        {
            "type": "discount_rate_principal_blank_for_bau",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": 'BAU rows must have blank "Discount Rate Principal (%)"',
            "bau_rows_found": int(bau_mask.sum()),
            "violations_count": int(discount_violation_mask.sum()),
            "details": discount_violations,
        },
        {
            "type": "program_dates_blank_for_bau",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": "BAU PAP_Code must have blank Program Start Date and Program End Date",
            "bau_rows_found": int(bau_mask.sum()),
            "rows_cleared": int(cleared_mask.sum()),
        }
    ]

    return df, audit_results


def apply_short_date_format_to_columns(ws, header_row=1):
    date_column_candidates = [
        ["Date Updated (nego date/review)", "Date Updated"],
        ["PTP Date/Due date", "PTP Date"],
    ]

    for header_candidates in date_column_candidates:
        header_map = build_header_map(ws, header_row=header_row)
        target_col = None

        for header in header_candidates:
            if header in header_map:
                target_col = header_map[header]
                break

        if target_col is None:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row_idx, target_col)
            formatted_value = coerce_to_excel_datetime(cell.value)

            if formatted_value is None:
                continue

            write_cell_value_preserving_format(
                cell,
                formatted_value,
                number_format_override=SHORT_DATE_NUMBER_FORMAT,
            )


def clear_refused_related_columns(ws, header_row=1):
    header_map = build_header_map(ws, header_row=header_row)
    status_col = None

    for header in STATUS_HEADER_CANDIDATES:
        if header in header_map:
            status_col = header_map[header]
            break

    if status_col is None:
        status_col = 8

    columns_to_clear = ["K", "L", "N", "O"]

    for row_idx in range(header_row + 1, ws.max_row + 1):
        status_value = str(ws.cell(row_idx, status_col).value or "").strip().upper()
        if status_value != "REFUSED/NOT AVAILING":
            continue

        for column_letter in columns_to_clear:
            write_cell_value_preserving_format(ws[f"{column_letter}{row_idx}"], None)


# ==============================
# WRITE BACK TO ORIGINAL WORKBOOK
# ==============================

def write_updated_values_to_original_sheet(ws, df, header_row=1, payment_format_overrides=None):
    column_specs = [
        ("Payment", ["Payment"], "Payment"),
        ("Status", STATUS_HEADER_CANDIDATES, "Status"),
        ("Date Updated", DATE_UPDATED_HEADER_CANDIDATES, "Date Updated"),
        ("Program Start Date", ["Program Start Date"], "Program Start Date"),
        ("Program End Date", ["Program End Date"], "Program End Date"),
        (FACE_AMOUNT_AUDIT_HEADER, [FACE_AMOUNT_AUDIT_HEADER], FACE_AMOUNT_AUDIT_HEADER),
        ("Discount Rate Charges (%)", ["Discount Rate Charges (%)"], "Discount Rate Charges (%)"),
    ]
    header_map = {
        df_col: ensure_worksheet_column(
            ws,
            header_candidates=header_candidates,
            create_header=create_header,
            header_row=header_row,
        )
        for df_col, header_candidates, create_header in column_specs
    }
    payment_format_overrides = payment_format_overrides or {}

    for df_col, _, _ in column_specs:
        col_idx = header_map[df_col]

        for excel_row, df_index in enumerate(df.index, start=header_row + 1):
            cell = ws.cell(excel_row, col_idx)
            value = df.at[df_index, df_col]
            number_format_override = None

            if df_col == "Payment":
                number_format_override = payment_format_overrides.get(df_index)
                if number_format_override is None:
                    number_format_override = get_payment_number_format(value)
            elif df_col == "Date Updated":
                number_format_override = SHORT_DATE_NUMBER_FORMAT
            elif df_col == FACE_AMOUNT_AUDIT_HEADER:
                number_format_override = WHOLE_NUMBER_FORMAT

            write_cell_value_preserving_format(
                cell,
                value,
                number_format_override=number_format_override,
            )


def render_audit_results(sheet_name, audit_results):
    st.subheader(f"AUDIT - {sheet_name}")

    for audit in audit_results:
        if audit["type"] == "under_nego_payment_review":
            if audit["rows_reviewed"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with status "UNDER NEGO" or "REFUSED/NOT AVAILING" were found.'
                )
            elif audit["rows_with_payment"] > 0:
                st.success(
                    f'[{audit["category"]}] Priority check applied: '
                    f'{audit["rows_with_payment"]} row(s) with status "UNDER NEGO" or "REFUSED/NOT AVAILING" '
                    'had payment, so the status was set or kept as "UNDER NEGO".'
                )
            else:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'Rows with status "UNDER NEGO" or "REFUSED/NOT AVAILING" were found, but no payment was detected, so the status was kept as is.'
                )

            if audit["rows_changed_to_under_nego"] > 0:
                st.warning(
                    f'[{audit["category"]}] Priority check applied: '
                    f'{audit["rows_changed_to_under_nego"]} row(s) were changed from "REFUSED/NOT AVAILING" to "UNDER NEGO".'
                )

            if audit["partial_rows_blank"] > 0:
                st.warning(
                    f'[{audit["category"]}] Priority check applied: '
                    f'{audit["partial_rows_blank"]} row(s) with PAP_Code "PARTIAL" had FACE AMOUNT cleared and kept blank.'
                )
        elif audit["type"] == "epa_term_required_for_epa":
            if audit["epa_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP equal to "EPA" were found for EPA TERM.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} EPA row(s) are missing "EPA TERM".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["epa_rows_found"]} EPA row(s) were checked and "EPA TERM" is present.'
                )
        elif audit["type"] == "discount_rate_principal_required_for_tpap":
            if audit["tpap_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "TPAP" were found for Discount Rate Principal (%).'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} TPAP row(s) are missing "Discount Rate Principal (%)".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["tpap_rows_found"]} TPAP row(s) were checked and "Discount Rate Principal (%)" is present.'
                )
        elif audit["type"] == "face_amount_required_for_tpap":
            if audit["tpap_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "TPAP" were found for FACE AMOUNT (OTP/EPA).'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} TPAP row(s) are missing "FACE AMOUNT (OTP/EPA)".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["tpap_rows_found"]} TPAP row(s) were checked and "FACE AMOUNT (OTP/EPA)" is present.'
                )
        elif audit["type"] == "face_amount_calculated_from_prin":
            if audit["source_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with a value in "Discount Rate Principal (%)" were found for FACE AMOUNT calculation.'
                )
            elif audit["rows_updated"] > 0:
                st.success(
                    f'[{audit["category"]}] Priority check applied: '
                    f'{audit["rows_updated"]} row(s) had "FACE AMOUNT (OTP/EPA)" calculated from '
                    '"Discount Rate Principal (%)" x "PRIN", rounded up to the nearest hundred.'
                )
            else:
                st.warning(
                    f'[{audit["category"]}] Priority check completed: '
                    'Rows with "Discount Rate Principal (%)" were found, but no FACE AMOUNT values were calculated.'
                )

            if audit["missing_prin_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["missing_prin_count"]} row(s) have "Discount Rate Principal (%)" '
                    'but missing or invalid "PRIN".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
        elif audit["type"] == "program_dates_required_for_tpap":
            if audit["tpap_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "TPAP" were found for Program Start Date and Program End Date.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} TPAP row(s) are missing Program Start Date and/or Program End Date.'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["tpap_rows_found"]} TPAP row(s) were checked and all required program dates are present.'
                )
        elif audit["type"] == "discount_rate_charges_zero_for_bau":
            if audit["bau_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "BAU" were found for Discount Rate Charges (%).'
                )
            elif audit["rows_updated"] > 0:
                st.warning(
                    f'[{audit["category"]}] Priority check applied: '
                    f'{audit["rows_updated"]} BAU row(s) with blank "Discount Rate Charges (%)" were updated to "0%".'
                )
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["bau_rows_found"]} BAU row(s) were checked and "Discount Rate Charges (%)" already had values.'
                )
        elif audit["type"] == "discount_rate_principal_blank_for_bau":
            if audit["bau_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "BAU" were found for Discount Rate Principal (%).'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} BAU row(s) have a value in "Discount Rate Principal (%)".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["bau_rows_found"]} BAU row(s) were checked and "Discount Rate Principal (%)" is blank.'
                )
        elif audit["rows_cleared"] > 0:
            st.warning(
                f'[{audit["category"]}] Priority check applied: '
                f'{audit["rows_cleared"]} row(s) with PAP_Code containing "BAU" '
                "had Program Start Date and Program End Date cleared."
            )
        elif audit["bau_rows_found"] > 0:
            st.success(
                f'[{audit["category"]}] Priority check passed: '
                f'{audit["bau_rows_found"]} BAU row(s) were checked and the program dates were already blank.'
            )
        else:
            st.info(
                f'[{audit["category"]}] Priority check completed: '
                'No rows with PAP_Code containing "BAU" were found.'
            )


# ==============================
# MAIN
# ==============================

def run():
    st.title("CC and PL Sheets Audit Automation")

    uploaded_tpap_file = st.file_uploader(
        "Upload TPAP Monitoring Excel file",
        type=["xlsx", "xls", "xlsm"]
    )

    uploaded_bpi_file = st.file_uploader(
        "Upload BPI Payment Extraction Excel file",
        type=["xlsx", "xls", "xlsm"]
    )

    if uploaded_tpap_file is None or uploaded_bpi_file is None:
        st.info("Please upload both TPAP Monitoring and BPI Payment Extraction Excel files.")
        return

    st.write("Processing files...")

    try:
        tpap_sheets, bpi_df, tpap_wb, tpap_ws_map, bpi_payment_number_formats = load_excel_files(
            uploaded_tpap_file,
            uploaded_bpi_file,
        )

        cc_df = tpap_sheets["CC"]
        pl_df = tpap_sheets["PL"]
        cc_ws = tpap_ws_map["CC"]
        pl_ws = tpap_ws_map["PL"]

        if not validate_required_columns(
            cc_df,
            bpi_df=bpi_df,
            sheet_name="CC",
            require_bpi_columns=True,
        ):
            return

        if not validate_required_columns(pl_df, sheet_name="PL"):
            return

        cc_df = ensure_columns_exist(cc_df)
        pl_df = ensure_columns_exist(pl_df)
        pl_df = populate_pl_payment_from_weekly_columns(pl_df)

        cc_df, payment_format_overrides = update_payment_data(
            cc_df,
            bpi_df,
            bpi_payment_number_formats=bpi_payment_number_formats,
        )
        cc_df = apply_payment_logic(cc_df)
        pl_df = apply_payment_logic(pl_df)
        cc_df, cc_audit_results = apply_audit_checks(cc_df, sheet_name="CC")
        pl_df, pl_audit_results = apply_audit_checks(pl_df, sheet_name="PL")

        write_updated_values_to_original_sheet(
            cc_ws,
            cc_df,
            header_row=1,
            payment_format_overrides=payment_format_overrides,
        )
        clear_refused_related_columns(cc_ws, header_row=1)
        apply_short_date_format_to_columns(cc_ws, header_row=1)

        write_updated_values_to_original_sheet(
            pl_ws,
            pl_df,
            header_row=1,
        )
        clear_refused_related_columns(pl_ws, header_row=1)
        apply_short_date_format_to_columns(pl_ws, header_row=1)

        output = BytesIO()
        tpap_wb.save(output)
        encrypted_output = encrypt_output_workbook(output, OUTPUT_FILE_PASSWORD)
        output_file_name = build_output_filename()

        render_audit_results("CC", cc_audit_results)
        render_audit_results("PL", pl_audit_results)

        st.success("Automation completed successfully!")
        st.caption(f'Download password: "{OUTPUT_FILE_PASSWORD}"')

        st.download_button(
            label="Download Processed File",
            data=encrypted_output.getvalue(),
            file_name=output_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"An error occurred: {e}")


if __name__ == "__main__":
    run()
