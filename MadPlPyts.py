import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.datetime import from_excel
from io import BytesIO
from datetime import datetime, date, timedelta
import calendar
import re
import msoffcrypto

# Auto password
DEFAULT_PASSWORD = "MAD_Mar2026"

# --- SETTINGS ---
SHEET_PREFIX = "payment as of"
DATE_COL = "F"
AMOUNT_COL = "E"
LAN2_COL = "C"
DATA_START_ROW = 2
NEW_SHEET_NAME = "Sheet1"
# ----------------


def find_payment_as_of_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower().startswith(SHEET_PREFIX):
            return name
    return None


def normalize_lan2(val):
    if val is None:
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).strip()
    return str(val).strip()


def parse_excel_date(val):
    if val is None or val == "":
        return None

    if isinstance(val, datetime):
        return val.date()

    if isinstance(val, date):
        return val

    if isinstance(val, (int, float)):
        try:
            converted = from_excel(val)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except Exception:
            return None

    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None

        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass

    return None


def month_str_to_num(m):
    m = m.strip().lower()
    month_map = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    return month_map.get(m)


def extract_month_year_from_sheetname(sheet_name):
    """
    Example:
    'payment as of MAR 2026' -> (3, 2026)
    """
    s = sheet_name.strip()
    s2 = re.sub(r'(?i)^\s*payment\s+as\s+of\s*', '', s).strip()

    parts = re.split(r"\s+", s2)
    if len(parts) >= 2:
        m_num = month_str_to_num(parts[0])
        try:
            y = int(parts[1])
        except Exception:
            y = None

        if m_num and y:
            return m_num, y

    today = datetime.now().date()
    return today.month, today.year


def first_monday(year, month):
    d = date(year, month, 1)
    while d.weekday() != 0:  # Monday = 0
        d += timedelta(days=1)
    return d


def build_week_ranges(year, month):
    """
    Builds 4 weekly ranges (Mon-Fri) starting from the first Monday of the month.
    """
    fm = first_monday(year, month)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    ranges = []
    for i in range(4):
        start = fm + timedelta(days=7 * i)
        if start > last_day:
            break

        end = start + timedelta(days=4)  # Mon-Fri
        if end > last_day:
            end = last_day

        ranges.append((start, end))

    return ranges, fm, last_day


def week_bucket(d, fm):
    """
    Maps any date to week 1-4 based on 7-day blocks from first Monday.
    Dates before first Monday -> week 1
    Dates after week 4 -> week 4
    """
    delta = (d - fm).days
    if delta < 0:
        return 1

    w = (delta // 7) + 1
    if w < 1:
        return 1
    if w > 4:
        return 4
    return w


def set_date_cell(ws, cell_addr, d):
    ws[cell_addr] = d
    ws[cell_addr].number_format = "mm/dd/yyyy"


def add_amount_to_cell(ws, cell_addr, amount):
    current = ws[cell_addr].value

    if current is None or current == "":
        ws[cell_addr] = amount
    else:
        ws[cell_addr] = current + amount


def set_latest_date(ws, cell_addr, new_date):
    """
    Keep only the latest date for that LAN2 in that week.
    """
    current_date = parse_excel_date(ws[cell_addr].value)

    if current_date is None or new_date > current_date:
        set_date_cell(ws, cell_addr, new_date)


def process_and_input_data_in_existing_sheet(file_stream):
    wb = load_workbook(file_stream)

    sheet_name = find_payment_as_of_sheet(wb)
    if not sheet_name:
        raise Exception("No sheet found that starts with 'payment as of'.")

    ws = wb[sheet_name]

    # Get month/year from sheet name
    m, y = extract_month_year_from_sheetname(sheet_name)
    week_ranges, fm, _ = build_week_ranges(y, m)

    # Ensure Sheet1 exists
    if NEW_SHEET_NAME not in wb.sheetnames:
        new_sheet = wb.create_sheet(NEW_SHEET_NAME)
    else:
        new_sheet = wb[NEW_SHEET_NAME]

    # Clear Sheet1 completely
    if new_sheet.max_row > 1:
        new_sheet.delete_rows(2, new_sheet.max_row - 1)

    for cell in ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1"]:
        new_sheet[cell] = None

    # Headers
    new_sheet["A1"] = "LAN2"

    def label_week(i):
        if i < len(week_ranges):
            s, e = week_ranges[i]
            return f"Payment week {i+1} ({s.strftime('%b %d')} to {e.strftime('%d')})"
        return f"Payment week {i+1}"

    new_sheet["B1"] = label_week(0)
    new_sheet["C1"] = "Payment Date (Week 1)"
    new_sheet["D1"] = label_week(1)
    new_sheet["E1"] = "Payment Date (Week 2)"
    new_sheet["F1"] = label_week(2)
    new_sheet["G1"] = "Payment Date (Week 3)"
    new_sheet["H1"] = label_week(3)
    new_sheet["I1"] = "Payment Date (Week 4)"
    new_sheet["J1"] = "TOTAL"

    # Map each LAN2 to one output row
    lan2_to_row = {}
    out_row = 2

    for row in range(DATA_START_ROW, ws.max_row + 1):
        lan_value = ws[f"{LAN2_COL}{row}"].value
        amount_value = ws[f"{AMOUNT_COL}{row}"].value
        payment_date = parse_excel_date(ws[f"{DATE_COL}{row}"].value)

        lan_key = normalize_lan2(lan_value)

        if not lan_key:
            continue
        if payment_date is None:
            continue
        if amount_value is None or amount_value == 0:
            continue

        if lan_key not in lan2_to_row:
            lan2_to_row[lan_key] = out_row
            new_sheet[f"A{out_row}"] = lan_key
            out_row += 1

        target_row = lan2_to_row[lan_key]
        week_no = week_bucket(payment_date, fm)

        if week_no == 1:
            add_amount_to_cell(new_sheet, f"B{target_row}", amount_value)
            set_latest_date(new_sheet, f"C{target_row}", payment_date)
        elif week_no == 2:
            add_amount_to_cell(new_sheet, f"D{target_row}", amount_value)
            set_latest_date(new_sheet, f"E{target_row}", payment_date)
        elif week_no == 3:
            add_amount_to_cell(new_sheet, f"F{target_row}", amount_value)
            set_latest_date(new_sheet, f"G{target_row}", payment_date)
        else:
            add_amount_to_cell(new_sheet, f"H{target_row}", amount_value)
            set_latest_date(new_sheet, f"I{target_row}", payment_date)

    # TOTAL = week1 + week2 + week3 + week4
    for r in range(2, out_row):
        new_sheet[f"J{r}"] = f"=SUM(B{r},D{r},F{r},H{r})"

    # Formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Header formatting
    for col in range(1, 11):
        col_letter = chr(64 + col)
        new_sheet[f"{col_letter}1"].font = bold_font
        if col == 1:
            new_sheet[f"{col_letter}1"].fill = red_fill
        new_sheet[f"{col_letter}1"].alignment = center_alignment

    # Center align used cells
    for col in range(1, 11):
        col_letter = chr(64 + col)
        for r in range(1, out_row):
            new_sheet[f"{col_letter}{r}"].alignment = center_alignment

    # Adjust column widths
    for col_cells in new_sheet.iter_cols(min_col=1, max_col=10):
        max_length = 0
        column_letter = col_cells[0].column_letter

        for cell in col_cells:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass

        new_sheet.column_dimensions[column_letter].width = max_length + 2

    return wb


def load_workbook_with_optional_password(uploaded, password_str=DEFAULT_PASSWORD):
    uploaded.seek(0)
    raw = uploaded.read()

    # Try normal open first
    try:
        return process_and_input_data_in_existing_sheet(BytesIO(raw))
    except Exception as e:
        if not password_str:
            raise e

    # If encrypted, decrypt using default password
    uploaded.seek(0)
    decrypted = BytesIO()
    office_file = msoffcrypto.OfficeFile(uploaded)
    office_file.load_key(password=password_str)
    office_file.decrypt(decrypted)
    decrypted.seek(0)

    return process_and_input_data_in_existing_sheet(decrypted)


# ==============================
# MAIN
# ==============================
def run():
    st.title("MAD PL PYTS AUTOMATION")

    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

    if uploaded_file:
        try:
            wb = load_workbook_with_optional_password(uploaded_file)

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            st.success("Done! Same LAN2 is placed in one row, amounts are summed by week, and the latest payment date per week is kept.")

            st.download_button(
                "Download Updated Excel",
                data=out,
                file_name="MADRID PL PYTS(AUTOMATED).xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as err:
            st.error(f"Error: {err}")


if __name__ == "__main__":
    run()
