import math
import re
from decimal import Decimal, InvalidOperation, ROUND_CEILING, ROUND_HALF_UP
from copy import copy
from io import BytesIO
from datetime import datetime, date
from pathlib import Path
import streamlit as st
import pandas as pd
import msoffcrypto
import openpyxl
import xlrd
from msoffcrypto.format.ooxml import OOXMLFile
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile


# ==============================
# CONFIGURATION
# ==============================

BPI_PASSWORD = "Madrid*123"
PRIN_DATABASE_PASSWORD = "BPI"
OUTPUT_FILE_NAME_TEMPLATE = "TPAP Monitoring - MADRID RECOVERY AND SPECIAL PROJECT - {date}.xlsx"
OUTPUT_FILE_DATE_FORMAT = "%d%B%Y"
PRIN_DATABASE_SHEET_NAME = "DB"
PROTECTED_STATUSES = {"COMPLIED"}
AUDIT_SKIP_STATUSES = {"DEFAULTED", "UNDER NEGO", "REFUSED/NOT AVAILING"}
AUDIT_STATUS_COLUMN = "STAT (AVAILED, UNDER NEGO, REFUSED/NOT AVAILING, COMPLIED, DEFAULTED)"
SHORT_DATE_NUMBER_FORMAT = "MM/DD/YYYY"
WHOLE_NUMBER_FORMAT = "#,##0"
PERCENT_NUMBER_FORMAT = "0.##%"
STATUS_HEADER_CANDIDATES = [AUDIT_STATUS_COLUMN, "Status"]
DATE_UPDATED_HEADER_CANDIDATES = ["Date Updated (nego date/review)", "Date Updated"]
PTP_DATE_HEADER = "PTP Date/Due date"
PTP_DATE_HEADER_CANDIDATES = [PTP_DATE_HEADER, "PTP Date", "PTP Due Date", "PTPDUEDATE"]
WRITE_OFF_DATE_HEADER = "WRITE OFF DATE"
WRITE_OFF_DATE_HEADER_CANDIDATES = [WRITE_OFF_DATE_HEADER, "WRITEOFF DATE", "WRITE OFF", "WRITEOFF"]
SOURCE_OF_CONTACT_HEADER = "SOURCE OF CONTACT"
SOURCE_OF_CONTACT_HEADER_CANDIDATES = [SOURCE_OF_CONTACT_HEADER, "Source of Contact"]
PRIN_HEADER = "PRIN"
OB_HEADER = "OB"
FACE_AMOUNT_AUDIT_HEADER = "FACE AMOUNT (OTP/EPA)"
UNDER_NEGO_PAYMENT_AUDIT_STATUSES = {"UNDER NEGO", "REFUSED/NOT AVAILING"}
UNDER_NEGO_STATUS = "UNDER NEGO"
REFUSED_STATUS = "REFUSED/NOT AVAILING"
AVAILED_STATUS = "AVAILED"
DEFAULTED_STATUS = "DEFAULTED"
COMPLIED_STATUS = "COMPLIED"
STATUSES_REQUIRING_RECO_DETAILS = (DEFAULTED_STATUS, AVAILED_STATUS, COMPLIED_STATUS)
STATUS_REQUIRED_RECO_COLUMNS = [WRITE_OFF_DATE_HEADER, "AREA", "AGENCY", "RECO LEVEL", SOURCE_OF_CONTACT_HEADER]
REQUIRED_AGENCY_VALUE = "MADRID"
PARTIAL_PAP_CODE = "PARTIAL"
TPAP_OTP_PAP_CODE = "TPAP_OTP"
BAU_PAP_CODE = "BAU"
BAU_EPA_PAP_CODE = "BAU_EPA"
EPA_PAP_VALUE = "EPA"
OTP_PAP_VALUE = "OTP"
PARTIAL_PAP_VALUE = "PARTIAL"
ACCOUNT_DISPLAY_HEADER = "Customer No."
PRIN_DATABASE_ACCOUNT_HEADER_CANDIDATES = ["account number", "account no", "account no.", "customer no.", "customer number"]
PRIN_DATABASE_PRIN_HEADER_CANDIDATES = [PRIN_HEADER]
PRIN_DATABASE_OB_HEADER_CANDIDATES = [OB_HEADER]
PL_PAYMENT_WEEK_COLUMN_PATTERN = re.compile(r"^Payment Amount .*?\(Week [1-4]\)$", re.IGNORECASE)
OOXML_FILE_SIGNATURE = b"PK\x03\x04"
OLE_FILE_SIGNATURE = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

# ==============================
# HELPERS
# ==============================

def get_current_quarter(current_datetime=None):
    current_datetime = current_datetime or datetime.today()
    return ((current_datetime.month - 1) // 3) + 1


def build_tpap_password(current_datetime=None):
    current_datetime = current_datetime or datetime.today()
    quarter = get_current_quarter(current_datetime)
    return f"MAD_{quarter}Q{current_datetime.year}"


def decrypt_file(uploaded_file, password):
    uploaded_file.seek(0)
    decrypted_file = BytesIO()
    office_file = msoffcrypto.OfficeFile(uploaded_file)
    office_file.load_key(password=password)
    office_file.decrypt(decrypted_file)
    decrypted_file.seek(0)
    return decrypted_file


def build_output_filename(current_datetime=None):
    current_datetime = current_datetime or datetime.today()
    date_text = current_datetime.strftime(OUTPUT_FILE_DATE_FORMAT).upper()
    return OUTPUT_FILE_NAME_TEMPLATE.format(date=date_text)


def encrypt_output_workbook(workbook_buffer, password):
    workbook_buffer.seek(0)
    encrypted_output = BytesIO()
    OOXMLFile(workbook_buffer).encrypt(password, encrypted_output)
    encrypted_output.seek(0)
    return encrypted_output


def format_excel_load_error(error):
    if isinstance(error, BadZipFile):
        return "File is not a valid OpenXML workbook archive."

    if isinstance(error, InvalidFileException):
        error_message = str(error).strip()
        return error_message or "File format is not supported by openpyxl."

    error_message = str(error).strip()
    if error_message:
        return error_message

    return error.__class__.__name__


def inspect_excel_upload(file_name, raw_bytes):
    normalized_name = Path(file_name or "DATABASE FOR TPAP PRIN file").name
    lowered_name = normalized_name.lower()
    header_bytes = raw_bytes[:2048]
    stripped_header_bytes = header_bytes.lstrip().lower()

    if lowered_name.startswith("~$"):
        return {
            "file_name": normalized_name,
            "container": "excel_temp_lock",
            "description": "Excel temporary lock file",
        }

    if not raw_bytes:
        return {
            "file_name": normalized_name,
            "container": "empty",
            "description": "empty file",
        }

    if raw_bytes.startswith(OOXML_FILE_SIGNATURE):
        return {
            "file_name": normalized_name,
            "container": "ooxml_zip",
            "description": "OpenXML workbook",
        }

    if raw_bytes.startswith(OLE_FILE_SIGNATURE):
        return {
            "file_name": normalized_name,
            "container": "ole_compound",
            "description": "legacy or encrypted Excel workbook",
        }

    if stripped_header_bytes.startswith((b"<!doctype html", b"<html")):
        return {
            "file_name": normalized_name,
            "container": "html",
            "description": "HTML document",
        }

    if b"\x00" not in header_bytes:
        has_delimiter = any(delimiter in header_bytes for delimiter in (b",", b";", b"\t"))
        has_line_break = any(line_break in header_bytes for line_break in (b"\r", b"\n"))
        if has_delimiter and has_line_break:
            return {
                "file_name": normalized_name,
                "container": "delimited_text",
                "description": "delimited text file",
            }

    return {
        "file_name": normalized_name,
        "container": "unknown",
        "description": "unrecognized file type",
    }


def build_excel_upload_precheck_error(file_info):
    file_name = file_info["file_name"]
    container = file_info["container"]

    if container == "excel_temp_lock":
        return (
            f'The uploaded file "{file_name}" looks like Excel\'s temporary lock file (~$...), '
            "not the actual workbook. Close the source workbook if it is open, then upload the real file."
        )

    if container == "empty":
        return f'The uploaded file "{file_name}" is empty. Please upload a valid .xls, .xlsx, or .xlsm workbook.'

    if container == "html":
        return (
            f'The uploaded file "{file_name}" is an HTML page, not an Excel workbook. '
            "Please upload a valid .xls, .xlsx, or .xlsm workbook."
        )

    if container == "delimited_text":
        return (
            f'The uploaded file "{file_name}" looks like a text or CSV file, not an Excel workbook. '
            "Please upload a valid .xls, .xlsx, or .xlsm workbook."
        )

    return None


def normalize_key(value):
    """Normalize account/customer numbers so matching is more reliable."""
    if pd.isna(value):
        return None

    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return str(int(value))

    if isinstance(value, int):
        return str(value)

    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"\s+", "", text)
    return text


def build_leading_zero_fallback_lookup(records_by_key):
    fallback_lookup = dict(records_by_key)
    alias_to_source_key = {}
    conflicting_aliases = set()

    for source_key in records_by_key:
        if not source_key or not str(source_key).isdigit():
            continue

        alias_key = str(source_key).lstrip("0") or "0"
        if alias_key == source_key or alias_key in records_by_key:
            continue

        existing_source_key = alias_to_source_key.get(alias_key)
        if existing_source_key is None:
            alias_to_source_key[alias_key] = source_key
        elif existing_source_key != source_key:
            conflicting_aliases.add(alias_key)

    for alias_key in conflicting_aliases:
        alias_to_source_key.pop(alias_key, None)

    for alias_key, source_key in alias_to_source_key.items():
        fallback_lookup[alias_key] = records_by_key[source_key]

    return fallback_lookup


def is_placeholder_blank_text(value):
    return isinstance(value, str) and value.strip() == "-"


def has_meaningful_value(value):
    if pd.isna(value):
        return False
    if is_placeholder_blank_text(value):
        return False
    return str(value).strip() != ""


def is_zero_equivalent_value(value):
    numeric_value = parse_numeric_value(value)
    return numeric_value is not None and numeric_value == 0


def has_discount_rate_charges_value(value):
    if not has_meaningful_value(value):
        return False
    return not is_zero_equivalent_value(value)


def has_required_bau_discount_rate_charge(value):
    if not has_meaningful_value(value):
        return False
    return is_zero_equivalent_value(value)


def is_ignored_discount_principal_value(value):
    if not has_meaningful_value(value):
        return True

    text = str(value).strip()
    if text.upper() == "**FORMULATED":
        return True

    return is_zero_equivalent_value(value)


def has_discount_rate_principal_value(value):
    if is_ignored_discount_principal_value(value):
        return False
    return parse_percentage_value(value) is not None


def parse_numeric_value(value):
    if pd.isna(value):
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    normalized = (
        text.replace("â‚±", "")
        .replace("₱", "")
        .replace(",", "")
        .replace(" ", "")
    )
    if normalized.endswith("%"):
        normalized = normalized[:-1]

    try:
        return float(normalized)
    except ValueError:
        return None


def parse_percentage_value(value):
    numeric_value = parse_numeric_value(value)
    if numeric_value is None:
        return None

    text = "" if pd.isna(value) else str(value).strip()
    if text.endswith("%") or abs(numeric_value) > 1:
        return numeric_value / 100

    return numeric_value


def normalize_currency_value(value):
    numeric_value = parse_numeric_value(value)
    if numeric_value is None:
        return None

    decimal_value = Decimal(str(numeric_value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(decimal_value)


def sanitize_placeholder_values(df):
    return df.apply(lambda column: column.map(lambda value: pd.NA if is_placeholder_blank_text(value) else value))


def normalize_discount_rate_principal_value(value):
    if not has_meaningful_value(value):
        return pd.NA

    text = str(value).strip() if isinstance(value, str) else ""
    if text.upper() == "**FORMULATED":
        return value

    parsed_value = parse_percentage_value(value)
    if parsed_value is None:
        return value

    decimal_value = Decimal(str(parsed_value)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    return float(decimal_value)


def round_up_to_nearest_hundred(value):
    numeric_value = normalize_currency_value(value)
    if numeric_value is None:
        return None

    decimal_value = Decimal(str(numeric_value))
    if decimal_value == 0:
        return 0

    hundred = Decimal("100")
    rounded_value = (decimal_value / hundred).to_integral_value(rounding=ROUND_CEILING) * hundred
    return int(rounded_value)


def round_otp_face_amounts_to_nearest_hundred(df):
    df = df.copy()
    if FACE_AMOUNT_AUDIT_HEADER not in df.columns or "PAP" not in df.columns:
        return df

    pap_series = df["PAP"].fillna("").astype(str).str.strip().str.upper()
    face_amount_values = df[FACE_AMOUNT_AUDIT_HEADER].apply(parse_numeric_value)
    otp_face_amount_mask = pap_series.eq(OTP_PAP_VALUE) & face_amount_values.notna()

    if not otp_face_amount_mask.any():
        return df

    rounded_face_amounts = face_amount_values.loc[otp_face_amount_mask].apply(round_up_to_nearest_hundred)
    df.loc[otp_face_amount_mask, FACE_AMOUNT_AUDIT_HEADER] = rounded_face_amounts

    if "Face Amount" in df.columns:
        df.loc[otp_face_amount_mask, "Face Amount"] = rounded_face_amounts

    return df


def has_payment_value(value):
    numeric_value = parse_numeric_value(value)
    if numeric_value is not None:
        return numeric_value != 0
    return has_meaningful_value(value)


def get_pl_payment_columns(df):
    return [
        column_name
        for column_name in df.columns
        if PL_PAYMENT_WEEK_COLUMN_PATTERN.match(str(column_name).strip())
    ]


def populate_pl_payment_from_weekly_columns(df):
    df = df.copy()
    payment_columns = get_pl_payment_columns(df)
    if not payment_columns:
        return df

    weekly_payments = df[payment_columns].apply(pd.to_numeric, errors="coerce")
    df["Payment"] = weekly_payments.sum(axis=1, min_count=1)
    return df


def apply_under_nego_payment_audit(df, sheet_name):
    df = df.copy()

    audit_status_source = df[AUDIT_STATUS_COLUMN] if AUDIT_STATUS_COLUMN in df.columns else get_audit_status_series(df)
    status_series = audit_status_source.fillna("").astype(str).str.strip().str.upper()
    pap_code_series = df["PAP_Code"].fillna("").astype(str).str.strip().str.upper()
    eligible_status_mask = status_series.eq(UNDER_NEGO_STATUS)

    if sheet_name == "PL":
        payment_columns = get_pl_payment_columns(df)
        if not payment_columns:
            raise ValueError(
                "PL sheet is missing payment amount week columns required for the UNDER NEGO audit check."
            )
        payment_has_value_mask = df[payment_columns].apply(lambda column: column.apply(has_payment_value)).any(axis=1)
    else:
        payment_columns = ["Payment"]
        payment_has_value_mask = df["Payment"].apply(has_payment_value)

    rows_with_payment_mask = eligible_status_mask & payment_has_value_mask

    return df, {
        "type": "under_nego_payment_review",
        "category": "AUDIT",
        "priority": "HIGH",
        "payment_columns": payment_columns,
        "rows_reviewed": int(eligible_status_mask.sum()),
        "rows_with_payment": int(rows_with_payment_mask.sum()),
        "rows_needing_under_nego": 0,
        "partial_face_amount_violations": 0,
        "details": pd.DataFrame(),
    }, pap_code_series.eq(PARTIAL_PAP_CODE)


def populate_face_amount_from_discount_principal(df, exclude_mask=None, header_row=1):
    df = df.copy()
    if exclude_mask is None:
        exclude_mask = pd.Series(False, index=df.index)

    status_series = get_audit_status_series(df).fillna("").astype(str).str.strip().str.upper()
    status_exclude_mask = status_series.isin({REFUSED_STATUS, UNDER_NEGO_STATUS})
    combined_exclude_mask = exclude_mask | status_exclude_mask

    discount_source_mask = ~df["Discount Rate Principal (%)"].apply(is_ignored_discount_principal_value)
    discount_rates = pd.to_numeric(
        df["Discount Rate Principal (%)"].apply(parse_percentage_value),
        errors="coerce",
    )
    principal_values = pd.to_numeric(
        df[PRIN_HEADER].apply(parse_numeric_value),
        errors="coerce",
    )
    current_face_amount_values = pd.to_numeric(
        df[FACE_AMOUNT_AUDIT_HEADER].apply(parse_numeric_value),
        errors="coerce",
    )

    valid_discount_mask = discount_source_mask & discount_rates.notna() & ~combined_exclude_mask
    computed_mask = valid_discount_mask & principal_values.notna()
    computed_values = (discount_rates * principal_values).apply(normalize_currency_value)
    rounded_computed_values = computed_values.apply(round_up_to_nearest_hundred)
    mismatch_mask = computed_mask & (
        current_face_amount_values.isna() | current_face_amount_values.lt(rounded_computed_values)
    )
    mismatch_details = build_audit_detail_table(
        df,
        mismatch_mask,
        ["Customer No.", "PAP_Code", FACE_AMOUNT_AUDIT_HEADER, "Discount Rate Principal (%)", PRIN_HEADER],
        extra_columns={
            "Computed Amount": computed_values[mismatch_mask].tolist(),
            "Minimum FACE AMOUNT (OTP/EPA)": rounded_computed_values[mismatch_mask].tolist(),
        },
        header_row=header_row,
    )

    missing_prin_mask = valid_discount_mask & principal_values.isna()
    missing_prin_details = build_audit_detail_table(
        df,
        missing_prin_mask,
        ["Customer No.", "PAP_Code", "Discount Rate Principal (%)", PRIN_HEADER],
        header_row=header_row,
    )

    return df, {
        "type": "face_amount_check_from_prin",
        "category": "AUDIT",
        "priority": "HIGH",
        "title": '"FACE AMOUNT (OTP/EPA)" must not be lower than rounded-up "Discount Rate Principal (%)" x "PRIN"',
        "source_rows_found": int(valid_discount_mask.sum()),
        "rows_checked": int(computed_mask.sum()),
        "violations_count": int(mismatch_mask.sum()),
        "mismatch_details": mismatch_details,
        "missing_prin_count": int(missing_prin_mask.sum()),
        "details": missing_prin_details,
        "status_skipped_rows": int((discount_source_mask & status_exclude_mask).sum()),
    }


def get_audit_status_series(df):
    if "Status" in df.columns:
        return df["Status"]

    if AUDIT_STATUS_COLUMN in df.columns:
        return df[AUDIT_STATUS_COLUMN]

    return pd.Series("", index=df.index, dtype="object")


def build_audit_detail_table(df, mask, columns, extra_columns=None, header_row=1):
    details = df.loc[mask, columns].copy()
    if details.empty:
        return details

    if "Customer No." in details.columns:
        details.rename(columns={"Customer No.": ACCOUNT_DISPLAY_HEADER}, inplace=True)
    if "PTP Date" in details.columns:
        details.rename(columns={"PTP Date": PTP_DATE_HEADER}, inplace=True)
    if "PAP_Code" in df.columns and "PAP_Code" not in details.columns:
        pap_code_values = df.loc[mask, "PAP_Code"].tolist()
        insert_at = 1 if ACCOUNT_DISPLAY_HEADER not in details.columns else details.columns.get_loc(ACCOUNT_DISPLAY_HEADER) + 1
        details.insert(insert_at, "PAP_Code", pap_code_values)

    if extra_columns:
        for column_name, value in extra_columns.items():
            details[column_name] = value

    details.insert(0, "Excel Row", details.index + header_row + 1)
    details.reset_index(drop=True, inplace=True)
    return details


def build_finding_detail_table(df, mask, columns, finding_builder, extra_columns=None, header_row=1):
    details = build_audit_detail_table(df, mask, columns, extra_columns=extra_columns, header_row=header_row)
    if details.empty:
        return details

    findings = [finding_builder(row_index) for row_index in df.index[mask]]
    details["Finding"] = findings
    return details


def dataframe_to_copyable_text(df):
    if df is None or df.empty:
        return ""

    export_df = df.copy()
    export_df = export_df.where(pd.notna(export_df), "")
    return export_df.to_csv(index=False, lineterminator="\n")


def build_copyable_audit_report(sheet_name, audit_results):
    report_sections = [
        f"AUDIT ERRORS - {sheet_name}",
        f"Generated: {datetime.today().strftime('%m/%d/%Y %I:%M:%S %p')}",
    ]
    has_errors = False

    for audit in audit_results:
        title = audit.get("title", audit.get("type", "Audit Rule"))
        section_lines = []

        if audit["type"] == "under_nego_payment_review":
            if audit.get("rows_needing_under_nego", 0) > 0:
                section_lines.append(
                    f'- {audit["rows_needing_under_nego"]} row(s) should have status "{UNDER_NEGO_STATUS}".'
                )
            if audit.get("partial_face_amount_violations", 0) > 0:
                section_lines.append(
                    f'- {audit["partial_face_amount_violations"]} row(s) with PAP_Code "{PARTIAL_PAP_CODE}" should have blank "{FACE_AMOUNT_AUDIT_HEADER}".'
                )
            if not audit.get("details", pd.DataFrame()).empty:
                section_lines.append(dataframe_to_copyable_text(audit["details"]).strip())
        elif audit["type"] == "partial_requires_under_nego_and_blank_due_date":
            rows_status_updated = audit.get("rows_status_updated", 0)
            rows_due_date_cleared = audit.get("rows_due_date_cleared", 0)
            if rows_status_updated > 0 or rows_due_date_cleared > 0:
                section_lines.append(
                    f'- {rows_status_updated} row(s) are not "{UNDER_NEGO_STATUS}" and {rows_due_date_cleared} row(s) have a value in "{PTP_DATE_HEADER}".'
                )
                if not audit.get("details", pd.DataFrame()).empty:
                    section_lines.append(dataframe_to_copyable_text(audit["details"]).strip())
        elif audit["type"] == "face_amount_check_from_prin":
            if audit.get("violations_count", 0) > 0:
                section_lines.append(
                    f'- {audit["violations_count"]} row(s) have "{FACE_AMOUNT_AUDIT_HEADER}" lower than '
                    f'"Discount Rate Principal (%)" x "{PRIN_HEADER}", rounded up to the nearest hundred.'
                )
                if not audit.get("mismatch_details", pd.DataFrame()).empty:
                    section_lines.append(dataframe_to_copyable_text(audit["mismatch_details"]).strip())
            if audit.get("missing_prin_count", 0) > 0:
                section_lines.append(
                    f'- {audit["missing_prin_count"]} row(s) have "Discount Rate Principal (%)" but missing or invalid "{PRIN_HEADER}".'
                )
                if not audit.get("details", pd.DataFrame()).empty:
                    section_lines.append(dataframe_to_copyable_text(audit["details"]).strip())
        elif audit["type"] == "prin_mismatch_vs_database":
            if audit.get("violations_count", 0) > 0:
                section_lines.append(
                    f'- {audit["violations_count"]} row(s) have "{PRIN_HEADER}" and/or "{OB_HEADER}" that do not match DATABASE FOR TPAP PRIN.'
                )
                if audit.get("prin_mismatch_count", 0) > 0:
                    section_lines.append(
                        f'- {audit["prin_mismatch_count"]} row(s) have a different "{PRIN_HEADER}" from the database.'
                    )
                if audit.get("ob_mismatch_count", 0) > 0:
                    section_lines.append(
                        f'- {audit["ob_mismatch_count"]} row(s) have a different "{OB_HEADER}" from the database.'
                    )
                if audit.get("missing_database_count", 0) > 0:
                    section_lines.append(
                        f'- {audit["missing_database_count"]} row(s) were not found in DATABASE FOR TPAP PRIN.'
                    )
                if not audit.get("details", pd.DataFrame()).empty:
                    section_lines.append(dataframe_to_copyable_text(audit["details"]).strip())
        elif audit["type"] == "status_requires_reco_details":
            if audit.get("violations_count", 0) > 0:
                section_lines.append(
                    f'- {audit["violations_count"]} row(s) with status "{DEFAULTED_STATUS}", "{AVAILED_STATUS}", '
                    f'or "{COMPLIED_STATUS}" are missing one or more required values in '
                    '"WRITE OFF DATE", "AREA", "AGENCY", "RECO LEVEL", or "SOURCE OF CONTACT".'
                )
                if not audit.get("details", pd.DataFrame()).empty:
                    section_lines.append(dataframe_to_copyable_text(audit["details"]).strip())
        elif audit["type"] == "agency_must_be_madrid":
            if audit.get("violations_count", 0) > 0:
                section_lines.append(
                    f'- {audit["violations_count"]} row(s) have "AGENCY" not equal to "{REQUIRED_AGENCY_VALUE}".'
                )
                if not audit.get("details", pd.DataFrame()).empty:
                    section_lines.append(dataframe_to_copyable_text(audit["details"]).strip())
        elif audit["type"] == "otp_face_amount_rounded_to_hundred":
            if audit.get("violations_count", 0) > 0:
                section_lines.append(
                    f'- {audit["violations_count"]} row(s) with PAP "{OTP_PAP_VALUE}" have '
                    f'"{FACE_AMOUNT_AUDIT_HEADER}" not rounded up to the nearest hundred.'
                )
                if not audit.get("details", pd.DataFrame()).empty:
                    section_lines.append(dataframe_to_copyable_text(audit["details"]).strip())
            if audit.get("skipped_complied_equal_rows", 0) > 0:
                section_lines.append(
                    f'- {audit["skipped_complied_equal_rows"]} row(s) were skipped because status is "{COMPLIED_STATUS}" '
                    f'and Payment is equal to "{FACE_AMOUNT_AUDIT_HEADER}".'
                )
        else:
            violations_count = audit.get("violations_count", 0)
            if violations_count > 0:
                section_lines.append(f"- {violations_count} row(s) failed this rule.")
                if not audit.get("details", pd.DataFrame()).empty:
                    section_lines.append(dataframe_to_copyable_text(audit["details"]).strip())

        if section_lines:
            has_errors = True
            report_sections.append("")
            report_sections.append(f"RULE: {title}")
            report_sections.extend(section_lines)

    if not has_errors:
        report_sections.append("")
        report_sections.append("No audit errors found.")

    return "\n".join(report_sections).strip()


def normalize_export_dataframe(df):
    if df is None or df.empty:
        return pd.DataFrame()

    export_df = df.copy()
    for column_name in export_df.columns:
        export_df[column_name] = export_df[column_name].apply(
            lambda value: "" if pd.isna(value) else format_display_value(value)
        )
    return export_df


def format_finding_value(value):
    if pd.isna(value):
        return "blank"

    formatted_value = format_display_value(value)
    text = str(formatted_value).strip()
    return text if text else "blank"


def build_field_value_list(detail_row, field_names):
    field_values = []
    for field_name in field_names:
        normalized_field_name = str(field_name).strip()
        if not normalized_field_name:
            continue

        field_value = format_finding_value(get_detail_value(detail_row, normalized_field_name))
        if field_value == "blank":
            continue

        field_values.append(f'"{normalized_field_name}"={field_value}')

    return field_values


def build_required_field_messages(detail_row):
    messages = []

    for column_name in detail_row.index:
        normalized_column_name = str(column_name).strip()
        if not normalized_column_name.startswith("Required ") or normalized_column_name == "Required Format":
            continue

        required_field_name = normalized_column_name.replace("Required ", "", 1)
        required_value = format_finding_value(get_detail_value(detail_row, normalized_column_name))
        current_value = format_finding_value(get_detail_value(detail_row, required_field_name))

        if required_value == "blank":
            if current_value != "blank":
                messages.append(
                    f'"{required_field_name}" must be blank but current value is {current_value}'
                )
            continue

        messages.append(
            f'"{required_field_name}" must be {required_value} but current value is {current_value}'
        )

    return messages


def get_audit_violation_count(audit):
    if "violations_count" in audit:
        return int(audit.get("violations_count", 0) or 0)

    if audit.get("type") == "under_nego_payment_review":
        return int(audit.get("rows_needing_under_nego", 0) or 0) + int(
            audit.get("partial_face_amount_violations", 0) or 0
        )

    if audit.get("type") == "partial_requires_under_nego_and_blank_due_date":
        return int(audit.get("rows_status_updated", 0) or 0) + int(
            audit.get("rows_due_date_cleared", 0) or 0
        )

    return 0


def get_audit_detail_exports(audit):
    detail_exports = []

    if audit.get("type") == "face_amount_check_from_prin":
        mismatch_details = audit.get("mismatch_details", pd.DataFrame())
        if isinstance(mismatch_details, pd.DataFrame) and not mismatch_details.empty:
            detail_exports.append(("Mismatch Details", mismatch_details))

        missing_prin_details = audit.get("details", pd.DataFrame())
        if isinstance(missing_prin_details, pd.DataFrame) and not missing_prin_details.empty:
            detail_exports.append(("Missing PRIN", missing_prin_details))
        return detail_exports

    details = audit.get("details", pd.DataFrame())
    if isinstance(details, pd.DataFrame) and not details.empty:
        detail_exports.append(("Details", details))

    return detail_exports


def get_detail_value(detail_row, column_name):
    value = detail_row.get(column_name, "")
    return "" if pd.isna(value) else value


def build_audit_finding_text(audit, detail_label, detail_row):
    audit_type = audit.get("type", "")
    finding_parts = []
    direct_finding = str(get_detail_value(detail_row, "Finding")).strip()
    current_status = format_finding_value(get_detail_value(detail_row, "Status"))
    current_pap = format_finding_value(get_detail_value(detail_row, "PAP"))
    current_pap_code = format_finding_value(get_detail_value(detail_row, "PAP_Code"))
    required_field_messages = build_required_field_messages(detail_row)
    handled_required_field_message_types = {
        "epa_pap_requires_bau_epa_code",
        "partial_pap_requires_partial_code_and_under_nego",
        "partial_requires_under_nego_and_blank_due_date",
        "otp_face_amount_rounded_to_hundred",
    }

    if audit_type in {"refused_status_field_rules", "under_nego_status_field_rules"} and direct_finding:
        invalid_fields = [field.strip() for field in direct_finding.split(",") if field.strip()]
        invalid_field_values = build_field_value_list(detail_row, invalid_fields)
        if invalid_field_values:
            finding_parts.append(
                f'Status is "{current_status}" but these fields should be blank: {", ".join(invalid_field_values)}'
            )
        else:
            finding_parts.append(f'Status is "{current_status}" but fields that should be blank have values.')
    elif audit_type == "status_requires_reco_details":
        missing_fields = str(get_detail_value(detail_row, "Missing Fields")).strip()
        current_field_values = build_field_value_list(detail_row, STATUS_REQUIRED_RECO_COLUMNS)
        if current_field_values:
            finding_parts.append(
                f'Status is "{current_status}" so these required fields are blank: {missing_fields}; '
                f'Other related values on this row: {", ".join(current_field_values)}'
            )
        else:
            finding_parts.append(
                f'Status is "{current_status}" so these required fields are blank: {missing_fields}'
            )
    elif audit_type == "agency_must_be_madrid":
        finding_parts.append(
            f'"AGENCY" must be "{REQUIRED_AGENCY_VALUE}" but current value is '
            f'{format_finding_value(get_detail_value(detail_row, "AGENCY"))}'
        )
    elif audit_type == "epa_pap_requires_bau_epa_code":
        finding_parts.append(
            f'PAP is "{current_pap}" so "PAP_Code" must be "{BAU_EPA_PAP_CODE}", '
            f'but current value is "{current_pap_code}"'
        )
    elif audit_type == "partial_pap_requires_partial_code_and_under_nego":
        row_issues = []
        if current_pap_code != PARTIAL_PAP_CODE:
            row_issues.append(f'"PAP_Code" must be "{PARTIAL_PAP_CODE}" but current value is "{current_pap_code}"')
        if current_status != UNDER_NEGO_STATUS:
            row_issues.append(f'"Status" must be "{UNDER_NEGO_STATUS}" but current value is "{current_status}"')
        if row_issues:
            finding_parts.append("; ".join(row_issues))
        else:
            finding_parts.append(
                f'PAP is "{current_pap}" and this row does not satisfy the required PAP_Code/Status rule.'
            )
    elif audit_type == "partial_requires_under_nego_and_blank_due_date":
        if required_field_messages:
            finding_parts.append("; ".join(required_field_messages))
        else:
            finding_parts.append(
                f'PAP_Code is "{current_pap_code}" and this row must be "{UNDER_NEGO_STATUS}" with blank "{PTP_DATE_HEADER}".'
            )
    elif audit_type == "otp_face_amount_rounded_to_hundred":
        finding_parts.append(
            f'PAP is "{current_pap}" and {FACE_AMOUNT_AUDIT_HEADER} must be rounded up to the nearest hundred; '
            f'current value is {format_finding_value(get_detail_value(detail_row, FACE_AMOUNT_AUDIT_HEADER))}; '
            f'required value is {format_finding_value(get_detail_value(detail_row, f"Required {FACE_AMOUNT_AUDIT_HEADER}"))}'
        )
    elif direct_finding:
        finding_parts.append(direct_finding)

    if audit_type not in handled_required_field_message_types:
        finding_parts.extend(required_field_messages)

    missing_fields = str(get_detail_value(detail_row, "Missing Fields")).strip()
    if missing_fields and audit_type not in {"status_requires_reco_details"}:
        missing_field_names = [field.strip() for field in missing_fields.split(",") if field.strip()]
        current_field_values = build_field_value_list(detail_row, missing_field_names)
        if current_field_values:
            finding_parts.append(
                f"Missing required field(s): {missing_fields}; Other related values on this row: {', '.join(current_field_values)}"
            )
        else:
            finding_parts.append(f"Missing required field(s): {missing_fields}")

    required_format = str(get_detail_value(detail_row, "Required Format")).strip()
    if required_format:
        finding_parts.append(
            f'Current Number Format: "{format_finding_value(get_detail_value(detail_row, "Number Format"))}"; '
            f'Required Number Format: "{required_format}"'
        )

    if audit_type == "face_amount_check_from_prin":
        if detail_label == "Mismatch Details":
            finding_parts.append(
                f'Current {FACE_AMOUNT_AUDIT_HEADER}: {format_finding_value(get_detail_value(detail_row, FACE_AMOUNT_AUDIT_HEADER))}; '
                f'"Discount Rate Principal (%)": {format_finding_value(get_detail_value(detail_row, "Discount Rate Principal (%)"))}; '
                f'"{PRIN_HEADER}": {format_finding_value(get_detail_value(detail_row, PRIN_HEADER))}; '
                f'Computed Amount: {format_finding_value(get_detail_value(detail_row, "Computed Amount"))}; '
                f'Minimum Required {FACE_AMOUNT_AUDIT_HEADER}: {format_finding_value(get_detail_value(detail_row, f"Minimum {FACE_AMOUNT_AUDIT_HEADER}"))}'
            )
        elif detail_label == "Missing PRIN":
            finding_parts.append(
                f'"Discount Rate Principal (%)" has value {format_finding_value(get_detail_value(detail_row, "Discount Rate Principal (%)"))} '
                f'but "{PRIN_HEADER}" is {format_finding_value(get_detail_value(detail_row, PRIN_HEADER))}'
            )

    elif audit_type == "blank_pap_or_pap_code_sets_refused":
        blank_fields = [
            field_name
            for field_name in ("PAP", "PAP_Code")
            if format_finding_value(get_detail_value(detail_row, field_name)) == "blank"
        ]
        if blank_fields:
            verb = "is" if len(blank_fields) == 1 else "are"
            finding_parts.append(
                f'{" and ".join(f"""\"{field_name}\"""" for field_name in blank_fields)} {verb} blank on this row'
            )

    elif audit_type == "otp_pap_cannot_have_bau_epa_code":
        finding_parts.append(
            f'PAP is "{current_pap}" but PAP_Code is "{current_pap_code}"; "{BAU_EPA_PAP_CODE}" is not allowed for OTP rows'
        )

    elif audit_type == "tpap_otp_refused_status_invalid":
        finding_parts.append(
            f'PAP_Code is "{current_pap_code}" but Status is "{current_status}"; "{REFUSED_STATUS}" is not allowed for "{TPAP_OTP_PAP_CODE}" rows'
        )

    elif audit_type == "discount_rate_principal_required_for_tpap":
        finding_parts.append(
            f'PAP_Code is "{current_pap_code}" and "Discount Rate Principal (%)" is {format_finding_value(get_detail_value(detail_row, "Discount Rate Principal (%)"))}; a value is required for TPAP rows'
        )

    elif audit_type == "discount_rate_charges_blank_for_tpap":
        finding_parts.append(
            f'PAP_Code is "{current_pap_code}" and "Discount Rate Charges (%)" must be blank, but current value is {format_finding_value(get_detail_value(detail_row, "Discount Rate Charges (%)"))}'
        )

    elif audit_type == "discount_rate_charges_zero_for_bau":
        finding_parts.append(
            f'PAP_Code is "{current_pap_code}" and "Discount Rate Charges (%)" must be 0% (or 0), but current value is {format_finding_value(get_detail_value(detail_row, "Discount Rate Charges (%)"))}'
        )

    elif audit_type == "program_dates_required_for_tpap":
        missing_fields = str(get_detail_value(detail_row, "Missing Fields")).strip()
        if missing_fields:
            finding_parts.append(f'TPAP row is missing: {missing_fields}')
        finding_parts.append(
            f'Current "Program Start Date": {format_finding_value(get_detail_value(detail_row, "Program Start Date"))}; '
            f'Current "Program End Date": {format_finding_value(get_detail_value(detail_row, "Program End Date"))}'
        )

    elif audit_type == "availed_status_not_past_due":
        finding_parts.append(
            f'Status is "{current_status}" but "{PTP_DATE_HEADER}" is {format_finding_value(get_detail_value(detail_row, PTP_DATE_HEADER))}, which is already past due'
        )

    elif audit_type == "defaulted_status_rules":
        finding_parts.append(
            f'Current Payment: {format_finding_value(get_detail_value(detail_row, "Payment"))}; '
            f'Current {FACE_AMOUNT_AUDIT_HEADER}: {format_finding_value(get_detail_value(detail_row, FACE_AMOUNT_AUDIT_HEADER))}; '
            f'Current "{PTP_DATE_HEADER}": {format_finding_value(get_detail_value(detail_row, PTP_DATE_HEADER))}'
        )

    elif audit_type == "complied_status_rules":
        finding_parts.append(
            f'Current Payment: {format_finding_value(get_detail_value(detail_row, "Payment"))}; '
            f'Current {FACE_AMOUNT_AUDIT_HEADER}: {format_finding_value(get_detail_value(detail_row, FACE_AMOUNT_AUDIT_HEADER))}'
        )

    elif audit_type == "date_format_mmddyyyy":
        finding_parts.append(
            f'Column "{get_detail_value(detail_row, "Column")}" has value {format_finding_value(get_detail_value(detail_row, "Value"))}; expected a valid date entry or Excel date using MM/DD/YYYY'
        )

    elif audit_type == "prin_centavo_format":
        finding_parts.append(
            f'"{PRIN_HEADER}" has decimal value {format_finding_value(get_detail_value(detail_row, "Value"))} but is not displayed with 2 decimal places'
        )

    column_name = str(get_detail_value(detail_row, "Column")).strip()
    if column_name:
        finding_parts.append(
            f'Column "{column_name}" current value: {format_finding_value(get_detail_value(detail_row, "Value"))}'
        )

    cleaned_parts = []
    seen_parts = set()
    for part in finding_parts:
        normalized_part = str(part).strip()
        if not normalized_part or normalized_part in seen_parts:
            continue
        seen_parts.add(normalized_part)
        cleaned_parts.append(normalized_part)

    if cleaned_parts:
        return "; ".join(cleaned_parts)

    return audit.get("title", audit.get("type", "Audit issue found"))


def build_audit_findings_dataframe(sheet_name, audit_results):
    findings_rows = []

    for audit in audit_results:
        detail_exports = get_audit_detail_exports(audit)
        if detail_exports:
            for detail_label, detail_df in detail_exports:
                export_df = normalize_export_dataframe(detail_df)
                for _, detail_row in export_df.iterrows():
                    customer_number = normalize_key(get_detail_value(detail_row, ACCOUNT_DISPLAY_HEADER)) or ""
                    findings_rows.append(
                        {
                            "Sheet": sheet_name,
                            "Audit Category": audit.get("title", audit.get("type", "")),
                            "Customer No.": customer_number,
                            "Excel Row": get_detail_value(detail_row, "Excel Row"),
                            "Findings": build_audit_finding_text(audit, detail_label, detail_row),
                        }
                    )
            continue

        if get_audit_violation_count(audit) > 0:
            findings_rows.append(
                {
                    "Sheet": sheet_name,
                    "Audit Category": audit.get("title", audit.get("type", "")),
                    "Customer No.": "",
                    "Excel Row": "",
                    "Findings": audit.get("title", audit.get("type", "Audit issue found")),
                }
            )

    findings_df = pd.DataFrame(findings_rows)
    if findings_df.empty:
        findings_df = pd.DataFrame(
            columns=["Sheet", "Audit Category", "Customer No.", "Excel Row", "Findings"]
        )

    return findings_df


def autosize_worksheet_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))

        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)


def apply_audit_findings_worksheet_formatting(worksheet):
    header_map = build_header_map(worksheet, header_row=1)
    customer_no_col = header_map.get("Customer No.")
    excel_row_col = header_map.get("Excel Row")

    if excel_row_col is not None:
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=excel_row_col)
            numeric_value = parse_numeric_value(cell.value)
            if numeric_value is None:
                continue
            cell.value = int(numeric_value)
            cell.number_format = "0"

    if customer_no_col is None:
        return

    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=customer_no_col)
        customer_number = normalize_key(cell.value)
        if not customer_number or not customer_number.isdigit():
            continue

        if len(customer_number) > 15:
            cell.value = customer_number
            continue

        cell.value = int(customer_number)
        cell.number_format = "0" * len(customer_number)


def build_audit_excel_workbook(sheet_audits):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for source_sheet_name, audit_results in sheet_audits.items():
            findings_df = build_audit_findings_dataframe(source_sheet_name, audit_results)
            findings_df.to_excel(
                writer,
                index=False,
                sheet_name=f"{source_sheet_name} Audit Findings"[:31],
            )

        for worksheet in writer.book.worksheets:
            apply_audit_findings_worksheet_formatting(worksheet)
            autosize_worksheet_columns(worksheet)

    output.seek(0)
    return output


def build_audit_template_workbook(sheet_audits):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for source_sheet_name, audit_results in sheet_audits.items():
            template_rows = []
            for audit in audit_results:
                template_rows.append(
                    {
                        "Sheet": source_sheet_name,
                        "Audit Category": audit.get("title", audit.get("type", "")),
                        "Customer No.": "",
                        "Excel Row": "",
                        "Findings": "",
                        "Reviewer Remarks": "",
                        "Supervisor Comments": "",
                    }
                )

            template_df = pd.DataFrame(template_rows)
            if template_df.empty:
                template_df = pd.DataFrame(
                    columns=[
                        "Sheet",
                        "Audit Category",
                        "Customer No.",
                        "Excel Row",
                        "Findings",
                        "Reviewer Remarks",
                        "Supervisor Comments",
                    ]
                )

            template_df.to_excel(writer, index=False, sheet_name=f"{source_sheet_name} Audit Template"[:31])

        for worksheet in writer.book.worksheets:
            autosize_worksheet_columns(worksheet)

    output.seek(0)
    return output


def normalize_excel_number_format(number_format):
    return str(number_format or "").strip().upper().replace("\\", "")


def is_mmddyyyy_number_format(number_format):
    normalized = normalize_excel_number_format(number_format)
    if not normalized:
        return False

    # Ignore Excel locale markers and formatting sections so common date formats
    # like `m/d/yyyy`, `[$-409]mm/dd/yyyy;@`, or `mm/dd/yyyy h:mm` still pass.
    normalized = normalized.split(";", 1)[0]
    normalized = re.sub(r"\[\$-[^\]]+\]", "", normalized)
    normalized = re.sub(r'"[^"]*"', "", normalized)
    normalized = normalized.replace("_", "").replace("*", "").replace(" ", "")
    normalized_lower = normalized.lower()

    return bool(re.search(r"m{1,4}[-/]d{1,4}[-/]y{4}", normalized_lower))


def is_valid_mmddyyyy_date_text(text):
    stripped_text = str(text).strip()
    if not stripped_text:
        return False

    # Accept common user-entered text dates like 3/1/2026 or 03/01/2026.
    return bool(re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", stripped_text))


def get_number_format_decimal_places(number_format):
    normalized = normalize_excel_number_format(number_format)
    if "." not in normalized:
        return 0

    decimal_part = normalized.split(".", 1)[1]
    count = 0
    for char in decimal_part:
        if char in {"0", "#"}:
            count += 1
            continue
        break
    return count


def get_value_decimal_places(value):
    if pd.isna(value) or isinstance(value, bool):
        return 0

    if isinstance(value, int):
        return 0

    if isinstance(value, float):
        if math.isnan(value):
            return 0
        decimal_value = Decimal(str(value))
        return max(0, -decimal_value.as_tuple().exponent)

    if isinstance(value, Decimal):
        return max(0, -value.as_tuple().exponent)

    text = str(value).strip()
    if not text:
        return 0

    normalized = (
        text.replace("Ã¢â€šÂ±", "")
        .replace("â‚±", "")
        .replace("$", "")
        .replace(",", "")
        .replace(" ", "")
    )
    if normalized.endswith("%"):
        normalized = normalized[:-1]

    try:
        decimal_value = Decimal(normalized)
    except InvalidOperation:
        return 0

    return max(0, -decimal_value.as_tuple().exponent)


def format_display_value(value):
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return pd.Timestamp(value).strftime("%m/%d/%Y")
    return value


def get_first_existing_column(df, header_candidates):
    return find_matching_column_name(df.columns, header_candidates)


def find_matching_column_name(columns, header_candidates):
    normalized_lookup = {
        re.sub(r"\s+", " ", str(column_name).strip()).casefold(): column_name
        for column_name in columns
    }
    for header in header_candidates:
        normalized_header = re.sub(r"\s+", " ", str(header).strip()).casefold()
        if normalized_header in normalized_lookup:
            return normalized_lookup[normalized_header]
    return None


def ensure_canonical_column(df, canonical_name, header_candidates, default_value=pd.NA):
    source_col = get_first_existing_column(df, header_candidates)

    if source_col is not None:
        if canonical_name != source_col or canonical_name not in df.columns:
            df[canonical_name] = df[source_col]
    elif canonical_name not in df.columns:
        df[canonical_name] = default_value

    return df


def coerce_to_excel_datetime(value):
    if pd.isna(value) or is_placeholder_blank_text(value):
        return None

    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.to_pydatetime()

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
        except Exception:
            return None

        if isinstance(converted, datetime):
            return converted
        if isinstance(converted, date):
            return datetime.combine(converted, datetime.min.time())
        return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None

        explicit_formats = (
            "%m/%d/%Y",
            "%m/%d/%y",
            "%d/%m/%Y",
            "%d/%m/%y",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%m-%d-%Y",
            "%d-%m-%Y",
            "%b %d %Y",
            "%B %d %Y",
            "%d %b %Y",
            "%d %B %Y",
        )
        for date_format in explicit_formats:
            try:
                return datetime.strptime(text, date_format)
            except ValueError:
                continue

        for dayfirst in (False, True):
            parsed = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
            if pd.notna(parsed):
                return parsed.to_pydatetime()

    return None


def normalize_date_cell_value(value):
    if not has_meaningful_value(value):
        return pd.NaT

    parsed_value = coerce_to_excel_datetime(value)
    if parsed_value is None:
        return value

    return pd.Timestamp(parsed_value).normalize()


def normalize_tpap_dataframe(df):
    df = sanitize_placeholder_values(df.copy())

    date_columns = [
        "Date Updated",
        "PTP Date",
        "Program Start Date",
        "Program End Date",
        WRITE_OFF_DATE_HEADER,
    ]
    for column_name in date_columns:
        if column_name in df.columns:
            df[column_name] = df[column_name].apply(normalize_date_cell_value)

    if "Discount Rate Principal (%)" in df.columns:
        df["Discount Rate Principal (%)"] = df["Discount Rate Principal (%)"].apply(
            normalize_discount_rate_principal_value
        )

    return df


def get_decimal_places(value):
    try:
        decimal_value = Decimal(str(value)).normalize()
    except (InvalidOperation, ValueError, TypeError):
        return 0

    return max(0, -decimal_value.as_tuple().exponent)


def get_payment_number_format(value, source_number_format=None):
    normalized_source_format = str(source_number_format).strip() if source_number_format else ""
    numeric_value = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric_value):
        return normalized_source_format or None

    decimal_places = get_decimal_places(numeric_value)
    comma_number_format = "#,##0"
    if decimal_places > 0:
        comma_number_format += "." + ("0" * min(decimal_places, 6))

    if normalized_source_format and normalized_source_format.lower() != "general":
        if "," in normalized_source_format:
            return normalized_source_format
        return comma_number_format

    return comma_number_format


def build_bpi_payment_number_format_lookup(bpi_ws):
    header_map = build_header_map(bpi_ws, header_row=1)
    if "Account No" not in header_map:
        raise ValueError("Column 'Account No' not found in BPI Payment Extraction sheet.")

    account_col = header_map["Account No"]
    payment_col = 4
    number_format_lookup = {}

    for row_idx in range(2, bpi_ws.max_row + 1):
        account_key = normalize_key(bpi_ws.cell(row_idx, account_col).value)
        payment_cell = bpi_ws.cell(row_idx, payment_col)
        payment_value = pd.to_numeric(payment_cell.value, errors="coerce")

        if account_key is None or pd.isna(payment_value):
            continue

        number_format_lookup[account_key] = payment_cell.number_format

    return number_format_lookup


def get_prin_database_required_columns(db_headers):
    account_column = find_matching_column_name(db_headers, PRIN_DATABASE_ACCOUNT_HEADER_CANDIDATES)
    prin_column = find_matching_column_name(db_headers, PRIN_DATABASE_PRIN_HEADER_CANDIDATES)
    ob_column = find_matching_column_name(db_headers, PRIN_DATABASE_OB_HEADER_CANDIDATES)

    missing_columns = []
    if account_column is None:
        missing_columns.append('"account number"')
    if prin_column is None:
        missing_columns.append(f'"{PRIN_HEADER}"')
    if ob_column is None:
        missing_columns.append(f'"{OB_HEADER}"')
    if missing_columns:
        raise ValueError(
            "DATABASE FOR TPAP PRIN sheet \"DB\" is missing required column(s): "
            + ", ".join(missing_columns)
        )

    return account_column, prin_column, ob_column


def load_prin_database_from_openpyxl(candidate_bytes):
    db_workbook = openpyxl.load_workbook(BytesIO(candidate_bytes), data_only=True, read_only=True)
    try:
        if PRIN_DATABASE_SHEET_NAME not in db_workbook.sheetnames:
            raise ValueError(
                f'Sheet "{PRIN_DATABASE_SHEET_NAME}" not found in DATABASE FOR TPAP PRIN file.'
            )

        db_ws = db_workbook[PRIN_DATABASE_SHEET_NAME]
        header_row = next(db_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ValueError('DATABASE FOR TPAP PRIN sheet "DB" is empty.')

        db_headers = ["" if header is None else str(header).strip() for header in header_row]
        account_column, prin_column, ob_column = get_prin_database_required_columns(db_headers)

        column_indices = {
            header_name: db_headers.index(header_name)
            for header_name in {account_column, prin_column, ob_column}
        }
        min_col_index = min(column_indices.values())
        max_col_index = max(column_indices.values())
        prin_lookup = {}

        for row_values in db_ws.iter_rows(
            min_row=2,
            min_col=min_col_index + 1,
            max_col=max_col_index + 1,
            values_only=True,
        ):
            if row_values is None:
                continue

            account_offset = column_indices[account_column] - min_col_index
            prin_offset = column_indices[prin_column] - min_col_index
            ob_offset = column_indices[ob_column] - min_col_index

            account_key = normalize_key(row_values[account_offset] if account_offset < len(row_values) else None)
            if not account_key:
                continue

            prin_lookup[account_key] = {
                PRIN_HEADER: row_values[prin_offset] if prin_offset < len(row_values) else None,
                OB_HEADER: row_values[ob_offset] if ob_offset < len(row_values) else None,
            }
    finally:
        db_workbook.close()

    return build_leading_zero_fallback_lookup(prin_lookup)


def load_prin_database_from_xlrd(candidate_bytes):
    legacy_workbook = xlrd.open_workbook(file_contents=candidate_bytes, on_demand=True)
    try:
        if PRIN_DATABASE_SHEET_NAME not in legacy_workbook.sheet_names():
            raise ValueError(
                f'Sheet "{PRIN_DATABASE_SHEET_NAME}" not found in DATABASE FOR TPAP PRIN file.'
            )

        legacy_sheet = legacy_workbook.sheet_by_name(PRIN_DATABASE_SHEET_NAME)
        if legacy_sheet.nrows == 0:
            raise ValueError('DATABASE FOR TPAP PRIN sheet "DB" is empty.')

        db_headers = ["" if header is None else str(header).strip() for header in legacy_sheet.row_values(0)]
        account_column, prin_column, ob_column = get_prin_database_required_columns(db_headers)

        column_indices = {
            header_name: db_headers.index(header_name)
            for header_name in {account_column, prin_column, ob_column}
        }
        min_col_index = min(column_indices.values())
        max_col_index = max(column_indices.values())
        prin_lookup = {}

        for row_index in range(1, legacy_sheet.nrows):
            row_values = legacy_sheet.row_values(
                row_index,
                start_colx=min_col_index,
                end_colx=max_col_index + 1,
            )

            account_offset = column_indices[account_column] - min_col_index
            prin_offset = column_indices[prin_column] - min_col_index
            ob_offset = column_indices[ob_column] - min_col_index

            account_key = normalize_key(row_values[account_offset] if account_offset < len(row_values) else None)
            if not account_key:
                continue

            prin_lookup[account_key] = {
                PRIN_HEADER: row_values[prin_offset] if prin_offset < len(row_values) else None,
                OB_HEADER: row_values[ob_offset] if ob_offset < len(row_values) else None,
            }
    finally:
        legacy_workbook.release_resources()

    return build_leading_zero_fallback_lookup(prin_lookup)


def load_prin_database_lookup(database_file):
    file_name = getattr(database_file, "name", "DATABASE FOR TPAP PRIN file")
    database_file.seek(0)
    raw_bytes = database_file.read()
    file_info = inspect_excel_upload(file_name, raw_bytes)
    precheck_error = build_excel_upload_precheck_error(file_info)
    if precheck_error:
        raise ValueError(precheck_error)

    prin_lookup = None
    load_errors = []
    decrypted_bytes = None
    decrypt_error = None

    try:
        decrypted_file = decrypt_file(BytesIO(raw_bytes), PRIN_DATABASE_PASSWORD)
        decrypted_bytes = decrypted_file.getvalue()
    except Exception as error:
        decrypt_error = error

    workbook_attempts = []
    container = file_info["container"]
    if container == "ooxml_zip":
        workbook_attempts.append(("OpenXML direct", load_prin_database_from_openpyxl, raw_bytes))
        if decrypted_bytes is not None:
            workbook_attempts.append(("OpenXML password", load_prin_database_from_openpyxl, decrypted_bytes))
        workbook_attempts.append(("Legacy Excel direct", load_prin_database_from_xlrd, raw_bytes))
        if decrypted_bytes is not None:
            workbook_attempts.append(("Legacy Excel password", load_prin_database_from_xlrd, decrypted_bytes))
    elif container == "ole_compound":
        if decrypted_bytes is not None:
            workbook_attempts.append(("OpenXML password", load_prin_database_from_openpyxl, decrypted_bytes))
        workbook_attempts.append(("Legacy Excel direct", load_prin_database_from_xlrd, raw_bytes))
        if decrypted_bytes is not None:
            workbook_attempts.append(("Legacy Excel password", load_prin_database_from_xlrd, decrypted_bytes))
    else:
        workbook_attempts.append(("OpenXML direct", load_prin_database_from_openpyxl, raw_bytes))
        if decrypted_bytes is not None:
            workbook_attempts.append(("OpenXML password", load_prin_database_from_openpyxl, decrypted_bytes))
        workbook_attempts.append(("Legacy Excel direct", load_prin_database_from_xlrd, raw_bytes))
        if decrypted_bytes is not None:
            workbook_attempts.append(("Legacy Excel password", load_prin_database_from_xlrd, decrypted_bytes))

    for attempt_label, workbook_loader, candidate_bytes in workbook_attempts:
        try:
            prin_lookup = workbook_loader(candidate_bytes)
            break
        except Exception as workbook_error:
            load_errors.append(f"{attempt_label} open error: {format_excel_load_error(workbook_error)}")

    if prin_lookup is None:
        if decrypt_error is not None:
            load_errors.insert(0, f'Password decrypt error: {format_excel_load_error(decrypt_error)}')

        detected_type_note = ""
        if file_info["container"] != "unknown":
            detected_type_note = f' Detected file type: {file_info["description"]}.'

        raise ValueError(
            "DATABASE FOR TPAP PRIN file could not be opened. "
            'Please upload a valid .xls, .xlsx, or .xlsm file, or the correct password-protected workbook using password "BPI".'
            + detected_type_note
            + " "
            + " Errors: "
            + " | ".join(load_errors)
        )

    return prin_lookup


def load_excel_files(tpap_file, bpi_file):
    tpap_password = build_tpap_password()
    tpap_decrypted = decrypt_file(tpap_file, tpap_password)
    bpi_decrypted = decrypt_file(bpi_file, BPI_PASSWORD)

    tpap_wb = openpyxl.load_workbook(tpap_decrypted)
    required_tpap_sheets = ["CC", "PL"]
    missing_tpap_sheets = [sheet_name for sheet_name in required_tpap_sheets if sheet_name not in tpap_wb.sheetnames]
    if missing_tpap_sheets:
        raise ValueError(
            "Missing required TPAP Monitoring sheet(s): "
            + ", ".join(f"'{sheet_name}'" for sheet_name in missing_tpap_sheets)
        )
    tpap_ws_map = {sheet_name: tpap_wb[sheet_name] for sheet_name in required_tpap_sheets}
    tpap_header_rows = {
        sheet_name: detect_header_row(tpap_ws_map[sheet_name]) for sheet_name in required_tpap_sheets
    }

    bpi_wb = openpyxl.load_workbook(bpi_decrypted, data_only=True)
    if "PAYMENTS" not in bpi_wb.sheetnames:
        raise ValueError("Sheet 'PAYMENTS' not found in BPI Payment Extraction file.")
    bpi_ws = bpi_wb["PAYMENTS"]
    bpi_payment_number_formats = build_bpi_payment_number_format_lookup(bpi_ws)

    tpap_decrypted.seek(0)
    bpi_decrypted.seek(0)

    tpap_sheets = {}
    for sheet_name in required_tpap_sheets:
        tpap_decrypted.seek(0)
        tpap_sheets[sheet_name] = pd.read_excel(
            tpap_decrypted,
            sheet_name=sheet_name,
            header=tpap_header_rows[sheet_name] - 1,
            engine="openpyxl",
        )
    bpi_df = pd.read_excel(bpi_decrypted, sheet_name="PAYMENTS", engine="openpyxl")

    for sheet_name, tpap_df in tpap_sheets.items():
        tpap_df.columns = tpap_df.columns.astype(str).str.strip()
        tpap_sheets[sheet_name] = tpap_df
    bpi_df.columns = bpi_df.columns.astype(str).str.strip()

    return tpap_sheets, bpi_df, tpap_wb, tpap_ws_map, tpap_header_rows, bpi_payment_number_formats


def validate_required_columns(tpap_df, bpi_df=None, sheet_name="CC", require_bpi_columns=False):
    required_tpap_columns = [
        ("Customer No.", ["Customer No."]),
        ("PAP", ["PAP"]),
        ("PAP_Code", ["PAP_Code"]),
        (PRIN_HEADER, [PRIN_HEADER]),
        (OB_HEADER, [OB_HEADER]),
        (PTP_DATE_HEADER, PTP_DATE_HEADER_CANDIDATES),
        ("EPA TERM", ["EPA TERM"]),
        ("Program Start Date", ["Program Start Date"]),
        ("Program End Date", ["Program End Date"]),
        (FACE_AMOUNT_AUDIT_HEADER, [FACE_AMOUNT_AUDIT_HEADER]),
        ("Discount Rate Principal (%)", ["Discount Rate Principal (%)"]),
        ("Discount Rate Charges (%)", ["Discount Rate Charges (%)"]),
        (WRITE_OFF_DATE_HEADER, WRITE_OFF_DATE_HEADER_CANDIDATES),
        ("AREA", ["AREA"]),
        ("AGENCY", ["AGENCY"]),
        ("RECO LEVEL", ["RECO LEVEL"]),
    ]
    missing_tpap_columns = [
        canonical_name
        for canonical_name, header_candidates in required_tpap_columns
        if get_first_existing_column(tpap_df, header_candidates) is None
    ]
    if missing_tpap_columns:
        st.error(
            f"Error: Missing required column(s) in TPAP Monitoring sheet '{sheet_name}': "
            + ", ".join(f"'{col}'" for col in missing_tpap_columns)
        )
        return False

    if require_bpi_columns and (bpi_df is None or "Account No" not in bpi_df.columns):
        st.error("Error: 'Account No' column not found in BPI Payment Extraction sheet.")
        return False

    return True


def ensure_columns_exist(df):
    df = df.copy()
    df = ensure_canonical_column(df, "Payment", ["Payment"])
    df = ensure_canonical_column(df, "Face Amount", ["Face Amount", FACE_AMOUNT_AUDIT_HEADER])
    df = ensure_canonical_column(df, "Status", STATUS_HEADER_CANDIDATES)
    df = ensure_canonical_column(df, "Date Updated", DATE_UPDATED_HEADER_CANDIDATES)
    df = ensure_canonical_column(df, "PTP Date", PTP_DATE_HEADER_CANDIDATES)
    df = ensure_canonical_column(df, WRITE_OFF_DATE_HEADER, WRITE_OFF_DATE_HEADER_CANDIDATES)
    df = ensure_canonical_column(df, SOURCE_OF_CONTACT_HEADER, SOURCE_OF_CONTACT_HEADER_CANDIDATES)
    return df


def get_bpi_payment_column(bpi_df):
    """
    Use COLUMN D from the PAYMENTS sheet of BPI PYTS EXTRACTION.
    Excel column D = pandas column index 3.
    """
    if len(bpi_df.columns) < 4:
        raise ValueError(
            "BPI PAYMENTS sheet does not have enough columns. "
            "Column D was expected but not found."
        )

    return bpi_df.columns[3]


def copy_cell_style(source_cell, target_cell):
    target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format


def write_cell_value_preserving_format(cell, value, number_format_override=None):
    original_style = copy(cell._style)
    original_font = copy(cell.font)
    original_fill = copy(cell.fill)
    original_border = copy(cell.border)
    original_alignment = copy(cell.alignment)
    original_protection = copy(cell.protection)
    original_number_format = cell.number_format

    if pd.isna(value):
        cell.value = None
    elif isinstance(value, pd.Timestamp):
        cell.value = value.to_pydatetime()
    else:
        cell.value = value

    cell._style = original_style
    cell.font = original_font
    cell.fill = original_fill
    cell.border = original_border
    cell.alignment = original_alignment
    cell.protection = original_protection
    cell.number_format = number_format_override or original_number_format


def build_header_map(ws, header_row=1):
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if value is not None:
            header_map[str(value).strip()] = col_idx
    return header_map


def detect_header_row(ws, expected_headers=None, max_scan_rows=10):
    expected_headers = expected_headers or [
        "Customer No.",
        "PAP",
        "PAP_Code",
        PRIN_HEADER,
        OB_HEADER,
        PTP_DATE_HEADER,
        FACE_AMOUNT_AUDIT_HEADER,
        "Discount Rate Principal (%)",
        "Discount Rate Charges (%)",
    ]

    best_row = 1
    best_match_count = -1

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        header_map = build_header_map(ws, header_row=row_idx)
        match_count = sum(1 for header in expected_headers if header in header_map)
        if match_count > best_match_count:
            best_row = row_idx
            best_match_count = match_count

    return best_row


def ensure_worksheet_column(ws, header_candidates, create_header=None, header_row=1):
    header_map = build_header_map(ws, header_row=header_row)

    for header in header_candidates:
        if header in header_map:
            return header_map[header]

    new_header = create_header or header_candidates[0]
    ensure_worksheet_columns(ws, [new_header], header_row=header_row)
    return build_header_map(ws, header_row=header_row)[new_header]


def ensure_worksheet_columns(ws, required_headers, header_row=1):
    header_map = build_header_map(ws, header_row)

    for header in required_headers:
        if header in header_map:
            continue

        template_col = ws.max_column
        new_col = ws.max_column + 1

        template_header_cell = ws.cell(header_row, template_col)
        new_header_cell = ws.cell(header_row, new_col)
        new_header_cell.value = header
        copy_cell_style(template_header_cell, new_header_cell)

        template_letter = get_column_letter(template_col)
        new_letter = get_column_letter(new_col)
        ws.column_dimensions[new_letter].width = ws.column_dimensions[template_letter].width

        for row_idx in range(header_row + 1, ws.max_row + 1):
            copy_cell_style(ws.cell(row_idx, template_col), ws.cell(row_idx, new_col))

        header_map[header] = new_col

    return header_map


# ==============================
# PAYMENT UPDATE
# ==============================

def update_payment_data(tpap_df, bpi_df, bpi_payment_number_formats=None):
    payment_col = get_bpi_payment_column(bpi_df)

    temp_bpi = bpi_df.copy()
    temp_bpi["Account Key"] = temp_bpi["Account No"].apply(normalize_key)
    temp_bpi[payment_col] = pd.to_numeric(temp_bpi[payment_col], errors="coerce")

    # Keep only valid account number rows
    temp_bpi = temp_bpi.dropna(subset=["Account Key"])

    # Keep only rows with actual payment values
    temp_bpi = temp_bpi.dropna(subset=[payment_col])

    # Sum all payment rows that belong to the same account number.
    temp_bpi = temp_bpi.groupby("Account Key", as_index=False)[payment_col].sum()

    payment_lookup = temp_bpi.set_index("Account Key")[payment_col]

    temp_tpap = tpap_df.copy()
    temp_tpap["Customer Key"] = temp_tpap["Customer No."].apply(normalize_key)

    mapped_payment = temp_tpap["Customer Key"].map(payment_lookup)
    payment_updated_mask = mapped_payment.notna()

    # Update only if BPI has a payment; otherwise keep current CC payment
    temp_tpap["Payment"] = temp_tpap["Payment"].where(mapped_payment.isna(), mapped_payment)

    payment_format_overrides = {}
    if bpi_payment_number_formats is None:
        bpi_payment_number_formats = {}

    for df_index in temp_tpap.index[payment_updated_mask]:
        customer_key = temp_tpap.at[df_index, "Customer Key"]
        source_number_format = bpi_payment_number_formats.get(customer_key)
        payment_value = temp_tpap.at[df_index, "Payment"]
        payment_format_overrides[df_index] = get_payment_number_format(
            payment_value,
            source_number_format=source_number_format,
        )

    temp_tpap.drop(columns=["Customer Key"], inplace=True)
    return temp_tpap, payment_format_overrides


def apply_prin_database_audit(df, prin_lookup, header_row=1):
    customer_keys = df["Customer No."].apply(normalize_key)
    customer_key_present_mask = customer_keys.notna() & customer_keys.ne("")
    database_record_values = customer_keys.map(prin_lookup)
    database_prin_values = database_record_values.apply(
        lambda value: value.get(PRIN_HEADER) if isinstance(value, dict) else None
    )
    database_ob_values = database_record_values.apply(
        lambda value: value.get(OB_HEADER) if isinstance(value, dict) else None
    )
    tpap_prin_values = df[PRIN_HEADER].apply(normalize_currency_value)
    tpap_ob_values = df[OB_HEADER].apply(normalize_currency_value)
    database_prin_normalized = database_prin_values.apply(normalize_currency_value)
    database_ob_normalized = database_ob_values.apply(normalize_currency_value)

    db_record_found_mask = customer_key_present_mask & customer_keys.isin(prin_lookup.keys())
    prin_match_mask = (
        (tpap_prin_values.isna() & database_prin_normalized.isna())
        | tpap_prin_values.eq(database_prin_normalized)
    )
    ob_match_mask = (
        (tpap_ob_values.isna() & database_ob_normalized.isna())
        | tpap_ob_values.eq(database_ob_normalized)
    )
    prin_mismatch_mask = db_record_found_mask & ~prin_match_mask
    ob_mismatch_mask = db_record_found_mask & ~ob_match_mask
    missing_database_mask = customer_key_present_mask & ~db_record_found_mask
    mismatch_mask = prin_mismatch_mask | ob_mismatch_mask
    violation_mask = mismatch_mask | missing_database_mask

    mismatch_fields = []
    for row_index in df.index[violation_mask]:
        if missing_database_mask.loc[row_index]:
            mismatch_fields.append("Account number not found")
            continue

        row_mismatches = []
        if prin_mismatch_mask.loc[row_index]:
            row_mismatches.append(PRIN_HEADER)
        if ob_mismatch_mask.loc[row_index]:
            row_mismatches.append(OB_HEADER)
        mismatch_fields.append(", ".join(row_mismatches))

    prin_database_details = build_finding_detail_table(
        df,
        violation_mask,
        ["Customer No.", "PAP", "PAP_Code", PRIN_HEADER, OB_HEADER],
        finding_builder=lambda row_index: (
            f'Account/Customer No. "{format_finding_value(df.at[row_index, "Customer No."])}" was not found in '
            f'"{PRIN_DATABASE_SHEET_NAME}" of DATABASE FOR TPAP PRIN.'
            if missing_database_mask.loc[row_index]
            else "; ".join(
                finding
                for finding in [
                    (
                        f'Current "{PRIN_HEADER}": {format_finding_value(df.at[row_index, PRIN_HEADER])}; '
                        f'Database "{PRIN_HEADER}": {format_finding_value(database_prin_values.loc[row_index])}'
                    )
                    if prin_mismatch_mask.loc[row_index]
                    else "",
                    (
                        f'Current "{OB_HEADER}": {format_finding_value(df.at[row_index, OB_HEADER])}; '
                        f'Database "{OB_HEADER}": {format_finding_value(database_ob_values.loc[row_index])}'
                    )
                    if ob_mismatch_mask.loc[row_index]
                    else "",
                ]
                if finding
            )
        ),
        extra_columns={
            "Mismatch Fields": mismatch_fields,
            f"Database {PRIN_HEADER}": [
                format_display_value(database_prin_values.loc[row_index]) for row_index in df.index[violation_mask]
            ],
            f"Database {OB_HEADER}": [
                format_display_value(database_ob_values.loc[row_index]) for row_index in df.index[violation_mask]
            ],
        },
        header_row=header_row,
    )

    return {
        "type": "prin_mismatch_vs_database",
        "category": "AUDIT",
        "priority": "HIGH",
        "title": f'"{PRIN_HEADER}" and "{OB_HEADER}" must match DATABASE FOR TPAP PRIN sheet "{PRIN_DATABASE_SHEET_NAME}"',
        "rows_checked": int(customer_key_present_mask.sum()),
        "violations_count": int(violation_mask.sum()),
        "mismatch_count": int(mismatch_mask.sum()),
        "prin_mismatch_count": int(prin_mismatch_mask.sum()),
        "ob_mismatch_count": int(ob_mismatch_mask.sum()),
        "missing_database_count": int(missing_database_mask.sum()),
        "details": prin_database_details,
    }


# ==============================
# PAYMENT LOGIC
# ==============================

def apply_payment_logic(df):
    df = df.copy()
    df = round_otp_face_amounts_to_nearest_hundred(df)

    current_datetime = datetime.today()
    today_value = pd.Timestamp(current_datetime.date())
    today_date = current_datetime.date()
    df["_status_changed"] = False

    df["Payment"] = pd.to_numeric(df["Payment"], errors="coerce")
    df["Face Amount"] = pd.to_numeric(df["Face Amount"], errors="coerce")
    pap_code_series = df["PAP_Code"].fillna("").astype(str).str.strip().str.upper()
    pap_series = df["PAP"].fillna("").astype(str).str.strip().str.upper()
    original_status_source = df[AUDIT_STATUS_COLUMN] if AUDIT_STATUS_COLUMN in df.columns else df["Status"]
    original_status_series = original_status_source.fillna("").astype(str).str.strip().str.upper()

    ptp_dates = pd.to_datetime(df["PTP Date"], errors="coerce") if "PTP Date" in df.columns else pd.Series([pd.NaT] * len(df))

    def set_status_if_changed(row_index, new_status):
        current_value = df.at[row_index, "Status"]
        current_value = "" if pd.isna(current_value) else str(current_value).strip().upper()
        if current_value == new_status:
            return

        df.at[row_index, "Status"] = new_status
        df.at[row_index, "Date Updated"] = today_value
        df.at[row_index, "_status_changed"] = True

    for index in df.index:
        current_status = str(df.at[index, "Status"]).strip().upper() if pd.notna(df.at[index, "Status"]) else ""
        original_status = original_status_series.loc[index] if index in original_status_series.index else current_status
        if current_status in PROTECTED_STATUSES:
            continue

        payment = df.at[index, "Payment"]
        face_amount = df.at[index, "Face Amount"]
        ptp_date = ptp_dates.loc[index] if index in ptp_dates.index else pd.NaT
        ptp_due_date = ptp_date.date() if pd.notna(ptp_date) else None
        ptp_is_past_due = ptp_due_date is not None and ptp_due_date < today_date
        has_actual_payment = pd.notna(payment) and payment > 0
        pap_code_value = pap_code_series.loc[index]

        if pap_code_value == PARTIAL_PAP_CODE or pap_code_value == "":
            continue

        if original_status in UNDER_NEGO_PAYMENT_AUDIT_STATUSES:
            if has_actual_payment and pd.notna(face_amount) and payment >= face_amount:
                set_status_if_changed(index, COMPLIED_STATUS)
            elif not has_actual_payment and ptp_is_past_due:
                set_status_if_changed(index, DEFAULTED_STATUS)
            continue

        if pap_series.loc[index] == "EPA":
            if has_actual_payment and pd.notna(face_amount) and payment >= face_amount:
                set_status_if_changed(index, COMPLIED_STATUS)
                continue

            if pd.notna(face_amount) and (pd.isna(payment) or payment < face_amount):
                if ptp_is_past_due:
                    set_status_if_changed(index, DEFAULTED_STATUS)
                else:
                    set_status_if_changed(index, AVAILED_STATUS)
                continue

        if has_actual_payment and pd.notna(face_amount) and payment >= face_amount:
            set_status_if_changed(index, COMPLIED_STATUS)
            continue

        if pd.notna(face_amount) and (pd.isna(payment) or payment < face_amount):
            if ptp_is_past_due:
                set_status_if_changed(index, DEFAULTED_STATUS)
            elif ptp_due_date is not None:
                set_status_if_changed(index, AVAILED_STATUS)
            continue

        if has_actual_payment:
            set_status_if_changed(index, AVAILED_STATUS)
            continue

        if pd.notna(ptp_date):
            if (pd.isna(payment) or payment <= 0) and ptp_due_date < today_date:
                set_status_if_changed(index, DEFAULTED_STATUS)

            elif (pd.isna(payment) or payment <= 0) and ptp_due_date >= today_date:
                set_status_if_changed(index, AVAILED_STATUS)

    return df


def apply_audit_checks(df, sheet_name="CC", header_row=1):
    df = df.copy()
    df = round_otp_face_amounts_to_nearest_hundred(df)
    df, under_nego_payment_audit, partial_face_amount_blank_mask = apply_under_nego_payment_audit(
        df,
        sheet_name=sheet_name,
    )
    df, face_amount_calculation_audit = populate_face_amount_from_discount_principal(
        df,
        exclude_mask=partial_face_amount_blank_mask,
        header_row=header_row,
    )

    pap_code_series = df["PAP_Code"].fillna("").astype(str)
    pap_code_upper_series = pap_code_series.str.strip().str.upper()
    pap_series = df["PAP"].fillna("").astype(str).str.strip().str.upper()
    status_series = df["Status"].fillna("").astype(str).str.strip().str.upper()
    ptp_has_value = df["PTP Date"].apply(has_meaningful_value)
    start_has_value = df["Program Start Date"].apply(has_meaningful_value)
    end_has_value = df["Program End Date"].apply(has_meaningful_value)
    epa_term_has_value = df["EPA TERM"].apply(has_meaningful_value)
    face_amount_has_value = df[FACE_AMOUNT_AUDIT_HEADER].apply(has_meaningful_value)
    discount_principal_has_value = df["Discount Rate Principal (%)"].apply(has_discount_rate_principal_value)
    discount_charges_has_value = df["Discount Rate Charges (%)"].apply(has_discount_rate_charges_value)
    payment_has_value_mask = df["Payment"].apply(has_payment_value)

    payment_numeric = pd.to_numeric(df["Payment"], errors="coerce")
    face_amount_numeric = pd.to_numeric(df["Face Amount"], errors="coerce")
    ptp_dates = pd.to_datetime(df["PTP Date"], errors="coerce")
    today_date = datetime.today().date()
    today_timestamp = pd.Timestamp(today_date)
    ptp_past_due_mask = ptp_dates.notna() & (ptp_dates < today_timestamp)

    partial_mask = pap_code_upper_series.eq(PARTIAL_PAP_CODE)
    partial_status_needs_update_mask = partial_mask & ~status_series.eq(UNDER_NEGO_STATUS)
    partial_due_date_present_mask = partial_mask & ptp_has_value
    partial_rule_violation_mask = partial_status_needs_update_mask | partial_due_date_present_mask
    partial_rule_details = build_audit_detail_table(
        df,
        partial_rule_violation_mask,
        ["Customer No.", "PAP_Code", "Status", "PTP Date"],
        extra_columns={
            "Required Status": UNDER_NEGO_STATUS,
            f"Required {PTP_DATE_HEADER}": "Blank",
        },
        header_row=header_row,
    )

    pap_blank_mask = pap_series.eq("")
    pap_code_blank_mask = pap_code_upper_series.eq("")
    blank_pap_or_code_mask = pap_blank_mask | pap_code_blank_mask
    blank_pap_or_code_status_violation_mask = blank_pap_or_code_mask & ~status_series.eq(REFUSED_STATUS)
    blank_pap_or_code_details = build_audit_detail_table(
        df,
        blank_pap_or_code_status_violation_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status"],
        extra_columns={"Required Status": REFUSED_STATUS},
        header_row=header_row,
    )

    epa_pap_mask = pap_series.eq(EPA_PAP_VALUE)
    epa_requires_bau_epa_mask = epa_pap_mask & ~pap_code_upper_series.eq(BAU_EPA_PAP_CODE)
    epa_requires_bau_epa_details = build_audit_detail_table(
        df,
        epa_requires_bau_epa_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status"],
        extra_columns={"Required PAP_Code": BAU_EPA_PAP_CODE},
        header_row=header_row,
    )

    partial_pap_mask = pap_series.eq(PARTIAL_PAP_VALUE)
    partial_pap_violation_mask = partial_pap_mask & (
        ~pap_code_upper_series.eq(PARTIAL_PAP_CODE) | ~status_series.eq(UNDER_NEGO_STATUS)
    )
    partial_pap_details = build_audit_detail_table(
        df,
        partial_pap_violation_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status"],
        extra_columns={
            "Required PAP_Code": PARTIAL_PAP_CODE,
            "Required Status": UNDER_NEGO_STATUS,
        },
        header_row=header_row,
    )

    otp_pap_mask = pap_series.eq(OTP_PAP_VALUE)
    otp_has_bau_epa_mask = otp_pap_mask & pap_code_upper_series.eq(BAU_EPA_PAP_CODE)
    otp_has_bau_epa_details = build_audit_detail_table(
        df,
        otp_has_bau_epa_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status"],
        header_row=header_row,
    )
    otp_face_amount_values = df[FACE_AMOUNT_AUDIT_HEADER].apply(parse_numeric_value)
    otp_face_amount_present_mask = otp_pap_mask & otp_face_amount_values.notna()
    otp_rounding_skip_mask = (
        otp_face_amount_present_mask
        & status_series.eq(COMPLIED_STATUS)
        & payment_numeric.notna()
        & face_amount_numeric.notna()
        & payment_numeric.round(2).eq(face_amount_numeric.round(2))
    )
    otp_face_amount_audit_mask = otp_face_amount_present_mask & ~otp_rounding_skip_mask
    otp_required_face_amount = otp_face_amount_values.apply(round_up_to_nearest_hundred)
    otp_face_amount_rounding_violation_mask = otp_face_amount_audit_mask & ~otp_face_amount_values.eq(
        otp_required_face_amount
    )
    otp_face_amount_rounding_details = build_audit_detail_table(
        df,
        otp_face_amount_rounding_violation_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status", FACE_AMOUNT_AUDIT_HEADER],
        extra_columns={
            f"Required {FACE_AMOUNT_AUDIT_HEADER}": otp_required_face_amount[
                otp_face_amount_rounding_violation_mask
            ].tolist()
        },
        header_row=header_row,
    )

    tpap_otp_mask = pap_code_upper_series.eq(TPAP_OTP_PAP_CODE)
    tpap_otp_refused_mask = tpap_otp_mask & status_series.eq(REFUSED_STATUS)
    tpap_otp_refused_details = build_audit_detail_table(
        df,
        tpap_otp_refused_mask,
        ["Customer No.", "PAP_Code", "Status", "PTP Date"],
        header_row=header_row,
    )

    bau_code_mask = pap_code_upper_series.eq(BAU_PAP_CODE)
    bau_has_required_discount_charge_mask = df["Discount Rate Charges (%)"].apply(has_required_bau_discount_rate_charge)
    bau_charge_missing_mask = bau_code_mask & ~bau_has_required_discount_charge_mask
    bau_charge_violations = build_audit_detail_table(
        df,
        bau_charge_missing_mask,
        ["Customer No.", "PAP", "PAP_Code", "Discount Rate Charges (%)"],
        header_row=header_row,
    )

    tpap_mask = pap_code_upper_series.str.contains("TPAP", case=False, na=False)
    tpap_missing_dates_mask = tpap_mask & ~(start_has_value & end_has_value)
    tpap_missing_discount_principal_mask = tpap_mask & ~discount_principal_has_value
    tpap_has_discount_charges_mask = tpap_mask & discount_charges_has_value

    tpap_date_violations = build_audit_detail_table(
        df,
        tpap_missing_dates_mask,
        ["Customer No.", "PAP", "PAP_Code", "Program Start Date", "Program End Date"],
        header_row=header_row,
    )
    if not tpap_date_violations.empty:
        missing_fields = []
        for row_index in df.index[tpap_missing_dates_mask]:
            fields = []
            if not start_has_value.loc[row_index]:
                fields.append("Program Start Date")
            if not end_has_value.loc[row_index]:
                fields.append("Program End Date")
            missing_fields.append(", ".join(fields))

        tpap_date_violations["Missing Fields"] = missing_fields

    tpap_discount_principal_violations = build_audit_detail_table(
        df,
        tpap_missing_discount_principal_mask,
        ["Customer No.", "PAP", "PAP_Code", "Discount Rate Principal (%)"],
        header_row=header_row,
    )

    tpap_discount_charges_violations = build_audit_detail_table(
        df,
        tpap_has_discount_charges_mask,
        ["Customer No.", "PAP", "PAP_Code", "Discount Rate Charges (%)"],
        header_row=header_row,
    )

    refused_status_mask = status_series.eq(REFUSED_STATUS)
    pap_has_value = df["PAP"].apply(has_meaningful_value)
    pap_code_has_value = df["PAP_Code"].apply(has_meaningful_value)
    refused_violation_mask = refused_status_mask & (
        pap_has_value
        | pap_code_has_value
        | payment_has_value_mask
        | start_has_value
        | end_has_value
        | discount_principal_has_value
        | ptp_has_value
        | face_amount_has_value
        | epa_term_has_value
    )
    refused_status_details = build_finding_detail_table(
        df,
        refused_violation_mask,
        [
            "Customer No.",
            "PAP",
            "PAP_Code",
            "Status",
            "Payment",
            "Program Start Date",
            "Program End Date",
            "PTP Date",
            FACE_AMOUNT_AUDIT_HEADER,
            "EPA TERM",
            "Discount Rate Principal (%)",
        ],
        finding_builder=lambda row_index: ", ".join(
            field_name
            for field_name, field_mask in [
                ("PAP", pap_has_value),
                ("PAP_Code", pap_code_has_value),
                ("Payment", payment_has_value_mask),
                ("Program Start Date", start_has_value),
                ("Program End Date", end_has_value),
                ("Discount Rate Principal (%)", discount_principal_has_value),
                (PTP_DATE_HEADER, ptp_has_value),
                (FACE_AMOUNT_AUDIT_HEADER, face_amount_has_value),
                ("EPA TERM", epa_term_has_value),
            ]
            if field_mask.loc[row_index]
        ),
        header_row=header_row,
    )

    under_nego_status_mask = status_series.eq(UNDER_NEGO_STATUS)
    under_nego_violation_mask = under_nego_status_mask & (
        start_has_value
        | end_has_value
        | discount_principal_has_value
        | ptp_has_value
        | face_amount_has_value
        | epa_term_has_value
    )
    under_nego_status_details = build_finding_detail_table(
        df,
        under_nego_violation_mask,
        [
            "Customer No.",
            "PAP",
            "PAP_Code",
            "Status",
            "Program Start Date",
            "Program End Date",
            "PTP Date",
            FACE_AMOUNT_AUDIT_HEADER,
            "EPA TERM",
            "Discount Rate Principal (%)",
        ],
        finding_builder=lambda row_index: ", ".join(
            field_name
            for field_name, field_mask in [
                ("Program Start Date", start_has_value),
                ("Program End Date", end_has_value),
                ("Discount Rate Principal (%)", discount_principal_has_value),
                (PTP_DATE_HEADER, ptp_has_value),
                (FACE_AMOUNT_AUDIT_HEADER, face_amount_has_value),
                ("EPA TERM", epa_term_has_value),
            ]
            if field_mask.loc[row_index]
        ),
        header_row=header_row,
    )

    reco_details_required_status_mask = status_series.isin(STATUSES_REQUIRING_RECO_DETAILS)
    reco_required_field_masks = {
        column_name: df[column_name].apply(has_meaningful_value)
        for column_name in STATUS_REQUIRED_RECO_COLUMNS
    }
    reco_required_all_present_mask = pd.Series(True, index=df.index)
    for field_mask in reco_required_field_masks.values():
        reco_required_all_present_mask &= field_mask
    reco_required_violation_mask = reco_details_required_status_mask & ~reco_required_all_present_mask
    reco_required_details = build_audit_detail_table(
        df,
        reco_required_violation_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status", *STATUS_REQUIRED_RECO_COLUMNS],
        extra_columns={
            "Missing Fields": [
                ", ".join(
                    column_name
                    for column_name, field_mask in reco_required_field_masks.items()
                    if not field_mask.loc[row_index]
                )
                for row_index in df.index[reco_required_violation_mask]
            ]
        },
        header_row=header_row,
    )

    agency_series = df["AGENCY"].fillna("").astype(str)
    agency_upper_series = agency_series.str.strip().str.upper()
    agency_has_value_mask = df["AGENCY"].apply(has_meaningful_value)
    invalid_agency_mask = agency_has_value_mask & ~agency_upper_series.eq(REQUIRED_AGENCY_VALUE)
    invalid_agency_details = build_audit_detail_table(
        df,
        invalid_agency_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status", "AGENCY"],
        extra_columns={"Required AGENCY": REQUIRED_AGENCY_VALUE},
        header_row=header_row,
    )

    availed_status_mask = status_series.eq(AVAILED_STATUS)
    availed_past_due_mask = availed_status_mask & ptp_past_due_mask
    availed_status_details = build_audit_detail_table(
        df,
        availed_past_due_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status", "Payment", FACE_AMOUNT_AUDIT_HEADER, "PTP Date"],
        header_row=header_row,
    )

    defaulted_status_mask = status_series.eq(DEFAULTED_STATUS)
    defaulted_valid_mask = defaulted_status_mask & ptp_past_due_mask & face_amount_numeric.notna() & (
        payment_numeric.isna() | (payment_numeric < face_amount_numeric)
    )
    defaulted_violation_mask = defaulted_status_mask & ~defaulted_valid_mask
    defaulted_status_details = build_finding_detail_table(
        df,
        defaulted_violation_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status", "Payment", FACE_AMOUNT_AUDIT_HEADER, "PTP Date"],
        finding_builder=lambda row_index: "; ".join(
            finding
            for finding, condition in [
                ("PTP Date/Due date is missing", pd.isna(ptp_dates.loc[row_index])),
                ("PTP Date/Due date is not past due", pd.notna(ptp_dates.loc[row_index]) and ptp_dates.loc[row_index] >= today_timestamp),
                (f'"{FACE_AMOUNT_AUDIT_HEADER}" is missing', pd.isna(face_amount_numeric.loc[row_index])),
                ("Payment is not less than FACE AMOUNT", pd.notna(face_amount_numeric.loc[row_index]) and pd.notna(payment_numeric.loc[row_index]) and payment_numeric.loc[row_index] >= face_amount_numeric.loc[row_index]),
            ]
            if condition
        ),
        header_row=header_row,
    )

    complied_status_mask = status_series.eq(COMPLIED_STATUS)
    complied_valid_mask = complied_status_mask & payment_numeric.notna() & face_amount_numeric.notna() & (
        payment_numeric >= face_amount_numeric
    )
    complied_violation_mask = complied_status_mask & ~complied_valid_mask
    complied_status_details = build_finding_detail_table(
        df,
        complied_violation_mask,
        ["Customer No.", "PAP", "PAP_Code", "Status", "Payment", FACE_AMOUNT_AUDIT_HEADER],
        finding_builder=lambda row_index: "; ".join(
            finding
            for finding, condition in [
                ("Payment is missing", pd.isna(payment_numeric.loc[row_index])),
                (f'"{FACE_AMOUNT_AUDIT_HEADER}" is missing', pd.isna(face_amount_numeric.loc[row_index])),
                ("Payment is less than FACE AMOUNT", pd.notna(payment_numeric.loc[row_index]) and pd.notna(face_amount_numeric.loc[row_index]) and payment_numeric.loc[row_index] < face_amount_numeric.loc[row_index]),
            ]
            if condition
        ),
        header_row=header_row,
    )

    audit_results = [
        under_nego_payment_audit,
        {
            "type": "epa_pap_requires_bau_epa_code",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'PAP "{EPA_PAP_VALUE}" must have PAP_Code "{BAU_EPA_PAP_CODE}"',
            "rows_found": int(epa_pap_mask.sum()),
            "violations_count": int(epa_requires_bau_epa_mask.sum()),
            "details": epa_requires_bau_epa_details,
        },
        {
            "type": "partial_pap_requires_partial_code_and_under_nego",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'PAP "{PARTIAL_PAP_VALUE}" must have PAP_Code "{PARTIAL_PAP_CODE}" and status "{UNDER_NEGO_STATUS}"',
            "rows_found": int(partial_pap_mask.sum()),
            "violations_count": int(partial_pap_violation_mask.sum()),
            "details": partial_pap_details,
        },
        {
            "type": "otp_pap_cannot_have_bau_epa_code",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'PAP "{OTP_PAP_VALUE}" must not have PAP_Code "{BAU_EPA_PAP_CODE}"',
            "rows_found": int(otp_pap_mask.sum()),
            "violations_count": int(otp_has_bau_epa_mask.sum()),
            "details": otp_has_bau_epa_details,
        },
        {
            "type": "otp_face_amount_rounded_to_hundred",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'PAP "{OTP_PAP_VALUE}" must have "{FACE_AMOUNT_AUDIT_HEADER}" rounded up to the nearest hundred',
            "rows_checked": int(otp_face_amount_audit_mask.sum()),
            "skipped_complied_equal_rows": int(otp_rounding_skip_mask.sum()),
            "violations_count": int(otp_face_amount_rounding_violation_mask.sum()),
            "details": otp_face_amount_rounding_details,
        },
        {
            "type": "partial_requires_under_nego_and_blank_due_date",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'PAP_Code "{PARTIAL_PAP_CODE}" rows must be "{UNDER_NEGO_STATUS}" and have blank "{PTP_DATE_HEADER}"',
            "partial_rows_found": int(partial_mask.sum()),
            "rows_status_updated": int(partial_status_needs_update_mask.sum()),
            "rows_due_date_cleared": int(partial_due_date_present_mask.sum()),
            "details": partial_rule_details,
        },
        {
            "type": "blank_pap_or_pap_code_sets_refused",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'Blank PAP or PAP_Code rows must be "{REFUSED_STATUS}"',
            "rows_found": int(blank_pap_or_code_mask.sum()),
            "violations_count": int(blank_pap_or_code_status_violation_mask.sum()),
            "details": blank_pap_or_code_details,
        },
        {
            "type": "tpap_otp_refused_status_invalid",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'PAP_Code "{TPAP_OTP_PAP_CODE}" must not have status "{REFUSED_STATUS}"',
            "tpap_otp_rows_found": int(tpap_otp_mask.sum()),
            "violations_count": int(tpap_otp_refused_mask.sum()),
            "details": tpap_otp_refused_details,
        },
        {
            "type": "discount_rate_principal_required_for_tpap",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": 'TPAP rows must have "Discount Rate Principal (%)"',
            "tpap_rows_found": int(tpap_mask.sum()),
            "violations_count": int(tpap_missing_discount_principal_mask.sum()),
            "details": tpap_discount_principal_violations,
        },
        {
            "type": "discount_rate_charges_blank_for_tpap",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": 'TPAP rows must have blank "Discount Rate Charges (%)"',
            "tpap_rows_found": int(tpap_mask.sum()),
            "violations_count": int(tpap_has_discount_charges_mask.sum()),
            "details": tpap_discount_charges_violations,
        },
        face_amount_calculation_audit,
        {
            "type": "program_dates_required_for_tpap",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": "TPAP rows must have Program Start Date and Program End Date",
            "tpap_rows_found": int(tpap_mask.sum()),
            "violations_count": int(tpap_missing_dates_mask.sum()),
            "details": tpap_date_violations,
        },
        {
            "type": "discount_rate_charges_zero_for_bau",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": 'PAP_Code "BAU" must have a value in "Discount Rate Charges (%)"',
            "bau_rows_found": int(bau_code_mask.sum()),
            "violations_count": int(bau_charge_missing_mask.sum()),
            "details": bau_charge_violations,
        },
        {
            "type": "refused_status_field_rules",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": (
                f'Status "{REFUSED_STATUS}" must have blank PAP, PAP_Code, Payment, '
                f'Program Start Date, Program End Date, "Discount Rate Principal (%)", '
                f'"{PTP_DATE_HEADER}", "{FACE_AMOUNT_AUDIT_HEADER}", and "EPA TERM"'
            ),
            "rows_found": int(refused_status_mask.sum()),
            "violations_count": int(refused_violation_mask.sum()),
            "details": refused_status_details,
        },
        {
            "type": "under_nego_status_field_rules",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": (
                f'Status "{UNDER_NEGO_STATUS}" must have blank Program Start Date, Program End Date, '
                f'"Discount Rate Principal (%)", "{PTP_DATE_HEADER}", "{FACE_AMOUNT_AUDIT_HEADER}", '
                f'and "EPA TERM"'
            ),
            "rows_found": int(under_nego_status_mask.sum()),
            "violations_count": int(under_nego_violation_mask.sum()),
            "details": under_nego_status_details,
        },
        {
            "type": "status_requires_reco_details",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": (
                f'Status "{DEFAULTED_STATUS}", "{AVAILED_STATUS}", and "{COMPLIED_STATUS}" must have '
                '"WRITE OFF DATE", "AREA", "AGENCY", "RECO LEVEL", and "SOURCE OF CONTACT"'
            ),
            "rows_found": int(reco_details_required_status_mask.sum()),
            "violations_count": int(reco_required_violation_mask.sum()),
            "details": reco_required_details,
        },
        {
            "type": "agency_must_be_madrid",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'"AGENCY" must only be "{REQUIRED_AGENCY_VALUE}"',
            "rows_found": int(agency_has_value_mask.sum()),
            "violations_count": int(invalid_agency_mask.sum()),
            "details": invalid_agency_details,
        },
        {
            "type": "availed_status_not_past_due",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'Status "{AVAILED_STATUS}" must not be past due',
            "rows_found": int(availed_status_mask.sum()),
            "violations_count": int(availed_past_due_mask.sum()),
            "details": availed_status_details,
        },
        {
            "type": "defaulted_status_rules",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'Status "{DEFAULTED_STATUS}" requires payment below face amount and a past-due "{PTP_DATE_HEADER}"',
            "rows_found": int(defaulted_status_mask.sum()),
            "violations_count": int(defaulted_violation_mask.sum()),
            "details": defaulted_status_details,
        },
        {
            "type": "complied_status_rules",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'Status "{COMPLIED_STATUS}" requires payment greater than or equal to face amount',
            "rows_found": int(complied_status_mask.sum()),
            "violations_count": int(complied_violation_mask.sum()),
            "details": complied_status_details,
        }
    ]

    return df, audit_results


def apply_special_format_audits(df, ws, header_row=1):
    header_map = build_header_map(ws, header_row=header_row)
    audit_results = []

    date_headers = [
        PTP_DATE_HEADER,
        "Program Start Date",
        "Program End Date",
        WRITE_OFF_DATE_HEADER,
        *DATE_UPDATED_HEADER_CANDIDATES,
    ]
    seen_date_headers = []
    for header in date_headers:
        if header in header_map and header not in seen_date_headers:
            seen_date_headers.append(header)

    date_format_rows = []
    for header in seen_date_headers:
        col_idx = header_map[header]
        for excel_row, df_index in enumerate(df.index, start=header_row + 1):
            cell = ws.cell(excel_row, col_idx)
            if not has_meaningful_value(cell.value):
                continue

            cell_value = cell.value
            number_format = cell.number_format
            valid_format = False

            if isinstance(cell_value, str):
                valid_format = is_valid_mmddyyyy_date_text(cell_value)
            elif isinstance(cell_value, (datetime, date)) or cell.is_date:
                # Treat real Excel date cells as valid dates. This avoids false
                # positives when Excel stores the same date with a different
                # built-in date display format.
                valid_format = True

            if valid_format:
                continue

            date_format_rows.append(
                {
                    "Excel Row": excel_row,
                    ACCOUNT_DISPLAY_HEADER: df.at[df_index, "Customer No."] if "Customer No." in df.columns else pd.NA,
                    "PAP_Code": df.at[df_index, "PAP_Code"] if "PAP_Code" in df.columns else pd.NA,
                    "Column": header,
                    "Value": format_display_value(cell_value),
                    "Number Format": number_format,
                }
            )

    prin_format_rows = []
    if PRIN_HEADER in header_map:
        prin_col_idx = header_map[PRIN_HEADER]
        for excel_row, df_index in enumerate(df.index, start=header_row + 1):
            cell = ws.cell(excel_row, prin_col_idx)
            numeric_value = parse_numeric_value(cell.value)
            if numeric_value is None:
                continue

            source_text = str(cell.value).strip() if isinstance(cell.value, str) else ""
            source_decimal_places = len(source_text.split(".", 1)[1]) if "." in source_text else 0
            value_decimal_places = get_value_decimal_places(cell.value)
            format_decimal_places = get_number_format_decimal_places(cell.number_format)
            has_decimal_representation = any(
                decimal_places > 0
                for decimal_places in (source_decimal_places, value_decimal_places, format_decimal_places)
            )
            if float(numeric_value).is_integer() and not has_decimal_representation:
                continue

            if max(source_decimal_places, value_decimal_places, format_decimal_places) >= 2:
                continue

            prin_format_rows.append(
                {
                    "Excel Row": excel_row,
                    ACCOUNT_DISPLAY_HEADER: df.at[df_index, "Customer No."] if "Customer No." in df.columns else pd.NA,
                    "PAP_Code": df.at[df_index, "PAP_Code"] if "PAP_Code" in df.columns else pd.NA,
                    "Column": PRIN_HEADER,
                    "Value": cell.value,
                    "Number Format": cell.number_format,
                    "Required Format": "#,##0.00",
                }
            )

    date_format_details = pd.DataFrame(date_format_rows)
    prin_format_details = pd.DataFrame(prin_format_rows)

    audit_results.append(
        {
            "type": "date_format_mmddyyyy",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'Date cells must use "{SHORT_DATE_NUMBER_FORMAT}"',
            "rows_checked": len(seen_date_headers),
            "violations_count": len(date_format_rows),
            "details": date_format_details,
        }
    )
    audit_results.append(
        {
            "type": "prin_centavo_format",
            "category": "AUDIT",
            "priority": "HIGH",
            "title": f'"{PRIN_HEADER}" must show centavo when it has a decimal value',
            "violations_count": len(prin_format_rows),
            "details": prin_format_details,
        }
    )

    return audit_results


def apply_short_date_format_to_columns(ws, header_row=1):
    date_column_candidates = [
        DATE_UPDATED_HEADER_CANDIDATES,
        PTP_DATE_HEADER_CANDIDATES,
        ["Program Start Date"],
        ["Program End Date"],
        WRITE_OFF_DATE_HEADER_CANDIDATES,
    ]

    for header_candidates in date_column_candidates:
        header_map = build_header_map(ws, header_row=header_row)
        target_col = None

        for header in header_candidates:
            if header in header_map:
                target_col = header_map[header]
                break

        if target_col is None:
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row_idx, target_col)
            formatted_value = coerce_to_excel_datetime(cell.value)

            if formatted_value is None:
                continue

            write_cell_value_preserving_format(
                cell,
                formatted_value,
                number_format_override=SHORT_DATE_NUMBER_FORMAT,
            )


def clear_placeholder_hyphens_in_worksheet(ws, header_row=1):
    for row_idx in range(header_row + 1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if is_placeholder_blank_text(cell.value):
                write_cell_value_preserving_format(cell, None)


def clear_refused_related_columns(ws, header_row=1):
    header_map = build_header_map(ws, header_row=header_row)
    status_col = None

    for header in STATUS_HEADER_CANDIDATES:
        if header in header_map:
            status_col = header_map[header]
            break

    if status_col is None:
        status_col = 8

    columns_to_clear = ["K", "L", "N", "O"]

    for row_idx in range(header_row + 1, ws.max_row + 1):
        status_value = str(ws.cell(row_idx, status_col).value or "").strip().upper()
        if status_value != "REFUSED/NOT AVAILING":
            continue

        for column_letter in columns_to_clear:
            write_cell_value_preserving_format(ws[f"{column_letter}{row_idx}"], None)


# ==============================
# WRITE BACK TO ORIGINAL WORKBOOK
# ==============================

def write_updated_values_to_original_sheet(ws, df, header_row=1, payment_format_overrides=None):
    column_specs = [
        ("Payment", ["Payment"], "Payment"),
        ("Status", STATUS_HEADER_CANDIDATES, "Status"),
        ("Date Updated", DATE_UPDATED_HEADER_CANDIDATES, DATE_UPDATED_HEADER_CANDIDATES[0]),
        ("PTP Date", PTP_DATE_HEADER_CANDIDATES, PTP_DATE_HEADER),
        ("Program Start Date", ["Program Start Date"], "Program Start Date"),
        ("Program End Date", ["Program End Date"], "Program End Date"),
        (WRITE_OFF_DATE_HEADER, WRITE_OFF_DATE_HEADER_CANDIDATES, WRITE_OFF_DATE_HEADER),
        (FACE_AMOUNT_AUDIT_HEADER, [FACE_AMOUNT_AUDIT_HEADER], FACE_AMOUNT_AUDIT_HEADER),
        ("Discount Rate Principal (%)", ["Discount Rate Principal (%)"], "Discount Rate Principal (%)"),
        ("Discount Rate Charges (%)", ["Discount Rate Charges (%)"], "Discount Rate Charges (%)"),
    ]
    status_changed_mask = df.get("_status_changed", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    header_map = {}
    existing_header_map = build_header_map(ws, header_row=header_row)

    for df_col, header_candidates, create_header in column_specs:
        if df_col == "Date Updated":
            existing_header = find_matching_column_name(existing_header_map.keys(), header_candidates)
            if existing_header is not None:
                header_map[df_col] = existing_header_map[existing_header]
            elif status_changed_mask.any():
                header_map[df_col] = ensure_worksheet_column(
                    ws,
                    header_candidates=header_candidates,
                    create_header=create_header,
                    header_row=header_row,
                )
            else:
                header_map[df_col] = None
            continue

        header_map[df_col] = ensure_worksheet_column(
            ws,
            header_candidates=header_candidates,
            create_header=create_header,
            header_row=header_row,
        )

    payment_format_overrides = payment_format_overrides or {}

    for df_col, _, _ in column_specs:
        col_idx = header_map[df_col]
        if col_idx is None:
            continue

        for excel_row, df_index in enumerate(df.index, start=header_row + 1):
            if df_col == "Date Updated" and not status_changed_mask.loc[df_index]:
                continue

            cell = ws.cell(excel_row, col_idx)
            value = df.at[df_index, df_col]
            number_format_override = None

            if df_col == "Payment":
                number_format_override = payment_format_overrides.get(df_index)
                if number_format_override is None:
                    number_format_override = get_payment_number_format(value)
            elif df_col in {
                "Date Updated",
                "PTP Date",
                "Program Start Date",
                "Program End Date",
                WRITE_OFF_DATE_HEADER,
            }:
                number_format_override = SHORT_DATE_NUMBER_FORMAT
            elif df_col == FACE_AMOUNT_AUDIT_HEADER:
                number_format_override = WHOLE_NUMBER_FORMAT
            elif df_col == "Discount Rate Principal (%)" and parse_percentage_value(value) is not None:
                number_format_override = PERCENT_NUMBER_FORMAT

            write_cell_value_preserving_format(
                cell,
                value,
                number_format_override=number_format_override,
            )


def render_audit_results(sheet_name, audit_results):
    st.subheader(f"AUDIT - {sheet_name}")

    for audit in audit_results:
        if audit["type"] == "under_nego_payment_review":
            if audit["rows_reviewed"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'No rows with status "{UNDER_NEGO_STATUS}" were found.'
                )
            elif audit["rows_with_payment"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'Rows with status "{UNDER_NEGO_STATUS}" were found, but no payment was detected.'
                )
            elif audit["rows_needing_under_nego"] > 0 or audit["partial_face_amount_violations"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["rows_with_payment"]} row(s) with status "{UNDER_NEGO_STATUS}" had payment and need review.'
                )
                if audit["rows_needing_under_nego"] > 0:
                    st.error(
                        f'[{audit["category"]}] Priority check failed: '
                        f'{audit["rows_needing_under_nego"]} row(s) should have status "{UNDER_NEGO_STATUS}".'
                    )
                if audit["partial_face_amount_violations"] > 0:
                    st.error(
                        f'[{audit["category"]}] Priority check failed: '
                        f'{audit["partial_face_amount_violations"]} row(s) with PAP_Code "{PARTIAL_PAP_CODE}" should have blank "{FACE_AMOUNT_AUDIT_HEADER}".'
                    )
                if not audit["details"].empty:
                    st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.info(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_with_payment"]} row(s) with status "{UNDER_NEGO_STATUS}" had payment.'
                )
        elif audit["type"] == "partial_requires_under_nego_and_blank_due_date":
            if audit["partial_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'No rows with PAP_Code "{PARTIAL_PAP_CODE}" were found.'
                )
            elif audit["rows_status_updated"] > 0 or audit["rows_due_date_cleared"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["partial_rows_found"]} row(s) with PAP_Code "{PARTIAL_PAP_CODE}" were checked. '
                    f'{audit["rows_status_updated"]} row(s) are not "{UNDER_NEGO_STATUS}" and '
                    f'{audit["rows_due_date_cleared"]} row(s) have a value in "{PTP_DATE_HEADER}".'
                )
                if not audit["details"].empty:
                    st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["partial_rows_found"]} row(s) with PAP_Code "{PARTIAL_PAP_CODE}" were already '
                    f'"{UNDER_NEGO_STATUS}" and already had blank "{PTP_DATE_HEADER}".'
                )
        elif audit["type"] == "blank_pap_or_pap_code_sets_refused":
            if audit["rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with blank "PAP" or "PAP_Code" were found.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) with blank "PAP" or "PAP_Code" should have status "{REFUSED_STATUS}".'
                )
                if not audit["details"].empty:
                    st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_found"]} row(s) with blank "PAP" or "PAP_Code" were already "{REFUSED_STATUS}".'
                )
        elif audit["type"] == "otp_face_amount_rounded_to_hundred":
            skipped_rows = audit.get("skipped_complied_equal_rows", 0)
            if audit["rows_checked"] == 0:
                if skipped_rows > 0:
                    st.info(
                        f'[{audit["category"]}] Priority check completed: '
                        f'No OTP rows required the round-up audit after skipping {skipped_rows} '
                        f'"{COMPLIED_STATUS}" row(s) where Payment equals "{FACE_AMOUNT_AUDIT_HEADER}".'
                    )
                else:
                    st.info(
                        f'[{audit["category"]}] Priority check completed: '
                        f'No rows with PAP "{OTP_PAP_VALUE}" and a value in "{FACE_AMOUNT_AUDIT_HEADER}" were found.'
                    )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) with PAP "{OTP_PAP_VALUE}" have '
                    f'"{FACE_AMOUNT_AUDIT_HEADER}" not rounded up to the nearest hundred.'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_checked"]} row(s) with PAP "{OTP_PAP_VALUE}" already have '
                    f'"{FACE_AMOUNT_AUDIT_HEADER}" rounded up to the nearest hundred.'
                )
            if skipped_rows > 0:
                st.info(
                    f'[{audit["category"]}] {skipped_rows} row(s) were skipped because status is "{COMPLIED_STATUS}" '
                    f'and Payment equals "{FACE_AMOUNT_AUDIT_HEADER}".'
                )
        elif audit["type"] == "tpap_otp_refused_status_invalid":
            if audit["tpap_otp_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'No rows with PAP_Code "{TPAP_OTP_PAP_CODE}" were found.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) with PAP_Code "{TPAP_OTP_PAP_CODE}" have invalid status "{REFUSED_STATUS}".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["tpap_otp_rows_found"]} row(s) with PAP_Code "{TPAP_OTP_PAP_CODE}" were checked and none had invalid "{REFUSED_STATUS}" status.'
                )
        elif audit["type"] == "refused_status_field_rules":
            if audit["rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'No rows with status "{REFUSED_STATUS}" were found.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) with status "{REFUSED_STATUS}" have a value in fields that should be blank.'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_found"]} row(s) with status "{REFUSED_STATUS}" were checked and the required fields are blank.'
                )
        elif audit["type"] == "under_nego_status_field_rules":
            if audit["rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'No rows with status "{UNDER_NEGO_STATUS}" were found.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) with status "{UNDER_NEGO_STATUS}" have a value in fields that should be blank.'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_found"]} row(s) with status "{UNDER_NEGO_STATUS}" were checked and the required fields are blank.'
                )
        elif audit["type"] == "status_requires_reco_details":
            if audit["rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'No rows with status "{DEFAULTED_STATUS}", "{AVAILED_STATUS}", or "{COMPLIED_STATUS}" were found.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) with status "{DEFAULTED_STATUS}", "{AVAILED_STATUS}", or "{COMPLIED_STATUS}" '
                    'are missing one or more required values in "WRITE OFF DATE", "AREA", "AGENCY", "RECO LEVEL", or "SOURCE OF CONTACT".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_found"]} row(s) with status "{DEFAULTED_STATUS}", "{AVAILED_STATUS}", or "{COMPLIED_STATUS}" '
                    'already have values in "WRITE OFF DATE", "AREA", "AGENCY", "RECO LEVEL", and "SOURCE OF CONTACT".'
                )
        elif audit["type"] == "agency_must_be_madrid":
            if audit["rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with a value in "AGENCY" were found.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) have "AGENCY" not equal to "{REQUIRED_AGENCY_VALUE}".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_found"]} row(s) with a value in "AGENCY" already use "{REQUIRED_AGENCY_VALUE}".'
                )
        elif audit["type"] == "epa_term_required_for_epa":
            if audit["epa_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP equal to "EPA" were found for EPA TERM.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} EPA row(s) are missing "EPA TERM".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["epa_rows_found"]} EPA row(s) were checked and "EPA TERM" is present.'
                )
        elif audit["type"] == "discount_rate_principal_required_for_tpap":
            if audit["tpap_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "TPAP" were found for Discount Rate Principal (%).'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} TPAP row(s) are missing "Discount Rate Principal (%)".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["tpap_rows_found"]} TPAP row(s) were checked and "Discount Rate Principal (%)" is present.'
                )
        elif audit["type"] == "face_amount_required_for_tpap":
            if audit["tpap_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "TPAP" were found for FACE AMOUNT (OTP/EPA).'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} TPAP row(s) are missing "FACE AMOUNT (OTP/EPA)".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["tpap_rows_found"]} TPAP row(s) were checked and "FACE AMOUNT (OTP/EPA)" is present.'
                )
        elif audit["type"] == "face_amount_check_from_prin":
            if audit["source_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with a usable percentage in "Discount Rate Principal (%)" were found for FACE AMOUNT calculation outside REFUSED/NOT AVAILING or UNDER NEGO.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) have "FACE AMOUNT (OTP/EPA)" lower than '
                    '"Discount Rate Principal (%)" x "PRIN", rounded up to the nearest hundred.'
                )
                st.dataframe(audit["mismatch_details"], use_container_width=True, hide_index=True)
            elif audit["rows_checked"] > 0:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_checked"]} row(s) were checked and "FACE AMOUNT (OTP/EPA)" is not lower than the rounded computed value.'
                )
            else:
                st.warning(
                    f'[{audit["category"]}] Priority check completed: '
                    'Rows with "Discount Rate Principal (%)" were found, but no valid FACE AMOUNT comparison could be performed.'
                )

            if audit["missing_prin_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["missing_prin_count"]} row(s) have "Discount Rate Principal (%)" '
                    'but missing or invalid "PRIN".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            elif audit.get("status_skipped_rows", 0) > 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'{audit["status_skipped_rows"]} row(s) were skipped because status is "{REFUSED_STATUS}" or "{UNDER_NEGO_STATUS}".'
                )
        elif audit["type"] == "prin_mismatch_vs_database":
            if audit["rows_checked"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    f'No rows with "{ACCOUNT_DISPLAY_HEADER}" were found for "{PRIN_HEADER}" and "{OB_HEADER}" database comparison.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} row(s) have "{PRIN_HEADER}" and/or "{OB_HEADER}" that do not match DATABASE FOR TPAP PRIN.'
                )
                if audit.get("prin_mismatch_count", 0) > 0:
                    st.error(
                        f'[{audit["category"]}] Priority check failed: '
                        f'{audit["prin_mismatch_count"]} row(s) have a different "{PRIN_HEADER}" from the database.'
                    )
                if audit.get("ob_mismatch_count", 0) > 0:
                    st.error(
                        f'[{audit["category"]}] Priority check failed: '
                        f'{audit["ob_mismatch_count"]} row(s) have a different "{OB_HEADER}" from the database.'
                    )
                if audit.get("missing_database_count", 0) > 0:
                    st.error(
                        f'[{audit["category"]}] Priority check failed: '
                        f'{audit["missing_database_count"]} row(s) were not found in DATABASE FOR TPAP PRIN.'
                    )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["rows_checked"]} row(s) were checked and "{PRIN_HEADER}" and "{OB_HEADER}" match DATABASE FOR TPAP PRIN.'
                )
        elif audit["type"] == "program_dates_required_for_tpap":
            if audit["tpap_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "TPAP" were found for Program Start Date and Program End Date.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} TPAP row(s) are missing Program Start Date and/or Program End Date.'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["tpap_rows_found"]} TPAP row(s) were checked and all required program dates are present.'
                )
        elif audit["type"] == "discount_rate_charges_zero_for_bau":
            if audit["bau_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "BAU" were found for Discount Rate Charges (%).'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} BAU row(s) are missing the required "0%" value in "Discount Rate Charges (%)".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["bau_rows_found"]} BAU row(s) were checked and "Discount Rate Charges (%)" is already set correctly.'
                )
        elif audit["type"] == "discount_rate_principal_blank_for_bau":
            if audit["bau_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "BAU" were found for Discount Rate Principal (%).'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} BAU row(s) have a value in "Discount Rate Principal (%)".'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["bau_rows_found"]} BAU row(s) were checked and "Discount Rate Principal (%)" is blank.'
                )
        elif audit["type"] == "program_dates_blank_for_bau":
            if audit["bau_rows_found"] == 0:
                st.info(
                    f'[{audit["category"]}] Priority check completed: '
                    'No rows with PAP_Code containing "BAU" were found.'
                )
            elif audit["violations_count"] > 0:
                st.error(
                    f'[{audit["category"]}] Priority check failed: '
                    f'{audit["violations_count"]} BAU row(s) should have blank Program Start Date and Program End Date.'
                )
                st.dataframe(audit["details"], use_container_width=True, hide_index=True)
            else:
                st.success(
                    f'[{audit["category"]}] Priority check passed: '
                    f'{audit["bau_rows_found"]} BAU row(s) were checked and the program dates are already blank.'
                )
        else:
            st.info(
                f'[{audit["category"]}] Priority check completed: '
                'No issues were found for this audit rule.'
            )

    copyable_report = build_copyable_audit_report(sheet_name, audit_results)
    report_date = datetime.today().strftime(OUTPUT_FILE_DATE_FORMAT).upper()
    with st.expander(f"Copy Or Download {sheet_name} Audit Errors"):
        st.caption("Copy the text below or download it as a .txt file.")
        st.text_area(
            f"{sheet_name} Audit Error Report",
            value=copyable_report,
            height=280,
            key=f"{sheet_name.lower()}_audit_error_report",
        )
        st.download_button(
            label=f"Download {sheet_name} Audit Errors",
            data=copyable_report,
            file_name=f"{sheet_name}_AUDIT_ERRORS_{report_date}.txt",
            mime="text/plain",
            key=f"{sheet_name.lower()}_audit_error_download",
        )


def update_processing_progress(progress_bar, status_placeholder, step_number, total_steps, message):
    progress_value = int((step_number / total_steps) * 100)
    progress_bar.progress(progress_value)
    status_placeholder.caption(f"Step {step_number} of {total_steps}: {message}")


# ==============================
# MAIN
# ==============================

def run():
    st.title("TPAP Monitoring Revoceries")

    uploaded_tpap_file = st.file_uploader(
        "Upload TPAP Monitoring Excel file",
        type=["xlsx", "xls", "xlsm"]
    )

    uploaded_bpi_file = st.file_uploader(
        "Upload BPI Payment Extraction Excel file",
        type=["xlsx", "xls", "xlsm"]
    )

    uploaded_prin_database_file = st.file_uploader(
        "Upload DATABASE FOR TPAP PRIN Excel file (optional)",
        type=["xlsx", "xls", "xlsm"]
    )

    if uploaded_tpap_file is None or uploaded_bpi_file is None:
        st.info("Please upload TPAP Monitoring and BPI Payment Extraction Excel files.")
        return

    total_steps = 8
    current_stage = "starting the automation"
    progress_bar = st.progress(0)
    progress_status = st.empty()
    progress_status.caption("Step 0 of 8: Waiting to start processing.")

    try:
        current_stage = "opening TPAP Monitoring and BPI workbooks"
        update_processing_progress(progress_bar, progress_status, 1, total_steps, "Opening TPAP Monitoring and BPI workbooks...")
        tpap_sheets, bpi_df, tpap_wb, tpap_ws_map, tpap_header_rows, bpi_payment_number_formats = load_excel_files(
            uploaded_tpap_file,
            uploaded_bpi_file,
        )

        prin_lookup = None
        if uploaded_prin_database_file is not None:
            current_stage = "loading the DATABASE FOR TPAP PRIN workbook"
            update_processing_progress(progress_bar, progress_status, 2, total_steps, "Loading the DATABASE FOR TPAP PRIN workbook...")
            prin_lookup = load_prin_database_lookup(uploaded_prin_database_file)
        else:
            current_stage = "skipping the DATABASE FOR TPAP PRIN workbook"
            update_processing_progress(progress_bar, progress_status, 2, total_steps, "Skipping DATABASE FOR TPAP PRIN workbook audit...")

        cc_df = tpap_sheets["CC"]
        pl_df = tpap_sheets["PL"]
        cc_ws = tpap_ws_map["CC"]
        pl_ws = tpap_ws_map["PL"]
        cc_header_row = tpap_header_rows["CC"]
        pl_header_row = tpap_header_rows["PL"]

        current_stage = "validating required columns"
        update_processing_progress(progress_bar, progress_status, 3, total_steps, "Validating required columns...")
        if not validate_required_columns(
            cc_df,
            bpi_df=bpi_df,
            sheet_name="CC",
            require_bpi_columns=True,
        ):
            progress_status.warning("Processing stopped because required columns are missing in the CC sheet or BPI file.")
            return

        if not validate_required_columns(pl_df, sheet_name="PL"):
            progress_status.warning("Processing stopped because required columns are missing in the PL sheet.")
            return

        current_stage = "preparing TPAP sheets"
        update_processing_progress(progress_bar, progress_status, 4, total_steps, "Preparing TPAP sheets and cleaning placeholder values...")
        cc_df = ensure_columns_exist(cc_df)
        pl_df = ensure_columns_exist(pl_df)
        cc_df = normalize_tpap_dataframe(cc_df)
        pl_df = normalize_tpap_dataframe(pl_df)
        pl_df = populate_pl_payment_from_weekly_columns(pl_df)
        clear_placeholder_hyphens_in_worksheet(cc_ws, header_row=cc_header_row)
        clear_placeholder_hyphens_in_worksheet(pl_ws, header_row=pl_header_row)

        current_stage = "updating payments"
        update_processing_progress(progress_bar, progress_status, 5, total_steps, "Updating payments from the BPI extraction...")
        cc_df, payment_format_overrides = update_payment_data(
            cc_df,
            bpi_df,
            bpi_payment_number_formats=bpi_payment_number_formats,
        )

        current_stage = "running audit checks"
        update_processing_progress(progress_bar, progress_status, 6, total_steps, "Running audit checks and payment logic...")
        cc_df, cc_audit_results = apply_audit_checks(cc_df, sheet_name="CC", header_row=cc_header_row)
        pl_df, pl_audit_results = apply_audit_checks(pl_df, sheet_name="PL", header_row=pl_header_row)
        if prin_lookup is not None:
            cc_audit_results.append(apply_prin_database_audit(cc_df, prin_lookup, header_row=cc_header_row))
            pl_audit_results.append(apply_prin_database_audit(pl_df, prin_lookup, header_row=pl_header_row))
        cc_df = apply_payment_logic(cc_df)
        pl_df = apply_payment_logic(pl_df)

        current_stage = "writing updates back to the workbook"
        update_processing_progress(progress_bar, progress_status, 7, total_steps, "Writing updates back to the workbook...")
        write_updated_values_to_original_sheet(
            cc_ws,
            cc_df,
            header_row=cc_header_row,
            payment_format_overrides=payment_format_overrides,
        )
        clear_refused_related_columns(cc_ws, header_row=cc_header_row)
        apply_short_date_format_to_columns(cc_ws, header_row=cc_header_row)

        write_updated_values_to_original_sheet(
            pl_ws,
            pl_df,
            header_row=pl_header_row,
        )
        clear_refused_related_columns(pl_ws, header_row=pl_header_row)
        apply_short_date_format_to_columns(pl_ws, header_row=pl_header_row)
        cc_audit_results.extend(apply_special_format_audits(cc_df, cc_ws, header_row=cc_header_row))
        pl_audit_results.extend(apply_special_format_audits(pl_df, pl_ws, header_row=pl_header_row))
        sheet_audits = {"CC": cc_audit_results, "PL": pl_audit_results}

        current_stage = "building the downloadable files"
        update_processing_progress(progress_bar, progress_status, 8, total_steps, "Building the downloadable files...")
        audit_findings_workbook = build_audit_excel_workbook(sheet_audits)
        audit_report_date = datetime.today().strftime(OUTPUT_FILE_DATE_FORMAT).upper()

        output_file_password = build_tpap_password()
        output = BytesIO()
        tpap_wb.save(output)

        encrypted_output = encrypt_output_workbook(output, output_file_password)
        output_file_name = build_output_filename()

        render_audit_results("CC", cc_audit_results)
        render_audit_results("PL", pl_audit_results)
        if prin_lookup is None:
            st.warning(
                'DATABASE FOR TPAP PRIN was not uploaded, so the "PRIN" and "OB" database comparison audit was skipped.'
            )
        st.subheader("Audit Excel Files")
        st.download_button(
            label="Download Audit Findings Excel",
            data=audit_findings_workbook.getvalue(),
            file_name=f"AUDIT_FINDINGS_{audit_report_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        progress_status.success("Processing completed. Your files are ready to download.")
        st.success("Automation completed successfully!")
        st.caption(f'Download password: "{output_file_password}"')

        st.download_button(
            label="Download Processed File",
            data=encrypted_output.getvalue(),
            file_name=output_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        progress_status.error(f"Processing stopped while {current_stage}.")
        st.error(f"An error occurred: {e}")


if __name__ == "__main__":
    run()
