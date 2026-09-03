import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from io import BytesIO
from datetime import datetime, date, timedelta
import calendar
from pathlib import Path
import re
import msoffcrypto
from zipfile import BadZipFile

# Auto password format: MAD_Apr2026, MAD_May2026, etc.
PASSWORD_PREFIX = "MAD"
OOXML_FILE_SIGNATURE = b"PK\x03\x04"
OLE_FILE_SIGNATURE = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

# --- SETTINGS ---
SHEET_PREFIX = "payment as of"
DATE_COL = "F"
AMOUNT_COL = "E"
LAN2_COL = "C"
DATA_START_ROW = 2
NEW_SHEET_NAME = "Sheet1"
# ----------------


def build_default_password(current_datetime=None, month=None, year=None):
    if month is None or year is None:
        current_datetime = current_datetime or datetime.today()
        month = current_datetime.month
        year = current_datetime.year

    return f"{PASSWORD_PREFIX}_{calendar.month_abbr[int(month)]}{int(year)}"


DEFAULT_PASSWORD = build_default_password()


def find_payment_as_of_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower().startswith(SHEET_PREFIX):
            return name
    return None


def normalize_lan2(val):
    if val is None:
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).strip()
    return str(val).strip()


def parse_excel_date(val):
    if val is None or val == "":
        return None

    if isinstance(val, datetime):
        return val.date()

    if isinstance(val, date):
        return val

    if isinstance(val, (int, float)):
        try:
            converted = from_excel(val)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except Exception:
            return None

    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None

        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass

    return None


def month_str_to_num(m):
    m = m.strip().lower()
    month_map = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    return month_map.get(m)


def extract_month_year_from_sheetname(sheet_name):
    """
    Example:
    'payment as of MAR 2026' -> (3, 2026)
    """
    s = sheet_name.strip()
    s2 = re.sub(r'(?i)^\s*payment\s+as\s+of\s*', '', s).strip()

    parts = re.split(r"\s+", s2)
    if len(parts) >= 2:
        m_num = month_str_to_num(parts[0])
        try:
            y = int(parts[1])
        except Exception:
            y = None

        if m_num and y:
            return m_num, y

    today = datetime.now().date()
    return today.month, today.year


def extract_month_year_from_filename(file_name):
    if not file_name:
        return None

    normalized_name = re.sub(r"[_\-.]+", " ", str(file_name))

    patterns = [
        r"(?i)\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t|tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\s+(\d{4})\b",
        r"(?i)\b(\d{4})\s+(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t|tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, normalized_name)
        if not match:
            continue

        first, second = match.groups()
        if first.isdigit():
            year = int(first)
            month = month_str_to_num(second)
        else:
            month = month_str_to_num(first)
            year = int(second)

        if month and year:
            return month, year

    return None


def build_password_candidates(file_name=None):
    candidates = []
    seen = set()

    inferred_month_year = extract_month_year_from_filename(file_name)
    if inferred_month_year:
        password = build_default_password(month=inferred_month_year[0], year=inferred_month_year[1])
        if password not in seen:
            candidates.append(password)
            seen.add(password)

    current_password = build_default_password()
    if current_password not in seen:
        candidates.append(current_password)

    return candidates


def format_workbook_open_error(error):
    if isinstance(error, BadZipFile):
        return "File is not a valid .xlsx/.xlsm workbook archive."

    if isinstance(error, InvalidFileException):
        message = str(error).strip()
        return message or "Unsupported workbook format."

    message = str(error).strip()
    if message:
        return message

    return error.__class__.__name__


def inspect_uploaded_excel(file_name, raw_bytes):
    normalized_name = Path(file_name or "uploaded workbook").name
    lowered_name = normalized_name.lower()
    header_bytes = raw_bytes[:2048]
    stripped_header_bytes = header_bytes.lstrip().lower()

    if lowered_name.startswith("~$"):
        return {
            "file_name": normalized_name,
            "container": "excel_temp_lock",
            "description": "Excel temporary lock file",
        }

    if not raw_bytes:
        return {
            "file_name": normalized_name,
            "container": "empty",
            "description": "empty file",
        }

    if raw_bytes.startswith(OOXML_FILE_SIGNATURE):
        return {
            "file_name": normalized_name,
            "container": "ooxml_zip",
            "description": "OpenXML workbook",
        }

    if raw_bytes.startswith(OLE_FILE_SIGNATURE):
        return {
            "file_name": normalized_name,
            "container": "ole_compound",
            "description": "password-protected or legacy Excel workbook",
        }

    if stripped_header_bytes.startswith((b"<!doctype html", b"<html")):
        return {
            "file_name": normalized_name,
            "container": "html",
            "description": "HTML document",
        }

    if b"\x00" not in header_bytes:
        has_delimiter = any(delimiter in header_bytes for delimiter in (b",", b";", b"\t"))
        has_line_break = any(line_break in header_bytes for line_break in (b"\r", b"\n"))
        if has_delimiter and has_line_break:
            return {
                "file_name": normalized_name,
                "container": "delimited_text",
                "description": "delimited text file",
            }

    return {
        "file_name": normalized_name,
        "container": "unknown",
        "description": "unrecognized file type",
    }


def build_upload_precheck_error(file_info):
    file_name = file_info["file_name"]
    container = file_info["container"]

    if container == "excel_temp_lock":
        return (
            f'The uploaded file "{file_name}" looks like Excel\'s temporary lock file (~$...), '
            "not the actual workbook. Close the source workbook if it is open, then upload the real file."
        )

    if container == "empty":
        return f'The uploaded file "{file_name}" is empty. Please upload a valid .xlsx or .xlsm workbook.'

    if container == "html":
        return (
            f'The uploaded file "{file_name}" is an HTML page, not an Excel workbook. '
            "Please upload a valid .xlsx or .xlsm workbook."
        )

    if container == "delimited_text":
        return (
            f'The uploaded file "{file_name}" looks like a CSV or text file, not an Excel workbook. '
            "Please upload a valid .xlsx or .xlsm workbook."
        )

    return None


def first_monday(year, month):
    d = date(year, month, 1)
    while d.weekday() != 0:  # Monday = 0
        d += timedelta(days=1)
    return d


def build_week_ranges(year, month):
    """
    Builds weekly ranges (Mon-Fri) starting from the first Monday of the month.
    Adds a Week 5 when the month still has remaining days after Week 4.
    """
    fm = first_monday(year, month)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    ranges = []
    start = fm
    while start <= last_day:
        end = start + timedelta(days=4)  # Mon-Fri
        if end > last_day:
            end = last_day

        ranges.append((start, end))
        start += timedelta(days=7)

    return ranges, fm, last_day


def week_bucket(d, fm, total_weeks):
    """
    Maps any date to week 1..N based on 7-day blocks from first Monday.
    Dates before first Monday -> week 1
    Dates after the last generated week -> last generated week
    """
    delta = (d - fm).days
    if delta < 0:
        return 1

    w = (delta // 7) + 1
    if w < 1:
        return 1
    if w > total_weeks:
        return total_weeks
    return w


def set_date_cell(ws, cell_addr, d):
    ws[cell_addr] = d
    ws[cell_addr].number_format = "mm/dd/yyyy"


def add_amount_to_cell(ws, cell_addr, amount):
    current = ws[cell_addr].value

    if current is None or current == "":
        ws[cell_addr] = amount
    else:
        ws[cell_addr] = current + amount


def set_latest_date(ws, cell_addr, new_date):
    """
    Keep only the latest date for that LAN2 in that week.
    """
    current_date = parse_excel_date(ws[cell_addr].value)

    if current_date is None or new_date > current_date:
        set_date_cell(ws, cell_addr, new_date)


def process_and_input_data_in_existing_sheet(file_stream):
    wb = load_workbook(file_stream)

    sheet_name = find_payment_as_of_sheet(wb)
    if not sheet_name:
        raise Exception("No sheet found that starts with 'payment as of'.")

    ws = wb[sheet_name]

    # Get month/year from sheet name
    m, y = extract_month_year_from_sheetname(sheet_name)
    week_ranges, fm, _ = build_week_ranges(y, m)

    # Ensure Sheet1 exists
    if NEW_SHEET_NAME in wb.sheetnames:
        existing_index = wb.sheetnames.index(NEW_SHEET_NAME)
        wb.remove(wb[NEW_SHEET_NAME])
        new_sheet = wb.create_sheet(NEW_SHEET_NAME, existing_index)
    else:
        new_sheet = wb.create_sheet(NEW_SHEET_NAME)

    # Headers
    new_sheet["A1"] = "LAN2"

    def label_week(i):
        if i < len(week_ranges):
            s, e = week_ranges[i]
            return f"Payment week {i+1} ({s.strftime('%b %d')} to {e.strftime('%d')})"
        return f"Payment week {i+1}"

    week_count = len(week_ranges)
    for i in range(week_count):
        amount_col = 2 + (i * 2)
        date_col = amount_col + 1
        new_sheet.cell(row=1, column=amount_col, value=label_week(i))
        new_sheet.cell(row=1, column=date_col, value=f"Payment Date (Week {i+1})")

    total_col = 2 + (week_count * 2)
    new_sheet.cell(row=1, column=total_col, value="TOTAL")

    # Map each LAN2 to one output row
    lan2_to_row = {}
    out_row = 2

    for row in range(DATA_START_ROW, ws.max_row + 1):
        lan_value = ws[f"{LAN2_COL}{row}"].value
        amount_value = ws[f"{AMOUNT_COL}{row}"].value
        payment_date = parse_excel_date(ws[f"{DATE_COL}{row}"].value)

        lan_key = normalize_lan2(lan_value)

        if not lan_key:
            continue
        if payment_date is None:
            continue
        if amount_value is None or amount_value == 0:
            continue

        if lan_key not in lan2_to_row:
            lan2_to_row[lan_key] = out_row
            new_sheet[f"A{out_row}"] = lan_key
            out_row += 1

        target_row = lan2_to_row[lan_key]
        week_no = week_bucket(payment_date, fm, week_count)
        amount_col_letter = get_column_letter(2 + ((week_no - 1) * 2))
        date_col_letter = get_column_letter(3 + ((week_no - 1) * 2))
        add_amount_to_cell(new_sheet, f"{amount_col_letter}{target_row}", amount_value)
        set_latest_date(new_sheet, f"{date_col_letter}{target_row}", payment_date)

    # TOTAL = sum of all payment week amount columns
    amount_column_letters = [get_column_letter(2 + (i * 2)) for i in range(week_count)]
    for r in range(2, out_row):
        total_formula = ",".join(f"{col_letter}{r}" for col_letter in amount_column_letters)
        new_sheet.cell(row=r, column=total_col, value=f"=SUM({total_formula})")

    # Formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Header formatting
    for col in range(1, total_col + 1):
        col_letter = get_column_letter(col)
        new_sheet[f"{col_letter}1"].font = bold_font
        if col == 1:
            new_sheet[f"{col_letter}1"].fill = red_fill
        new_sheet[f"{col_letter}1"].alignment = center_alignment

    # Center align used cells
    for col in range(1, total_col + 1):
        col_letter = get_column_letter(col)
        for r in range(1, out_row):
            new_sheet[f"{col_letter}{r}"].alignment = center_alignment

    # Adjust column widths
    for col_cells in new_sheet.iter_cols(min_col=1, max_col=total_col):
        max_length = 0
        column_letter = col_cells[0].column_letter

        for cell in col_cells:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass

        new_sheet.column_dimensions[column_letter].width = max_length + 2

    return wb


def load_workbook_with_optional_password(uploaded, password_str=None):
    password_candidates = [password_str] if password_str else build_password_candidates(getattr(uploaded, "name", ""))

    uploaded.seek(0)
    raw = uploaded.read()
    file_info = inspect_uploaded_excel(getattr(uploaded, "name", ""), raw)
    precheck_error = build_upload_precheck_error(file_info)
    if precheck_error:
        raise ValueError(precheck_error)

    # Try normal open first for standard .xlsx/.xlsm files.
    direct_open_error = None
    try:
        return process_and_input_data_in_existing_sheet(BytesIO(raw))
    except Exception as error:
        direct_open_error = error

    if file_info["container"] == "ooxml_zip":
        raise ValueError(
            "The uploaded workbook could not be opened as a normal .xlsx/.xlsm file. "
            f"Details: {format_workbook_open_error(direct_open_error)}"
        )

    # If encrypted, try the password inferred from the file name first, then fall back.
    last_error = None
    for candidate_password in password_candidates:
        try:
            decrypted = BytesIO()
            office_file = msoffcrypto.OfficeFile(BytesIO(raw))
            office_file.load_key(password=candidate_password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            return process_and_input_data_in_existing_sheet(decrypted)
        except Exception as decrypt_error:
            last_error = decrypt_error

    if last_error is not None:
        if file_info["container"] == "ole_compound":
            raise ValueError(
                "The uploaded workbook looks password-protected or legacy Excel format. "
                "If it is password-protected, make sure the file name month/year matches the expected password "
                '(for example, "MAD_Apr2026"). If it is an old .xls workbook, save it first as .xlsx or .xlsm. '
                f"Details: {format_workbook_open_error(last_error)}"
            )

        if direct_open_error is not None:
            raise ValueError(
                "The uploaded workbook could not be opened. "
                f"Direct open error: {format_workbook_open_error(direct_open_error)}. "
                f"Password open error: {format_workbook_open_error(last_error)}."
            )

        raise ValueError(format_workbook_open_error(last_error))

    raise ValueError("Unable to determine the correct password for the uploaded workbook.")


# ==============================
# MAIN
# ==============================
def run():
    st.title("MAD PL PYTS AUTOMATION")

    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xlsm"])

    if uploaded_file:
        try:
            wb = load_workbook_with_optional_password(uploaded_file)

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            st.success("Done! Same LAN2 is placed in one row, amounts are summed by week, and the latest payment date per week is kept.")

            st.download_button(
                "Download Updated Excel",
                data=out,
                file_name="MADRID PL PYTS(AUTOMATED).xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as err:
            st.error(f"Error: {err}")


if __name__ == "__main__":
    run()
