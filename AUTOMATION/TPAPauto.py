import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, date, timedelta
import msoffcrypto
from concurrent.futures import ThreadPoolExecutor
import re
from copy import copy

# =========================
# Streamlit Configuration
# =========================
st.set_page_config(page_title="Excel Automation")
st.title("Automate PL and CC Sheet Processing")

# =========================
# File Upload
# =========================
uploaded_main = st.file_uploader(
    "TPAP File",
    type=["xlsx", "xlsm"]
)

uploaded_sheet1 = st.file_uploader(
    "MADRID PL PYTS",
    type=["xlsx", "xlsm"]
)

# ✅ Different passwords per file
main_password = st.text_input(
    "TPAP File",
    value="MAD_1Q2026",
    type="password"
)

sheet1_password = st.text_input(
    "MADRID PL PYTS",
    value="MAD_Mar2026",
    type="password"
)

# =========================
# Settings
# =========================
TARGET_SHEET_PL = "PL"
TARGET_SHEET_CC = "CC"
TARGET_SHEET_SHEET1 = "Sheet1"
TARGET_SHEET_PAYMENT_AS_OF = "Payment as of"  # can be "payment as of FEB 2026"

# CC columns
COL_PTP_DUE = "P"         # PTP date / Due Date
COL_DATE_UPDATED = "G"    # Date Updated
COL_STAT = "H"            # Status
COL_PAYMENT = "J"         # Payment
COL_FACE = "R"            # Face Amount
DATA_START_ROW = 2

STATUS_DEFUALTED = "DEFUALTED"
STATUS_AVAILED = "AVAILED"
STATUS_COMPLIED = "COMPLIED"

# PL key
PL_KEY_COL = "A"

# PL write area: I = label, J = value
PL_LABEL_COL = 9   # I
PL_VALUE_COL = 10  # J

# =========================
# (Optional) Holidays
# =========================
HOLIDAYS = [
    date(2026, 1, 1),  # New Year's Day
    # add more if needed
]

# =========================
# Helpers
# =========================
def normalize_status(val) -> str:
    return str(val).strip() if val is not None else ""

def normalize_key(val):
    """Normalize keys so 123, 123.0, '123' match consistently."""
    if val is None:
        return None
    if isinstance(val, bool):
        return str(val).strip()
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val) if val.is_integer() else val
    if isinstance(val, str):
        s = val.strip()
        if s.isdigit():
            try:
                return int(s)
            except Exception:
                return s
        return s
    return str(val).strip()

def to_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace("₱", "").replace(",", "").replace(" ", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None

def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        # Added a safer mix of formats
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None

def is_past_due(due_dt, today_dt):
    return due_dt is not None and due_dt < today_dt

def apply_status(ws, row, new_status, today_date):
    ws[f"{COL_STAT}{row}"].value = new_status
    cell = ws[f"{COL_DATE_UPDATED}{row}"]
    cell.value = today_date
    cell.number_format = "mm/dd/yyyy"

def decrypt_bytes_if_needed(file_bytes: bytes, password_text: str, label: str) -> BytesIO:
    raw = BytesIO(file_bytes)
    raw.seek(0)

    if not password_text:
        raw.seek(0)
        return raw

    try:
        raw.seek(0)
        office = msoffcrypto.OfficeFile(raw)
        office.load_key(password=password_text)
        decrypted = BytesIO()
        office.decrypt(decrypted)
        decrypted.seek(0)
        return decrypted
    except Exception as e:
        raise Exception(
            f"Decryption failed for {label}. Check password (or leave blank if not protected). "
            f"Details: {e}"
        )

def load_wb_from_bytes(file_bytes: bytes, filename: str, password_text: str, label: str):
    keep_vba = filename.lower().endswith(".xlsm")
    stream = decrypt_bytes_if_needed(file_bytes, password_text, label)
    stream.seek(0)
    return load_workbook(stream, data_only=False, keep_vba=keep_vba)

def find_payment_as_of_sheet(wb, prefix: str):
    """Accepts exact 'Payment as of' or any sheet that starts with it."""
    if prefix in wb.sheetnames:
        return wb[prefix]

    p = str(prefix).strip().lower()
    for name in wb.sheetnames:
        
        if str(name).strip().lower().startswith(p):
            return wb[name]
    return None

def find_first_date_in_sheet(ws, scan_rows=30, scan_cols=15):
    """Finds the first cell that looks like a date."""
    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            v = ws.cell(row=r, column=c).value
            d = to_date(v)
            if d is not None:
                return d
    return None

def get_weekday_range_from_week_start(week_start: date):
    """
    Merge behavior:
    - Snap to Monday
    - Default end is Friday
    - If Monday is holiday/weekend, move start forward
    - If Friday is holiday/weekend, move end backward
    Output: ("Feb 02", "06")
    """
    monday = week_start - timedelta(days=week_start.weekday())
    friday = monday + timedelta(days=4)

    # move monday forward if it falls on weekend/holiday
    while monday.weekday() >= 5 or monday in HOLIDAYS:
        monday += timedelta(days=1)

    # move friday backward if it falls on weekend/holiday
    while friday.weekday() >= 5 or friday in HOLIDAYS:
        friday -= timedelta(days=1)

    start_txt = monday.strftime("%b %d")
    end_txt = friday.strftime("%d")
    return start_txt, end_txt

def build_sheet1_weeks_map(ws_sheet1):
    """
    key = LAN2 (col A)
    value = list of (week_index, week_value, week_date)

    Supports headers:
      "WEEK 1", "WEEK1", "WEEK 01", ...
      "WEEK 1 DATE", "WEEK1 DATE", ...
    Fallback:
      B=value, C=date, D=value, E=date, ...
    """
    headers = []
    for c in range(1, ws_sheet1.max_column + 1):
        hv = ws_sheet1.cell(row=1, column=c).value
        headers.append("" if hv is None else str(hv).strip())

    week_value_cols = {}
    week_date_cols = {}

    for idx, h in enumerate(headers, start=1):
        m_val = re.fullmatch(r"WEEK\s*0*(\d+)", h, flags=re.IGNORECASE)
        m_date = re.fullmatch(r"WEEK\s*0*(\d+)\s*DATE", h, flags=re.IGNORECASE)
        if m_val:
            week_value_cols[int(m_val.group(1))] = idx
        if m_date:
            week_date_cols[int(m_date.group(1))] = idx

    # Fallback if no headers found
    if not week_value_cols and not week_date_cols:
        w = 1
        col = 2
        max_col = ws_sheet1.max_column
        while col <= max_col:
            week_value_cols[w] = col
            if col + 1 <= max_col:
                week_date_cols[w] = col + 1
            col += 2
            w += 1

    max_week = max(set(week_value_cols.keys()) | set(week_date_cols.keys())) if (week_value_cols or week_date_cols) else 0

    mapping = {}
    for r in range(DATA_START_ROW, ws_sheet1.max_row + 1):
        lan2 = normalize_key(ws_sheet1.cell(row=r, column=1).value)
        if lan2 is None:
            continue

        weeks = []
        for w in range(1, max_week + 1):
            vc = week_value_cols.get(w)
            dc = week_date_cols.get(w)
            week_val = ws_sheet1.cell(row=r, column=vc).value if vc else None
            week_date = ws_sheet1.cell(row=r, column=dc).value if dc else None

            num = to_number(week_val)
            if week_val is None or (num is not None and num == 0):
                continue

            d = to_date(week_date)
            if d is None:
                continue

            weeks.append((w, week_val, d))

        if weeks:
            weeks.sort(key=lambda x: x[0])
            mapping[lan2] = weeks

    return mapping

# =========================
# CC Processing (Complete)
# =========================
def process_cc_sheet(ws_cc):
    today_dt = datetime.today().date()

    for row in range(DATA_START_ROW, ws_cc.max_row + 1):
        status = normalize_status(ws_cc[f"{COL_STAT}{row}"].value).upper()

        # Skip protected statuses
        if status in ("REFUSED/NOT AVAILING", "UNDER NEGO"):
            continue

        due_dt = to_date(ws_cc[f"{COL_PTP_DUE}{row}"].value)

        if is_past_due(due_dt, today_dt):
            apply_status(ws_cc, row, STATUS_DEFUALTED, today_dt)
            continue

        payment_val = to_number(ws_cc[f"{COL_PAYMENT}{row}"].value)
        if payment_val is None or payment_val == 0:
            apply_status(ws_cc, row, STATUS_AVAILED, today_dt)
            continue

        face_val = to_number(ws_cc[f"{COL_FACE}{row}"].value)
        if face_val is not None and payment_val >= face_val:
            apply_status(ws_cc, row, STATUS_COMPLIED, today_dt)
        else:
            apply_status(ws_cc, row, STATUS_AVAILED, today_dt)

# =========================
# PL Processing (Improved)
# =========================
def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    """Copy styling so inserted rows look like the base row."""
    for c in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        d._style = copy(s._style)
        d.font = copy(s.font)
        d.fill = copy(s.fill)
        d.border = copy(s.border)
        d.alignment = copy(s.alignment)
        d.number_format = s.number_format
        d.protection = copy(s.protection)

def get_customer_block_end(ws_pl, base_row: int, cust_key):
    """
    Find the end row of the current customer block:
    From base_row+1 downward while column A equals cust_key (inserted rows copy A,B,C).
    Stops when key changes to a different non-empty key.
    """
    r = base_row + 1
    last = ws_pl.max_row
    end = base_row

    while r <= last:
        k = normalize_key(ws_pl[f"{PL_KEY_COL}{r}"].value)
        if k is None:
            break
        if k != cust_key:
            break
        end = r
        r += 1

    return end

def find_label_row(ws_pl, start_row: int, end_row: int, contains_text: str):
    """
    Find a row within [start_row..end_row] where column I contains contains_text.
    Returns row index or None.
    """
    needle = contains_text.strip().lower()
    for r in range(start_row, end_row + 1):
        v = ws_pl.cell(row=r, column=PL_LABEL_COL).value
        if v is None:
            continue
        if needle in str(v).strip().lower():
            return r
    return None

def update_pl_from_sheet1(ws_pl, sheet1_weeks_map, insert_new_rows: bool, payment_as_of_date: date):
    """
    For each PL row:
      if PL!A matches Sheet1 LAN2:
        - Insert 2 rows per week if missing (Payment Week + Payment Date)
        - If exists already, UPDATE value/date/label range
    """
    if not insert_new_rows:
        st.warning("Insert mode is OFF. Turn it ON if you want row insertion per week.")
        return

    # collect matching customer rows
    matches = []
    for r in range(DATA_START_ROW, ws_pl.max_row + 1):
        cust = normalize_key(ws_pl[f"{PL_KEY_COL}{r}"].value)
        if cust is not None and cust in sheet1_weeks_map:
            matches.append(r)

    if not matches:
        st.info("No matching customers found in PL based on Sheet1 LAN2.")
        return

    max_col = ws_pl.max_column

    # bottom-up insertion keeps indices stable
    for base_row in reversed(matches):
        cust = normalize_key(ws_pl[f"{PL_KEY_COL}{base_row}"].value)
        weeks = sheet1_weeks_map.get(cust, [])
        if not weeks:
            continue

        # compute current block end (in case file already has inserted rows)
        block_end = get_customer_block_end(ws_pl, base_row, cust)
        insert_at = base_row + 1  # we will insert from top of block downward

        for (week_idx, week_val, week_date) in weeks:
            start_txt, end_txt = get_weekday_range_from_week_start(week_date)

            week_label_key = f"Payment Week {week_idx}"
            date_label_key = f"Payment Date (Week {week_idx})"

            # refresh block_end each loop because rows can be inserted
            block_end = get_customer_block_end(ws_pl, base_row, cust)

            existing_week_row = find_label_row(ws_pl, base_row + 1, block_end, week_label_key)
            existing_date_row = find_label_row(ws_pl, base_row + 1, block_end, date_label_key)

            # If both exist -> UPDATE values + normalize label text
            if existing_week_row and existing_date_row:
                ws_pl.cell(row=existing_week_row, column=PL_LABEL_COL).value = (
                    f"Payment Week {week_idx} ({start_txt} to {end_txt})"
                )
                ws_pl.cell(row=existing_week_row, column=PL_VALUE_COL).value = week_val

                dcell = ws_pl.cell(row=existing_date_row, column=PL_VALUE_COL)
                dcell.value = week_date
                dcell.number_format = "mm/dd/yyyy"
                continue

            # If one exists, insert only the missing one (keeps workbook clean)
            if existing_week_row and not existing_date_row:
                # insert date row below week row
                ws_pl.insert_rows(existing_week_row + 1, amount=1)
                copy_row_style(ws_pl, base_row, existing_week_row + 1, max_col)

                for col_idx in (1, 2, 3):
                    ws_pl.cell(row=existing_week_row + 1, column=col_idx).value = ws_pl.cell(row=base_row, column=col_idx).value

                # Update Payment Date label and value
                ws_pl.cell(row=existing_week_row + 1, column=PL_LABEL_COL).value = f"Payment Date (Week {week_idx})"
                dcell = ws_pl.cell(row=existing_week_row + 1, column=PL_VALUE_COL)
                dcell.value = week_date
                dcell.number_format = "mm/dd/yyyy"
                continue

            if existing_date_row and not existing_week_row:
                # insert week row above date row
                ws_pl.insert_rows(existing_date_row, amount=1)
                copy_row_style(ws_pl, base_row, existing_date_row, max_col)

                for col_idx in (1, 2, 3):
                    ws_pl.cell(row=existing_date_row, column=col_idx).value = ws_pl.cell(row=base_row, column=col_idx).value

                # Update Payment Week label and value
                ws_pl.cell(row=existing_date_row, column=PL_LABEL_COL).value = f"Payment Week {week_idx} ({start_txt} to {end_txt})"
                ws_pl.cell(row=existing_date_row, column=PL_VALUE_COL).value = week_val
                continue

            # Neither exists -> insert 2 rows at insert_at
            ws_pl.insert_rows(insert_at, amount=2)
            copy_row_style(ws_pl, base_row, insert_at, max_col)
            copy_row_style(ws_pl, base_row, insert_at + 1, max_col)

            # Copy identifiers (A,B,C)
            for col_idx in (1, 2, 3):
                ws_pl.cell(row=insert_at, column=col_idx).value = ws_pl.cell(row=base_row, column=col_idx).value
                ws_pl.cell(row=insert_at + 1, column=col_idx).value = ws_pl.cell(row=base_row, column=col_idx).value

            # Week row
            ws_pl.cell(row=insert_at, column=PL_LABEL_COL).value = f"Payment Week {week_idx} ({start_txt} to {end_txt})"
            ws_pl.cell(row=insert_at, column=PL_VALUE_COL).value = week_val
                           # Date row
            ws_pl.cell(row=insert_at + 1, column=PL_LABEL_COL).value = f"Payment Date (Week {week_idx})"
            dcell = ws_pl.cell(row=insert_at + 1, column=PL_VALUE_COL)
            dcell.value = week_date
            dcell.number_format = "mm/dd/yyyy"

            insert_at += 2  # Move the insertion point forward by 2 rows (one for week, one for date)

# =========================
# UI Controls
# =========================
# Removed the UI controls section with the checkboxes.

# =========================
# Run
# =========================
if uploaded_main and uploaded_sheet1:
    try:
        main_bytes = uploaded_main.getvalue()
        sheet1_bytes = uploaded_sheet1.getvalue()

        with ThreadPoolExecutor(max_workers=2) as executor:
            f_main = executor.submit(load_wb_from_bytes, main_bytes, uploaded_main.name, main_password, "MAIN file")
            f_sheet = executor.submit(load_wb_from_bytes, sheet1_bytes, uploaded_sheet1.name, sheet1_password, "SHEET1 file")
            wb_main = f_main.result()
            wb_sheet = f_sheet.result()

        # Validate sheets
        if TARGET_SHEET_PL not in wb_main.sheetnames:
            st.error("PL sheet not found in the MAIN file.")
            st.stop()

        if TARGET_SHEET_SHEET1 not in wb_sheet.sheetnames:
            st.error("Sheet1 sheet not found in the SHEET1 file.")
            st.stop()

        ws_payment_as_of = find_payment_as_of_sheet(wb_sheet, TARGET_SHEET_PAYMENT_AS_OF)
        if ws_payment_as_of is None:
            st.error(
                f"Payment as of sheet not found in the SHEET1 file. "
                f"Expected '{TARGET_SHEET_PAYMENT_AS_OF}' or sheet starting with it "
                f"(e.g., 'payment as of FEB 2026')."
            )
            st.stop()

        ws_pl = wb_main[TARGET_SHEET_PL]
        ws_sheet1 = wb_sheet[TARGET_SHEET_SHEET1]

        payment_as_of_date = find_first_date_in_sheet(ws_payment_as_of)
        if payment_as_of_date is None:
            st.error("Could not find a valid date in the 'Payment as of' sheet.")
            st.stop()

        st.write(f"Payment as of sheet detected: **{ws_payment_as_of.title}**")
        st.write(f"Payment as of date detected: **{payment_as_of_date.strftime('%d/%m/%Y')}**")

        # Build mapping (LAN2 -> weeks)
        sheet1_weeks_map = build_sheet1_weeks_map(ws_sheet1)
        if not sheet1_weeks_map:
            st.warning("No WEEK data detected in Sheet1 (mapping is empty). Nothing will be inserted/updated in PL.")

        # Update PL
        update_pl_from_sheet1(
            ws_pl=ws_pl,
            sheet1_weeks_map=sheet1_weeks_map,
            insert_new_rows=True,  # Default value set to True
            payment_as_of_date=payment_as_of_date
        )

        # Process CC if exists (always processed, no checkbox anymore)
        if TARGET_SHEET_CC in wb_main.sheetnames:
            process_cc_sheet(wb_main[TARGET_SHEET_CC])
        else:
            st.info("CC sheet not found in MAIN file; skipping CC processing.")

        # Save output
        out = BytesIO()
        wb_main.save(out)
        out.seek(0)

        is_xlsm = uploaded_main.name.lower().endswith(".xlsm")
        out_name = "updated_template.xlsm" if is_xlsm else "updated_template.xlsx"
        mime = "application/vnd.ms-excel.sheet.macroEnabled.12" if is_xlsm else \
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        st.success("Processing complete!")
        st.download_button(
            label="Download Updated Excel",
            data=out,
            file_name=out_name,
            mime=mime
        )

    except Exception as e:
        st.error(f"Processing failed: {e}")
else:
    st.info("Please upload BOTH files (Main workbook + Sheet1 workbook).")