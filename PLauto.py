import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, date, timedelta
import msoffcrypto

st.set_page_config(page_title="Excel Automation")
st.title("Automate CC and PL Sheet Processing")

# File upload
uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls", "xlsm"])
password = st.text_input("Excel password (leave blank if none)", type="password")

# Column settings
TARGET_SHEET_CC = "CC"
TARGET_SHEET_PL = "PL"
COL_PTP_DUE = "P"
COL_DATE_UPDATED = "G"
COL_STAT = "H"
COL_PAYMENT = "J"
COL_FACE = "R"
DATA_START_ROW = 2

STATUS_DEFUALTED = "DEFUALTED"
STATUS_AVAILED = "AVAILED"
PROTECTED_STATUSES = {"UNDER NEGO", "REFUSED/NOT AVAILING", "COMPLIED"}
DATE_FORMAT = "mm/dd/yyyy"

TPAP_PASSWORD = "MAD_1Q2026"

# Normalize status values
def normalize_status(val) -> str:
    return str(val).strip() if val is not None else ""

# Convert values to numbers
def to_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace("₱", "").replace(",", "").replace(" ", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None

# Convert payment values to numbers
def to_payment_number(val):
    num = to_number(val)
    return num if num is not None and num != 0 else None

# Convert string dates into date objects
def to_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m-%d/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None

# Check if the date is past due
def is_past_due(due_dt: date, today_dt: date) -> bool:
    return due_dt is not None and due_dt < today_dt

# Apply status and update date
def apply_status(ws, row: int, new_status: str, today_date: date):
    ws[f"{COL_STAT}{row}"] = new_status
    cell = ws[f"{COL_DATE_UPDATED}{row}"]
    cell.value = today_date
    cell.number_format = DATE_FORMAT

# Process CC sheet
def process_workbook_cc(file_stream):
    wb = load_workbook(file_stream)
    if TARGET_SHEET_CC not in wb.sheetnames:
        raise Exception("CC sheet not found")
    ws = wb[TARGET_SHEET_CC]
    ws.column_dimensions[COL_DATE_UPDATED].width = 18
    for r in range(DATA_START_ROW, ws.max_row + 1):
        ws[f"{COL_DATE_UPDATED}{r}"].number_format = DATE_FORMAT
    today_dt = datetime.today().date()
    for row in range(DATA_START_ROW, ws.max_row + 1):
        status = normalize_status(ws[f"{COL_STAT}{row}"].value).upper()
        if status in PROTECTED_STATUSES:
            continue
        due_dt = to_date(ws[f"{COL_PTP_DUE}{row}"].value)
        if is_past_due(due_dt, today_dt):
            apply_status(ws, row, STATUS_DEFUALTED, today_dt)
            continue
        payment_val = to_payment_number(ws[f"{COL_PAYMENT}{row}"].value)
        if payment_val is None:
            apply_status(ws, row, STATUS_AVAILED, today_dt)
            continue
        face_val = to_number(ws[f"{COL_FACE}{row}"].value)
        if face_val is not None and payment_val >= face_val:
            apply_status(ws, row, "COMPLIED", today_dt)
        else:
            apply_status(ws, row, STATUS_AVAILED, today_dt)
    return wb

# Process PL sheet (optional logic can be added here)
def process_pl_sheet(file_stream):
    wb = load_workbook(file_stream)
    if TARGET_SHEET_PL not in wb.sheetnames:
        raise Exception("PL sheet not found")
    ws = wb[TARGET_SHEET_PL]
    today_dt = datetime.today().date()
    for row in range(DATA_START_ROW, ws.max_row + 1):
        status = normalize_status(ws[f"{COL_STAT}{row}"].value).upper()
        if status in PROTECTED_STATUSES:
            continue
        due_dt = to_date(ws[f"{COL_PTP_DUE}{row}"].value)
        if is_past_due(due_dt, today_dt):
            apply_status(ws, row, STATUS_DEFUALTED, today_dt)
            continue
        payment_val = to_payment_number(ws[f"{COL_PAYMENT}{row}"].value)
        if payment_val is None:
            apply_status(ws, row, STATUS_AVAILED, today_dt)
            continue
        face_val = to_number(ws[f"{COL_FACE}{row}"].value)
        if face_val is not None and payment_val >= face_val:
            apply_status(ws, row, "COMPLIED", today_dt)
        else:
            apply_status(ws, row, STATUS_AVAILED, today_dt)
    return wb

# Select task (CC or PL sheet)
task = st.radio("Select Automation Task", ["Process 'CC' Sheet", "Process 'PL' Sheet"])

if uploaded_file:
    uploaded_file.seek(0)
    if task == "Process 'CC' Sheet":
        try:
            wb = process_workbook_cc(BytesIO(uploaded_file.read()))
        except Exception as e:
            if password:
                try:
                    uploaded_file.seek(0)
                    decrypted = BytesIO()
                    office_file = msoffcrypto.OfficeFile(uploaded_file)
                    office_file.load_key(password=password)
                    office_file.decrypt(decrypted)
                    decrypted.seek(0)
                    wb = process_workbook_cc(decrypted)
                except Exception as e2:
                    st.error(f"Failed to open password-protected file: {e2}")
                    st.stop()
            else:
                st.error(f"Failed to process file: {e}")
                st.stop()
        st.success("CC sheet updated successfully!")
    elif task == "Process 'PL' Sheet":
        try:
            wb = process_pl_sheet(BytesIO(uploaded_file.read()))
        except Exception as e:
            if password:
                try:
                    uploaded_file.seek(0)
                    decrypted = BytesIO()
                    office_file = msoffcrypto.OfficeFile(uploaded_file)
                    office_file.load_key(password=password)
                    office_file.decrypt(decrypted)
                    decrypted.seek(0)
                    wb = process_pl_sheet(decrypted)
                except Exception as e2:
                    st.error(f"Failed to open password-protected file: {e2}")
                    st.stop()
            else:
                st.error(f"Failed to process file: {e}")
                st.stop()
        st.success("PL sheet processed successfully!")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    st.download_button(
        label="Download Updated Excel",
        data=output,
        file_name="updated_template_with_cc_and_pl_sheets.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )